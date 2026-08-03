#!/usr/bin/env python3
"""Verify Paycheck-Budget-System.xlsx — structure + real formula recalc.

Checks the P1 review fixes stay fixed:
  1. bottom-line labels merged B:D (nothing clips), values in E with a
     $0.00-at-zero number format (the promised "$0" moment, not a dash)
  2. START HERE answers the second-paycheck question
  3. Debt Snowball has 10 debt rows with blank-safe formulas
  4. Year Dashboard has a native combo chart wired to the live ranges
  5. every key formula recalculates to the expected worked-example value
     (ground truth: LibreOffice headless recalc, not our own evaluator)
"""
import os, subprocess, sys, tempfile, zipfile
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Paycheck-Budget-System.xlsx")
MONEY0 = "$#,##0.00;($#,##0.00)"
failures = []

def check(cond, msg):
    print(("  ok  " if cond else "  FAIL") + " " + msg)
    if not cond:
        failures.append(msg)

wb = load_workbook(XLSX)

print("[1] clipped-label + $0-format fixes")
pb = wb["Paycheck Budget"]
merges = {str(m) for m in pb.merged_cells.ranges}
check("B33:D33" in merges and "B34:D34" in merges, "bottom-line labels merged across B:D")
check(pb["B33"].value == "Left to Assign (planned) — get this to $0", "hero label text intact")
check(pb["E33"].number_format == MONEY0 and pb["E34"].number_format == MONEY0,
      "hero cells render $0.00 at zero (never '-')")
check(isinstance(pb["E33"].value, str) and pb["E33"].value.startswith("="), "Left to Assign formula lives in E33")

print("[2] second-paycheck guidance")
sh = wb["START HERE"]
texts = " | ".join(str(c.value) for row in sh.iter_rows() for c in row if c.value)
check("TWO OR MORE CHECKS A MONTH" in texts, "START HERE has a second-check section")
check("Duplicate" in texts and "Move or Copy" in texts, "covers both Sheets and Excel copy flows")
check("One tab = one paycheck" in str(pb["B6"].value), "Paycheck Budget tab carries the pointer note")

print("[3] Debt Snowball capacity")
ds = wb["Debt Snowball"]
check(ds["B17"].border.left.style == "thin" and ds["C17"].fill.fill_type == "solid",
      "10th debt row (17) exists and is input-styled")
check(ds["B18"].value == "Total" and ds["C18"].value == "=SUM(C8:C17)", "total row sums all 10 rows")
check('IF(OR($C12="",$C12=0),""' in str(ds["G12"].value), "blank rows stay blank (no stray 'raise payment')")

print("[4] Year Dashboard chart")
with zipfile.ZipFile(XLSX) as z:
    charts = [n for n in z.namelist() if n.startswith("xl/charts/chart")]
    chart_xml = z.read(charts[0]).decode() if charts else ""
check(len(charts) == 1, "exactly one chart part in the workbook")
check("'Year Dashboard'!$C$7:$C$18" in chart_xml and "'Year Dashboard'!$D$7:$D$18" in chart_xml,
      "Income and Spent columns wired to live ranges")
check("'Year Dashboard'!$F$7:$F$18" in chart_xml and "<lineChart>" in chart_xml,
      "savings-rate line series wired to live range")
check("'Year Dashboard'!$B$7:$B$18" in chart_xml, "categories are the month names")
check(all(f'val="{c}"' in chart_xml for c in ("1F3A5F", "F2C14E", "2E7D6B")),
      "house palette (navy/gold/teal) applied")

print("[5] formula recalc (LibreOffice ground truth)")
with tempfile.TemporaryDirectory() as td:
    subprocess.run(["soffice", "--headless", "--convert-to", "xlsx", "--outdir", td, XLSX],
                   check=True, capture_output=True, timeout=180)
    calc = load_workbook(os.path.join(td, os.path.basename(XLSX)), data_only=True)
pbc, dsc, ydc = calc["Paycheck Budget"], calc["Debt Snowball"], calc["Year Dashboard"]
def near(a, b, eps=0.01):
    return a is not None and isinstance(a, (int, float)) and abs(a - b) < eps
check(near(pbc["C15"].value, 1590) and near(pbc["D15"].value, 1597.40), "bills subtotal 1590 / 1597.40")
check(near(pbc["C24"].value, 610) and near(pbc["D24"].value, 545.67), "spending subtotal 610 / 545.67")
check(near(pbc["C30"].value, 200) and near(pbc["D30"].value, 200), "savings subtotal 200 / 200")
check(near(pbc["E33"].value, 0), "Left to Assign = $0 on the worked example")
check(near(pbc["E34"].value, 56.93), "Actually remaining = $56.93")
check(dsc["G8"].value == 6 and dsc["G9"].value == 75 and dsc["G10"].value == 34,
      "payoff months 6 / 75 / 34 (listing QA pins)")
check(dsc["F11"].value in ("", None) and dsc["G11"].value in ("", None), "blank debt row stays blank")
check(near(dsc["C18"].value, 12890) and near(dsc["F18"].value, 520), "snowball totals 12,890 / 520")
check(near(ydc["F14"].value, 400 / 4800, 1e-6), "Aug savings rate 8.3%")
check(near(ydc["C19"].value, 4800) and near(ydc["E19"].value, 400), "year totals")

if failures:
    print(f"\nFAILED: {len(failures)} check(s)")
    sys.exit(1)
print("\nALL CHECKS PASSED")
