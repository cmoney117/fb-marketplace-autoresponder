#!/usr/bin/env python3
"""Build The Paycheck Budget spreadsheet product (starter digital listing).
NOTE: never use the phrase "Budget by Paycheck" anywhere — registered brand of The Budget Mom, LLC (see listing.md IP screen).

2026-08-03 P1 review fixes baked in:
  - bottom-line labels merged across B:D so "get this to $0" is never clipped;
    the values live in column E (under Difference)
  - hero cells render $0.00 (not the accounting dash) — the promised "$0" moment
  - START HERE answers "what about my second check?" (duplicate-the-tab flow)
  - Debt Snowball supports 10 debts (3 worked examples + 7 ready rows)
  - Year Dashboard: native combo chart — Income vs Spent columns + savings-rate
    line — driven by the live B6:F18 ranges
No macros — formulas + native features only (Excel 2010+ / Google Sheets)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils import get_column_letter

OUT = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/paycheck-budget-v1"
os.makedirs(OUT, exist_ok=True)

NAVY = "1F3A5F"; TEAL = "2E7D6B"; GOLD = "F2C14E"; LIGHT = "F4F6F8"; YELLOW = "FFF2CC"
H1 = Font(name="Arial", size=18, bold=True, color="FFFFFF")
H2 = Font(name="Arial", size=12, bold=True, color="FFFFFF")
B = Font(name="Arial", size=11)
BB = Font(name="Arial", size=11, bold=True)
SMALL = Font(name="Arial", size=10, italic=True, color="666666")
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00;($#,##0.00);"-"'
MONEY0 = "$#,##0.00;($#,##0.00)"  # hero cells: 0 renders $0.00, never "-"
PCT = "0.0%"

wb = Workbook()

# ---------- Tab 1: START HERE ----------
ws = wb.active; ws.title = "START HERE"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDE", (34, 30, 30, 30)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:E2"); ws["B2"] = "THE PAYCHECK BUDGET — Complete System"
ws["B2"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
ws["B2"].fill = HEAD_FILL; ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 34
rows = [
    ("", ""),
    ("Welcome! This system budgets the way real life works: one paycheck at a time.", ""),
    ("", ""),
    ("HOW TO USE (3 steps)", "SEC"),
    ("1. Go to the 'Paycheck Budget' tab. Enter your paycheck amount in the yellow cell.", ""),
    ("2. Give every dollar a job: fill the yellow Planned cells until 'Left to Assign' reads $0.", ""),
    ("3. As you spend, fill in Actual. The sheet shows exactly where you stand — no math, ever.", ""),
    ("", ""),
    ("THE OTHER TABS", "SEC"),
    ("• Debt Snowball — list up to 10 debts smallest to largest; see the payoff month for each one.", ""),
    ("• Year Dashboard — log each month once; watch the chart fill in and your savings rate trend up.", ""),
    ("", ""),
    ("TWO OR MORE CHECKS A MONTH? (most people!)", "SEC"),
    ("One tab = one paycheck. When the next check lands, make a copy of the 'Paycheck Budget' tab:", ""),
    ("• Google Sheets: right-click the tab → Duplicate.   • Excel: right-click the tab → Move or Copy → tick 'Create a copy'.", ""),
    ("Rename each copy for its check (e.g. 'Aug 1 check', 'Aug 15 check') and budget that check to $0.", ""),
    ("At the end of the month, add your checks together and log one line on the Year Dashboard.", ""),
    ("", ""),
    ("COLOR LEGEND", "SEC"),
    ("Yellow cells = yours to edit.  Everything else calculates automatically — please don't type over it.", ""),
    ("The numbers already in the sheet are a realistic example. Replace them with your own.", ""),
    ("", ""),
    ("GOOGLE SHEETS USERS", "SEC"),
    ("Upload this file to Google Drive, then open it — it converts automatically and all formulas work.", ""),
    ("The Year Dashboard chart converts to a Sheets chart too — colors may shift slightly; the numbers are identical.", ""),
    ("", ""),
    ("Questions or a formula acting up? Message us on Etsy any time — we answer fast and we'll fix it.", ""),
]
r = 4
for text, kind in rows:
    cell = ws.cell(row=r, column=2, value=text)
    if kind == "SEC":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell.font = H2; cell.fill = SEC_FILL
        ws.row_dimensions[r].height = 20
    else:
        cell.font = B
    r += 1

# ---------- Tab 2: Paycheck Budget ----------
ws = wb.create_sheet("Paycheck Budget")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDE", (28, 14, 14, 14)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:E2"); ws["B2"] = "PAYCHECK BUDGET"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
ws["B4"] = "Paycheck date:"; ws["B4"].font = BB
ws["C4"] = "2026-08-15"; ws["C4"].fill = INPUT_FILL; ws["C4"].font = B; ws["C4"].border = BORDER
ws["B5"] = "Paycheck amount:"; ws["B5"].font = BB
ws["C5"] = 2400; ws["C5"].fill = INPUT_FILL; ws["C5"].number_format = MONEY; ws["C5"].border = BORDER
ws["B6"] = "One tab = one paycheck. Next check? Duplicate this tab — the how-to is on START HERE."
ws["B6"].font = SMALL

def section(ws, start, title, items):
    ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=5)
    t = ws.cell(row=start, column=2, value=title); t.font = H2; t.fill = SEC_FILL
    hdrs = ["Item", "Planned", "Actual", "Difference"]
    for i, h in enumerate(hdrs):
        c = ws.cell(row=start + 1, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
    r = start + 2
    for name, planned, actual in items:
        ws.cell(row=r, column=2, value=name).font = B
        p = ws.cell(row=r, column=3, value=planned); p.fill = INPUT_FILL
        a = ws.cell(row=r, column=4, value=actual); a.fill = INPUT_FILL
        d = ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
        for cc in (ws.cell(row=r, column=2), p, a, d):
            cc.border = BORDER
        for cc in (p, a, d): cc.number_format = MONEY
        r += 1
    tr = r
    ws.cell(row=tr, column=2, value="Subtotal").font = BB
    ws.cell(row=tr, column=3, value=f"=SUM(C{start+2}:C{r-1})").font = BB
    ws.cell(row=tr, column=4, value=f"=SUM(D{start+2}:D{r-1})").font = BB
    ws.cell(row=tr, column=5, value=f"=C{tr}-D{tr}").font = BB
    for col in range(2, 6):
        c = ws.cell(row=tr, column=col); c.fill = ALT_FILL; c.border = BORDER
        if col > 2: c.number_format = MONEY
    return tr

bills_total = section(ws, 7, "BILLS (fixed — due before your next check)", [
    ("Rent / mortgage share", 950, 950), ("Electric", 85, 92.40), ("Car payment", 310, 310),
    ("Car insurance", 120, 120), ("Phone", 65, 65), ("Internet", 60, 60)])
spend_start = bills_total + 2
spend_total = section(ws, spend_start, "SPENDING (flexible)", [
    ("Groceries", 300, 268.51), ("Gas", 120, 104.22), ("Eating out", 80, 96.75),
    ("Fun money", 60, 45.00), ("Household / misc", 50, 31.19)])
save_start = spend_total + 2
save_total = section(ws, save_start, "SAVINGS & EXTRA DEBT", [
    ("Emergency fund", 100, 100), ("Extra to smallest debt", 100, 100)])
sr = save_total + 2
ws.merge_cells(start_row=sr, start_column=2, end_row=sr, end_column=5)
ws.cell(row=sr, column=2, value="THE BOTTOM LINE").font = H2
ws.cell(row=sr, column=2).fill = HEAD_FILL
# labels merged across B:D so the full text always shows; values sit in E
# (under Difference) and render $0.00 — never a dash — when you hit zero.
labels = [
    ("Left to Assign (planned) — get this to $0", f"=C5-C{bills_total}-C{spend_total}-C{save_total}"),
    ("Actually remaining (real money right now)", f"=C5-D{bills_total}-D{spend_total}-D{save_total}"),
]
for i, (lab, formula) in enumerate(labels):
    rr = sr + 1 + i
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
    ws.cell(row=rr, column=2, value=lab).font = BB
    fc = ws.cell(row=rr, column=5, value=formula); fc.font = BB; fc.number_format = MONEY0; fc.border = BORDER
ws.cell(row=sr + 3, column=2, value="Tip: 'Left to Assign' at $0 means every dollar has a job before you spend it.").font = SMALL
ws.freeze_panes = "A7"

# ---------- Tab 3: Debt Snowball ----------
ws = wb.create_sheet("Debt Snowball")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDEFG", (24, 14, 10, 14, 14, 18)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:G2"); ws["B2"] = "DEBT SNOWBALL"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
ws["B4"] = "List debts SMALLEST balance first. Put your monthly extra payment in the yellow cell — it attacks the top debt."
ws["B4"].font = SMALL
ws["B5"] = "Monthly extra payment:"; ws["B5"].font = BB
ws["C5"] = 100; ws["C5"].fill = INPUT_FILL; ws["C5"].number_format = MONEY; ws["C5"].border = BORDER
hdrs = ["Debt", "Balance", "APR", "Min payment", "You pay", "Months to payoff"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=7, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
# 10 debt rows: 3 worked examples + 7 blank yellow rows ready for real debts
debts = [("Store card", 640, 0.279, 35), ("Credit card", 2850, 0.246, 75), ("Car loan", 9400, 0.069, 310)]
N_DEBT_ROWS = 10
for i in range(N_DEBT_ROWS):
    r = 8 + i
    name, bal, apr, minp = debts[i] if i < len(debts) else (None, None, None, None)
    ws.cell(row=r, column=2, value=name).font = B
    b = ws.cell(row=r, column=3, value=bal); b.fill = INPUT_FILL; b.number_format = MONEY
    a = ws.cell(row=r, column=4, value=apr); a.fill = INPUT_FILL; a.number_format = PCT
    m = ws.cell(row=r, column=5, value=minp); m.fill = INPUT_FILL; m.number_format = MONEY
    yp = ws.cell(row=r, column=6, value=f'=IF(OR($C{r}="",$C{r}=0),"",E{r}+IF(ROW()=8,$C$5,0))'); yp.number_format = MONEY
    mo = ws.cell(row=r, column=7,
                 value=f'=IF(OR($C{r}="",$C{r}=0),"",IFERROR(ROUNDUP(NPER(D{r}/12,-F{r},C{r}),0),"raise payment"))')
    for col in range(2, 8): ws.cell(row=r, column=col).border = BORDER
tr = 8 + N_DEBT_ROWS
ws.cell(row=tr, column=2, value="Total").font = BB
ws.cell(row=tr, column=3, value=f"=SUM(C8:C{tr-1})").font = BB
ws.cell(row=tr, column=6, value=f"=SUM(F8:F{tr-1})").font = BB
for col in (3, 6): ws.cell(row=tr, column=col).number_format = MONEY
for col in range(2, 8): ws.cell(row=tr, column=col).fill = ALT_FILL; ws.cell(row=tr, column=col).border = BORDER
ws.cell(row=tr + 2, column=2, value="When a debt hits $0, roll its whole 'You pay' amount onto the next debt down. That's the snowball.").font = SMALL
ws.cell(row=tr + 3, column=2, value='"raise payment" means the payment doesn\'t cover the interest — increase it.').font = SMALL

# ---------- Tab 4: Year Dashboard ----------
ws = wb.create_sheet("Year Dashboard")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDEF", (14, 14, 14, 14, 14)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:F2"); ws["B2"] = "YEAR DASHBOARD"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
ws["B4"] = "Once a month, copy your totals here (2 minutes). Watch the savings-rate column — that's the number that changes your life."
ws["B4"].font = SMALL
hdrs = ["Month", "Income", "Spent", "Saved", "Savings rate"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=6, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
example = {"Aug": (4800, 4230, 400)}
for i, m in enumerate(months):
    r = 7 + i
    ws.cell(row=r, column=2, value=m).font = B
    inc, sp, sv = example.get(m, (None, None, None))
    for j, v in enumerate((inc, sp, sv)):
        c = ws.cell(row=r, column=3 + j, value=v); c.fill = INPUT_FILL; c.number_format = MONEY; c.border = BORDER
    rate = ws.cell(row=r, column=6, value=f'=IF(OR(C{r}="",C{r}=0),"",E{r}/C{r})')
    rate.number_format = PCT; rate.border = BORDER
    ws.cell(row=r, column=2).border = BORDER
tr = 19
ws.cell(row=tr, column=2, value="Year total").font = BB
for j, col in enumerate("CDE"):
    ws.cell(row=tr, column=3 + j, value=f"=SUM({col}7:{col}18)").font = BB
    ws.cell(row=tr, column=3 + j).number_format = MONEY
ws.cell(row=tr, column=6, value=f'=IF(OR(C{tr}="",C{tr}=0),"",E{tr}/C{tr})').font = BB
ws.cell(row=tr, column=6).number_format = PCT
for col in range(2, 7): ws.cell(row=tr, column=col).fill = ALT_FILL; ws.cell(row=tr, column=col).border = BORDER
ws.cell(row=tr + 2, column=2, value="August row is a worked example — clear it and enter your own months.").font = SMALL

# --- native combo chart: Income vs Spent columns + savings-rate line ---------
# Driven by the live ranges (B6:F18), so the chart fills in as months are logged.
bar = BarChart()
bar.type = "col"
bar.grouping = "clustered"
bar.style = 10
bar.title = "Your year at a glance"
bar.add_data(Reference(ws, min_col=3, max_col=4, min_row=6, max_row=18), titles_from_data=True)  # Income, Spent
bar.set_categories(Reference(ws, min_col=2, min_row=7, max_row=18))  # month names
bar.series[0].graphicalProperties = GraphicalProperties(solidFill=NAVY)   # Income
bar.series[1].graphicalProperties = GraphicalProperties(solidFill=GOLD)   # Spent
bar.y_axis.numFmt = "$#,##0"
bar.y_axis.majorGridlines = None
bar.y_axis.scaling.min = 0  # bars must grow from $0, not float
line = LineChart()
line.add_data(Reference(ws, min_col=6, max_col=6, min_row=6, max_row=18), titles_from_data=True)  # Savings rate
lp = LineProperties(solidFill=TEAL, w=28000)
line.series[0].graphicalProperties = GraphicalProperties(ln=lp)
line.series[0].smooth = False
line.series[0].marker.symbol = "circle"  # visible even with a single month logged
line.series[0].marker.size = 7
line.series[0].marker.graphicalProperties = GraphicalProperties(solidFill=TEAL)
line.y_axis.axId = 200
line.y_axis.title = None
line.y_axis.numFmt = "0%"
line.y_axis.majorGridlines = None
line.y_axis.scaling.min = 0
line.y_axis.crosses = "max"  # secondary percent axis on the right
bar += line
bar.legend.position = "b"
bar.width = 17
bar.height = 10
ws.add_chart(bar, "H4")  # beside the table: table + chart = one dashboard view

# print setup: landscape, fit to one page wide — Ctrl+P gives whole tabs, not confetti
from openpyxl.worksheet.properties import PageSetupProperties
for sheet in wb.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

path = os.path.join(OUT, "Paycheck-Budget-System.xlsx")
wb.save(path)
print("saved", path)
