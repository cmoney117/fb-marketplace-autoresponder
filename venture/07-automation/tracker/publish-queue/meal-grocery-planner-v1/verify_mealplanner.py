#!/usr/bin/env python3
"""Verify Meal-Plan-Grocery-Budget-Planner.xlsx against independently recomputed values.

Imports ONLY the worked-example source data from build_mealplanner.py, then:
  1. recomputes every key total in plain Python (no spreadsheet math reused);
  2. asserts the exact formula string (correct ranges) of every formula cell —
     and that no unexpected formula cell exists anywhere in the workbook;
  3. evaluates the closed set of numeric formulas with a mini evaluator
     (SUM / SUMIF / COUNTIF / COUNTA / COUNT / MIN + arithmetic) and compares;
  4. if the file has been recalculated (--recalc, see below), compares every
     cached value — including IF/INDEX/MATCH price-book + dinner-panel cells;
  5. checks the traffic-light + cheapest-price conditional formatting ranges,
     the data-validation dropdowns (P1-3), the native bar chart with
     house-palette series colors (P1-2), sheet protection with unlocked yellow
     inputs (P2-6), print setup on every tab (P2-7), that no total/hero formula
     cell renders 0 as "-" (P1-4), that no input (yellow) cell holds a formula,
     no macros, no banned post-2007 functions, and the exact tab list.

--recalc: converts the workbook through headless LibreOffice so every formula
gets a cached value (nice first open everywhere), then re-injects the
openpyxl-built chart XML because LibreOffice drops the brand series colors.
Run: python3 build_mealplanner.py && python3 verify_mealplanner.py --recalc

Exits non-zero on any failure.
"""
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries, coordinate_to_tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_mealplanner import (  # worked-example data only
    XLSX, WEEK1_MEALS, MEAL_ROWS, WEEKLY_BUDGET, ROWS_PER_CAT, GROCERY,
    STORES, PRICE_ROWS, PB_FIRST, PB_LAST, MONTH_BUDGET, LAST_MONTH_TOTAL,
    WEEK_ACTUALS, PANTRY, PAN_FIRST, PAN_LAST, PANTRY_PLACES,
    DINNER_WEEK_CELL, DINNER_FIRST_ROW, dinner_formula,
)

TOL = 0.005
TABS = ["START HERE", "Weekly Meal Plan", "Grocery List Builder",
        "Price Book", "Monthly Grocery Budget", "Pantry Inventory"]
BANNED = re.compile(r"\b(XLOOKUP|XMATCH|SORT|FILTER|UNIQUE|SEQUENCE|TEXTJOIN|CONCAT|IFS|SWITCH|MAXIFS|MINIFS)\s*\(")
CHART_PART = "xl/charts/chart1.xml"


# ------------------------------------------------------------------- recalc
def recalc():
    """LibreOffice-recalc XLSX in place, preserving the openpyxl chart XML."""
    with zipfile.ZipFile(XLSX) as z:
        chart_xml = z.read(CHART_PART)
    tmp = tempfile.mkdtemp(prefix="mealplanner-recalc-")
    try:
        subprocess.run(
            ["soffice", "--headless", f"-env:UserInstallation=file://{tmp}/lo-profile",
             "--convert-to", "xlsx", "--outdir", tmp, XLSX],
            check=True, capture_output=True, timeout=300)
        conv = os.path.join(tmp, os.path.basename(XLSX))
        assert os.path.exists(conv), "LibreOffice produced no output"
        patched = conv + ".patched"
        with zipfile.ZipFile(conv) as zin, zipfile.ZipFile(patched, "w", zipfile.ZIP_DEFLATED) as zout:
            names = zin.namelist()
            assert CHART_PART in names, "chart part lost in LibreOffice round-trip"
            for name in names:
                zout.writestr(name, chart_xml if name == CHART_PART else zin.read(name))
        shutil.move(patched, XLSX)
        print("recalculated via LibreOffice (chart XML preserved)")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if "--recalc" in sys.argv:
    recalc()

# ---------------------------------------------------------------- expectations
# Recomputed straight from the source data — no spreadsheet ranges involved.
meal_count_w1 = sum(1 for m in MEAL_ROWS for v in WEEK1_MEALS[m] if v)
subtotals = [round(sum(p for _, p, _ in items), 2) for _, items in GROCERY]
list_total = round(sum(subtotals), 2)
left_vs_budget = round(WEEKLY_BUDGET - list_total, 2)
in_cart = round(sum(p for _, items in GROCERY for _, p, got in items if got == "Yes"), 2)
week_budget = MONTH_BUDGET / 4
week_diffs = [round(week_budget - a, 2) for a in WEEK_ACTUALS]
month_spent = round(sum(WEEK_ACTUALS), 2)
month_left = round(MONTH_BUDGET - month_spent, 2)
pct_used = month_spent / MONTH_BUDGET
vs_last = round(LAST_MONTH_TOTAL - month_spent, 2)
pantry_items = sum(1 for row in PANTRY if row[0])
pantry_flags = sum(1 for row in PANTRY if row[3] == "Yes")
dinners_w1 = list(WEEK1_MEALS["Dinner"])  # week picker ships set to 1
pb_expected = []  # (cheapest store name, cheapest price) per worked-example row
for item, size, *prices in PRICE_ROWS:
    mn = min(prices)
    assert prices.count(mn) == 1, f"price-book row '{item}' has a tied minimum — INDEX/MATCH would be ambiguous"
    pb_expected.append((STORES[prices.index(mn)], mn))

# hard pins — the worked example must read exactly like listing.md says
assert meal_count_w1 == 28
assert subtotals == [25.31, 28.30, 25.24, 30.19, 10.35, 15.47]
assert list_total == 134.86 and left_vs_budget == 15.14 and in_cart == 22.92
assert week_diffs == [11.58, -11.85, 0.00, 17.10]
assert month_spent == 583.17 and month_left == 16.83 and vs_last == 71.83
assert abs(pct_used - 0.97195) < 1e-9
assert pantry_items == 10 and pantry_flags == 3
assert {d > 0 for d in week_diffs} == {True, False} and any(d == 0 for d in week_diffs) and any(d < 0 for d in week_diffs), \
    "worked example must exercise all three traffic lights"

# ------------------------------------------------- expected formulas + values
MP, GL, PB, MB, PI = TABS[1], TABS[2], TABS[3], TABS[4], TABS[5]
FORMULAS = {}   # (sheet, coord) -> exact formula string
VALUES = {}     # (sheet, coord) -> expected value ("" for blank-row guards)
EVAL_SKIP = set()  # cells the mini evaluator can't do (IF/INDEX/MATCH) — cached pass covers them

for w in range(4):  # meal-plan counters: week block starts at row 6, 14, 22, 30
    s = 6 + w * 8
    FORMULAS[(MP, f"C{s + 6}")] = f"=COUNTA(C{s + 2}:I{s + 5})"
    VALUES[(MP, f"C{s + 6}")] = meal_count_w1 if w == 0 else 0

sub_rows = [11 + (i + 1) * (ROWS_PER_CAT + 2) - 1 for i in range(len(GROCERY))]  # 20,30,40,50,60,70
assert sub_rows == [20, 30, 40, 50, 60, 70]
for i, sr in enumerate(sub_rows):
    FORMULAS[(GL, f"C{sr}")] = f"=SUM(C{sr - ROWS_PER_CAT}:C{sr - 1})"
    VALUES[(GL, f"C{sr}")] = subtotals[i]
FORMULAS[(GL, "C6")] = "=" + "+".join(f"C{sr}" for sr in sub_rows)
VALUES[(GL, "C6")] = list_total
FORMULAS[(GL, "C7")] = "=C4-C6"; VALUES[(GL, "C7")] = left_vs_budget
FORMULAS[(GL, "C8")] = '=SUMIF(D12:D69,"Yes",C12:C69)'; VALUES[(GL, "C8")] = in_cart
for i in range(7):  # dinner panel H4..H10 pulls the picked week's dinners
    coord = f"H{DINNER_FIRST_ROW + i}"
    FORMULAS[(GL, coord)] = dinner_formula(i)
    VALUES[(GL, coord)] = dinners_w1[i]
    EVAL_SKIP.add((GL, coord))

for r in range(PB_FIRST, PB_LAST + 1):
    FORMULAS[(PB, f"G{r}")] = f'=IF(COUNT(D{r}:F{r})=0,"",INDEX($D$6:$F$6,MATCH(MIN(D{r}:F{r}),D{r}:F{r},0)))'
    FORMULAS[(PB, f"H{r}")] = f'=IF(COUNT(D{r}:F{r})=0,"",MIN(D{r}:F{r}))'
    EVAL_SKIP.add((PB, f"G{r}")); EVAL_SKIP.add((PB, f"H{r}"))
    if r - PB_FIRST < len(PRICE_ROWS):
        VALUES[(PB, f"G{r}")], VALUES[(PB, f"H{r}")] = pb_expected[r - PB_FIRST]
    else:
        VALUES[(PB, f"G{r}")] = VALUES[(PB, f"H{r}")] = ""

for i in range(4):
    r = 9 + i
    FORMULAS[(MB, f"C{r}")] = "=$C$4/4"; VALUES[(MB, f"C{r}")] = week_budget
    FORMULAS[(MB, f"E{r}")] = f"=C{r}-D{r}"; VALUES[(MB, f"E{r}")] = week_diffs[i]
FORMULAS[(MB, "C13")] = "=SUM(C9:C12)"; VALUES[(MB, "C13")] = MONTH_BUDGET
FORMULAS[(MB, "D13")] = "=SUM(D9:D12)"; VALUES[(MB, "D13")] = month_spent
FORMULAS[(MB, "E13")] = "=C13-D13"; VALUES[(MB, "E13")] = month_left
FORMULAS[(MB, "C16")] = "=D13"; VALUES[(MB, "C16")] = month_spent
FORMULAS[(MB, "C17")] = "=C4-D13"; VALUES[(MB, "C17")] = month_left
FORMULAS[(MB, "C18")] = '=IF(C4=0,"",D13/C4)'; VALUES[(MB, "C18")] = pct_used
EVAL_SKIP.add((MB, "C18"))
FORMULAS[(MB, "C19")] = "=C5-D13"; VALUES[(MB, "C19")] = vs_last

FORMULAS[(PI, "C4")] = f"=COUNTA(B{PAN_FIRST}:B{PAN_LAST})"; VALUES[(PI, "C4")] = pantry_items
FORMULAS[(PI, "C5")] = f'=COUNTIF(E{PAN_FIRST}:E{PAN_LAST},"Yes")'; VALUES[(PI, "C5")] = pantry_flags

# ---------------------------------------------------------------- mini evaluator
def _cells(ws, ref):
    if ":" in ref:
        return [c for row in ws[ref] for c in row]
    return [ws[ref]]

def _val(ws, cell):
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        return evaluate(ws, v)
    return v if isinstance(v, (int, float)) else 0

def _match(v, crit):
    return isinstance(v, str) and v.strip().lower() == crit.strip().lower()

def evaluate(ws, formula):
    """Closed set this product uses outside EVAL_SKIP: SUM, COUNTA, COUNTIF,
    SUMIF, COUNT, MIN + cell arithmetic (+ - * /)."""
    expr = formula.lstrip("=").replace("$", "")
    def fn(m):
        name, args = m.group(1).upper(), [a.strip() for a in m.group(2).split(",")]
        if name == "SUM":
            return repr(sum(_val(ws, c) for c in _cells(ws, args[0])))
        if name == "COUNTA":
            return repr(sum(1 for c in _cells(ws, args[0]) if c.value not in (None, "")))
        if name == "COUNT":
            return repr(sum(1 for c in _cells(ws, args[0]) if isinstance(c.value, (int, float))))
        if name == "MIN":
            return repr(min(_val(ws, c) for c in _cells(ws, args[0]) if isinstance(c.value, (int, float))))
        if name == "COUNTIF":
            return repr(sum(1 for c in _cells(ws, args[0]) if _match(c.value, args[1].strip('"'))))
        if name == "SUMIF":
            crit_cells, sum_cells = _cells(ws, args[0]), _cells(ws, args[2])
            return repr(sum(_val(ws, s) for c, s in zip(crit_cells, sum_cells) if _match(c.value, args[1].strip('"'))))
        raise ValueError("unsupported function " + name)
    expr = re.sub(r"([A-Z]+)\(([^()]*)\)", fn, expr)
    expr = re.sub(r"\b([A-Z]{1,2}[0-9]{1,4})\b", lambda m: repr(_val(ws, ws[m.group(1)])), expr)
    return eval(expr)

# ------------------------------------------------------------------ the checks
YELLOW = "FFF2CC"
wb = load_workbook(XLSX)                    # formula strings
wbv = load_workbook(XLSX, data_only=True)   # cached values (present after --recalc)

assert wb.sheetnames == TABS, f"tab list changed: {wb.sheetnames}"

# 1+2: every formula cell accounted for, exact string, not yellow-filled,
#      locked (P2-6), and never formatting 0 as "-" (P1-4)
def is_yellow(cell):
    rgb = getattr(cell.fill.fgColor, "rgb", None)
    return isinstance(rgb, str) and rgb.endswith(YELLOW)

seen = 0
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                key = (sheet, cell.coordinate)
                assert key in FORMULAS, f"unexpected formula at {key}: {cell.value}"
                assert cell.value == FORMULAS[key], \
                    f"{key}: formula is {cell.value!r}, expected {FORMULAS[key]!r}"
                assert not BANNED.search(cell.value), f"{key}: banned function in {cell.value!r}"
                assert not is_yellow(cell), f"{key}: formula cell has the edit-me yellow fill"
                assert cell.protection.locked in (None, True), f"{key}: formula cell is unlocked"
                assert '"-"' not in cell.number_format, \
                    f"{key}: calc cell would render 0 as a dash ({cell.number_format!r})"
                seen += 1
missing = set(FORMULAS) - {(s, c) for s in wb.sheetnames for c in
                           [cell.coordinate for row in wb[s].iter_rows() for cell in row
                            if isinstance(cell.value, str) and cell.value.startswith("=")]}
assert seen == len(FORMULAS) and not missing, f"missing formula cells: {sorted(missing)}"

# 3: mini-evaluate the numeric formulas against recomputed values
evaluated = 0
for key, want in VALUES.items():
    if key in EVAL_SKIP:
        continue
    sheet, coord = key
    got = evaluate(wb[sheet], wb[sheet][coord].value)
    assert abs(got - want) < TOL, f"{key}: evaluated {got}, expected {want}"
    evaluated += 1

# 4: cached values (LibreOffice recalc) — covers IF/INDEX/MATCH + dinner cells
cached = 0
have_cached = wbv[GL]["C6"].value is not None
if have_cached:
    for key, want in VALUES.items():
        sheet, coord = key
        got = wbv[sheet][coord].value
        if want == "":
            assert got in (None, ""), f"{key}: cached {got!r}, expected blank"
        elif isinstance(want, str):
            assert got == want, f"{key}: cached {got!r}, expected {want!r}"
        else:
            assert isinstance(got, (int, float)) and abs(got - want) < TOL, \
                f"{key}: cached {got!r}, expected {want}"
        cached += 1
    for sheet in wbv.sheetnames:
        for row in wbv[sheet].iter_rows():
            for cell in row:
                assert not (isinstance(cell.value, str) and cell.value.startswith("#")), \
                    f"error value {cell.value} at {sheet}!{cell.coordinate}"

# 5a: traffic-light conditional formatting on the over/under cells
def cf_map(ws):
    return {str(cf.sqref): list(cf.rules) for cf in ws.conditional_formatting}

def assert_traffic(ws, rng, where):
    rules = cf_map(ws).get(rng)
    assert rules, f"no conditional formatting on {where} {rng}: have {list(cf_map(ws))}"
    ops = {r.operator for r in rules if r.type == "cellIs"}
    assert {"greaterThan", "equal", "lessThan"} <= ops, f"{where} {rng}: traffic ops are {ops}"

assert_traffic(wb[GL], "C7", GL)
assert_traffic(wb[MB], "E9:E13", MB)
assert_traffic(wb[MB], "C17", MB)
assert_traffic(wb[MB], "C19", MB)

# 5b: cheapest-price highlight + use-first highlight
pb_rules = cf_map(wb[PB])
rng = f"D{PB_FIRST}:F{PB_LAST}"
assert rng in pb_rules, f"no cheapest-price rule on {rng}: have {list(pb_rules)}"
assert any(f"MIN($D{PB_FIRST}:$F{PB_FIRST})" in (r.formula[0] if r.formula else "")
           for r in pb_rules[rng]), "cheapest-price formula rule missing/wrong"
pi_rules = cf_map(wb[PI])
rng = f"E{PAN_FIRST}:E{PAN_LAST}"
assert rng in pi_rules, f"no use-first rule on {rng}: have {list(pi_rules)}"
assert any(r.type == "cellIs" and r.operator == "equal" and r.formula == ['"Yes"']
           for r in pi_rules[rng]), "use-first rule must be equal-\"Yes\""

# 5c: no macros in the package
with zipfile.ZipFile(XLSX) as z:
    names = z.namelist()
    assert not any("vbaProject" in n for n in names), "workbook contains a macro project"
    chart_parts = [n for n in names if re.match(r"xl/charts/chart\d+\.xml$", n)]
    chart_xml = z.read(CHART_PART).decode("utf-8", "ignore")

# 5d: data-validation dropdowns replace every "type Yes exactly" cell (P1-3)
def covered(ws, coord):
    """formula1 of the list validation covering coord, else None."""
    col, row = coordinate_to_tuple(coord)[1], coordinate_to_tuple(coord)[0]
    for dv in ws.data_validations.dataValidation:
        for rng in str(dv.sqref).split():
            c1, r1, c2, r2 = range_boundaries(rng)
            if c1 <= col <= c2 and r1 <= row <= r2:
                return dv
    return None

def assert_dropdown(sheet, coords, options, strict):
    want = '"' + ",".join(options) + '"'
    for coord in coords:
        dv = covered(wb[sheet], coord)
        assert dv is not None, f"{sheet}!{coord}: no dropdown covers it"
        assert dv.type == "list" and dv.formula1 == want, \
            f"{sheet}!{coord}: dropdown is {dv.formula1!r}, expected {want!r}"
        assert bool(dv.showErrorMessage) == strict, \
            f"{sheet}!{coord}: strictness {dv.showErrorMessage} != {strict}"

got_probe = [f"D{sr - ROWS_PER_CAT}" for sr in sub_rows] + [f"D{sr - 1}" for sr in sub_rows]
assert_dropdown(GL, got_probe, ("Yes", "No"), True)
assert_dropdown(GL, [DINNER_WEEK_CELL], ("1", "2", "3", "4"), True)
assert_dropdown(PI, [f"E{PAN_FIRST}", f"E{PAN_LAST}"], ("Yes", "No"), True)
assert_dropdown(PI, [f"C{PAN_FIRST}", f"C{PAN_LAST}"], PANTRY_PLACES, False)
# subtotal rows must NOT sit inside the Got-it dropdown
for sr in sub_rows:
    dv = covered(wb[GL], f"D{sr}")
    assert dv is None, f"{GL}!D{sr}: subtotal row caught by a dropdown"

# 5e: native bar chart, real ranges, house palette (P1-2)
assert chart_parts == [CHART_PART], f"expected exactly one chart part, got {chart_parts}"
assert "barChart" in chart_xml, "chart is not a bar chart"
for ref in ("$C$9:$C$12", "$D$9:$D$12", "$B$9:$B$12"):
    assert ref in chart_xml, f"chart does not reference the real range {ref}"
assert len(re.findall(r"<(?:c:)?ser>", chart_xml)) == 2, "chart must plot exactly Budget + Actual"
for color in ("1F3A5F", "2E7D6B"):
    assert color in chart_xml, f"house-palette color {color} missing from chart"
assert len(wb[MB]._charts) == 1, "chart not anchored on the Monthly Grocery Budget tab"

# 5f: sheet protection — formulas locked, yellow inputs unlocked (P2-6)
yellow_unlocked = 0
for sheet in wb.sheetnames:
    ws = wb[sheet]
    assert ws.protection.sheet, f"{sheet}: sheet protection is off"
    assert not ws.protection.password, f"{sheet}: protection must be password-free"
    for row in ws.iter_rows():
        for cell in row:
            if is_yellow(cell):
                assert cell.protection.locked is False, \
                    f"{sheet}!{cell.coordinate}: yellow input cell is locked"
                yellow_unlocked += 1
assert yellow_unlocked > 200, f"suspiciously few unlocked inputs: {yellow_unlocked}"

# 5g: print setup on every tab (P2-7)
ORIENT = {TABS[0]: "portrait", MP: "landscape", GL: "landscape",
          PB: "landscape", MB: "portrait", PI: "landscape"}
for sheet in wb.sheetnames:
    ws = wb[sheet]
    assert ws.print_area, f"{sheet}: no print area"
    assert ws.page_setup.orientation == ORIENT[sheet], \
        f"{sheet}: orientation {ws.page_setup.orientation}, expected {ORIENT[sheet]}"
    assert int(ws.page_setup.fitToWidth or 0) == 1, f"{sheet}: not fit-to-width"
    pspr = ws.sheet_properties.pageSetUpPr
    assert pspr is not None and pspr.fitToPage, f"{sheet}: fitToPage flag missing"
for sheet, titles in ((GL, "$10:$10"), (PB, "$6:$6"), (PI, "$8:$8")):
    got = wb[sheet].print_title_rows
    assert got and got.replace("$", "") == titles.replace("$", ""), \
        f"{sheet}: repeat-header rows are {got!r}"

print(f"VERIFIED: {seen} formula strings exact · {evaluated} evaluated in Python · "
      f"{cached} cached values checked{'' if have_cached else ' (recalc not run — cached pass skipped)'}")
print(f"  grocery list: ${list_total:,.2f} of ${WEEKLY_BUDGET} -> left vs budget ${left_vs_budget:,.2f} (green); in cart ${in_cart:,.2f}")
print(f"  meal plan: week 1 counter {meal_count_w1}/28, weeks 2-4 = 0")
print(f"  dinner panel: week picker=1 -> {dinners_w1[0]!r} ... {dinners_w1[-1]!r} pulled live")
print(f"  month: spent ${month_spent:,.2f} of ${MONTH_BUDGET} -> left ${month_left:,.2f}, {pct_used:.1%} used, ${vs_last:,.2f} less than last month")
print(f"  weekly over/under {week_diffs} -> green/red/amber/green all exercised; 0 renders $0.00 not '-'")
print(f"  price book: {len(PRICE_ROWS)} rows, cheapest = {[s for s, _ in pb_expected]}")
print(f"  pantry: {pantry_items} items, {pantry_flags} use-first flags")
print(f"  dropdowns: Got it? Yes/No · week 1-4 · Use first? Yes/No · Where {'/'.join(PANTRY_PLACES)}")
print(f"  chart: clustered bars, Budget(C9:C12, navy) vs Actual(D9:D12, teal) on {MB}")
print(f"  protection: 6 sheets locked (no password), {yellow_unlocked} yellow input cells unlocked; print setup on all tabs")
