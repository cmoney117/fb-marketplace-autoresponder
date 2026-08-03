#!/usr/bin/env python3
"""Build the Meal Plan & Grocery Budget Planner spreadsheet product (Wave-1 sibling).

Deterministic: generates Meal-Plan-Grocery-Budget-Planner.xlsx (formulas only —
NO macros, works in Excel and free Google Sheets). Worked-example source data
lives in the module-level constants below so verify_mealplanner.py can import
them and recompute every expected value independently.

Run: python3 build_mealplanner.py
Then verify: python3 verify_mealplanner.py  (run scripts/recalc.py in between
for the cached-value pass, see listing.md QA status).
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule

OUT = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/meal-grocery-planner-v1"
XLSX = os.path.join(OUT, "Meal-Plan-Grocery-Budget-Planner.xlsx")

NAVY = "1F3A5F"; TEAL = "2E7D6B"; GOLD = "F2C14E"; LIGHT = "F4F6F8"; YELLOW = "FFF2CC"
H1 = Font(name="Arial", size=18, bold=True, color="FFFFFF")
H2 = Font(name="Arial", size=12, bold=True, color="FFFFFF")
B = Font(name="Arial", size=11)
BB = Font(name="Arial", size=11, bold=True)
SMALL = Font(name="Arial", size=10, italic=True, color="666666")
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00;($#,##0.00);"-"'
PCT = "0.0%"
# traffic-light fills (classic Excel good/neutral/bad — same trio as simple-budget-starter)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE"); GREEN_FONT = Font(name="Arial", color="006100")
YEL_FILL = PatternFill("solid", fgColor="FFEB9C"); YEL_FONT = Font(name="Arial", color="9C6500")
RED_FILL = PatternFill("solid", fgColor="FFC7CE"); RED_FONT = Font(name="Arial", color="9C0006")

# ------------------------------------------------------------------ source data
# One real week for a family of four. Numbers are typical mid-2020s US
# supermarket prices entered by us as example data (not sourced from any list).
DAYS = ("Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun")
WEEK1_MEALS = {  # 4 meal rows x 7 days = 28 cells, all filled in the example
    "Breakfast": ("Oatmeal + fruit", "Eggs & toast", "Yogurt parfaits", "Oatmeal + fruit",
                  "Smoothies", "Pancakes", "Egg scramble"),
    "Lunch": ("Turkey wraps", "Leftover chili", "PB&J + apples", "Tuna melts",
              "Leftover stir-fry", "Grilled cheese + soup", "Deli sandwiches"),
    "Dinner": ("Sheet-pan chicken & veg", "Big-batch chili", "Taco night", "Pasta marinara",
               "Chicken stir-fry", "Homemade pizza", "Leftovers night"),
    "Snack": ("Apples + peanut butter", "Popcorn", "Carrots + hummus", "Cheese & crackers",
              "Bananas", "Trail mix", "Frozen berry yogurt"),
}
MEAL_ROWS = ("Breakfast", "Lunch", "Dinner", "Snack")

WEEKLY_BUDGET = 150  # yellow cell on Grocery List Builder
ROWS_PER_CAT = 8     # item rows per category (extra rows left blank for the buyer)
GROCERY = [  # (category, [(item, est. price, got-it flag "Yes"/"")])
    ("PRODUCE", [
        ("Bananas", 2.28, "Yes"), ("Apples (3 lb bag)", 4.49, "Yes"),
        ("Carrots (2 lb)", 2.29, ""), ("Salad mix", 3.99, ""),
        ("Bell peppers (3)", 3.49, ""), ("Broccoli", 2.79, ""),
        ("Onions (3 lb bag)", 2.99, ""), ("Tomatoes", 2.99, "")]),
    ("MEAT & SEAFOOD", [
        ("Chicken thighs (3 lb)", 8.97, "Yes"), ("Ground beef (2 lb)", 9.98, ""),
        ("Turkey deli slices", 4.99, ""), ("Tuna (4 cans)", 4.36, "")]),
    ("DAIRY & EGGS", [
        ("Milk (2 gal)", 7.18, "Yes"), ("Eggs (18 ct)", 4.79, ""),
        ("Cheddar block", 5.49, ""), ("Yogurt (large tub)", 3.79, ""), ("Butter", 3.99, "")]),
    ("PANTRY & DRY GOODS", [
        ("Oats (large canister)", 4.29, ""), ("Pasta (2 boxes)", 2.58, ""),
        ("Marinara (2 jars)", 4.58, ""), ("Rice (5 lb)", 5.49, ""),
        ("Peanut butter", 3.79, ""), ("Bread (2 loaves)", 5.18, ""),
        ("Tortillas", 3.29, ""), ("Taco seasoning", 0.99, "")]),
    ("FROZEN", [
        ("Frozen veg mix (3 bags)", 3.87, ""), ("Frozen berries", 3.99, ""),
        ("Pizza dough", 2.49, "")]),
    ("HOUSEHOLD", [
        ("Dish soap", 2.99, ""), ("Paper towels", 5.99, ""), ("Trash bags", 6.49, "")]),
]

STORES = ("Store 1", "Store 2", "Store 3")  # yellow headers — buyer renames to real stores
PRICE_ROWS = [  # (item, size, price@store1, price@store2, price@store3) — unique min per row
    ("Milk (1 gal)", "1 gal", 3.89, 3.59, 3.29),
    ("Eggs (18 ct)", "18 ct", 4.97, 5.29, 4.39),
    ("Chicken breast", "per lb", 3.49, 2.99, 3.19),
    ("Ground beef 80/20", "per lb", 4.87, 4.99, 4.49),
    ("Sandwich bread", "loaf", 2.59, 2.49, 1.99),
    ("Bananas", "per lb", 0.54, 0.59, 0.44),
    ("Peanut butter", "16 oz", 2.48, 3.29, 2.79),
    ("Pasta", "16 oz", 1.24, 0.99, 1.09),
    ("Rice", "5 lb", 4.97, 5.49, 5.29),
    ("Cheddar cheese", "8 oz", 2.87, 2.50, 2.29),
]
PB_FIRST, PB_LAST = 7, 31  # 25 price-book rows (blank ones ready for the buyer)

MONTH_BUDGET = 600       # yellow cell — monthly grocery budget
LAST_MONTH_TOTAL = 655   # yellow cell — last month's grocery total
WEEK_ACTUALS = (138.42, 161.85, 150.00, 132.90)  # exercises green/red/amber/green lights

PANTRY = [  # (item, where, qty, use-first "Yes"/"", notes)
    ("Chicken thighs (2 lb)", "Freezer", "2 lb", "Yes", "plan a dinner around these first"),
    ("Rice (white)", "Pantry", "5 lb", "", ""),
    ("Black beans", "Pantry", "4 cans", "", ""),
    ("Pasta", "Pantry", "2 boxes", "", ""),
    ("Marinara sauce", "Fridge", "1 open jar", "Yes", "open — use this week"),
    ("Oats", "Pantry", "1 canister", "", ""),
    ("Yogurt (tub)", "Fridge", "1", "Yes", "expires this week"),
    ("Frozen veg mix", "Freezer", "3 bags", "", ""),
    ("Tortillas", "Fridge", "1 pack", "", ""),
    ("Bread", "Freezer", "1 loaf", "", ""),
]
PAN_FIRST, PAN_LAST = 9, 58  # 50 pantry rows


def banner(ws, ref, text, font=None):
    ws.merge_cells(ref)
    c = ws[ref.split(":")[0]]
    c.value = text
    c.font = font or H1
    c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[c.row].height = 30


def traffic(ws, rng):
    """Green above 0, amber at exactly 0, red below 0."""
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL, font=GREEN_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["0"], fill=YEL_FILL, font=YEL_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT))


def build():
    wb = Workbook()

    # ---------- Tab 1: START HERE ----------
    ws = wb.active; ws.title = "START HERE"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDE", (40, 30, 30, 30)): ws.column_dimensions[c].width = w
    ws.merge_cells("B2:E2"); ws["B2"] = "MEAL PLAN & GROCERY BUDGET PLANNER"
    ws["B2"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
    ws["B2"].fill = HEAD_FILL; ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 34
    rows = [
        ("", ""),
        ("Welcome! The cheapest groceries are the ones you planned. This planner runs the weekly loop that", ""),
        ("shrinks a grocery bill: check the pantry, plan the meals, price the list, then shop with a number in mind.", ""),
        ("", ""),
        ("HOW TO USE (4 steps)", "SEC"),
        ("1. 'Pantry Inventory' — flag Use first? = Yes on anything that should get eaten soon.", ""),
        ("2. 'Weekly Meal Plan' — plan 7 days around those use-first items. Dinners first; Week 1 is a full example.", ""),
        ("3. 'Grocery List Builder' — set your weekly budget in the yellow cell, list what you need with estimated", ""),
        ("    prices, and watch 'Left vs budget'. If it goes red, swap or cut items BEFORE you're in the store.", ""),
        ("4. 'Monthly Grocery Budget' — after each week's shop, log what you really spent. The month view shows", ""),
        ("    budget left and how the month compares to last month.", ""),
        ("", ""),
        ("THE OTHER TAB", "SEC"),
        ("• Price Book — jot prices for your 20 staples across 3 stores (rename the yellow store headers).", ""),
        ("   The cheapest price lights up green, so you learn where each staple genuinely costs less —", ""),
        ("   and you can tell a real sale from a shelf sign that just says SALE.", ""),
        ("", ""),
        ("THE WORKED EXAMPLE (already in the sheet — replace it with your own)", "SEC"),
        ("One real week for a family of four: a $150 weekly budget, a fully planned Week 1, and a priced", ""),
        ("grocery list that comes to $134.86 — so 'Left vs budget' shows $15.14 green before you leave the house.", ""),
        ("The monthly tab shows $583.17 spent of a $600 budget, $71.83 less than the month before.", ""),
        ("", ""),
        ("COLOR LEGEND", "SEC"),
        ("Yellow cells = yours to edit.  Everything else calculates automatically — please don't type over it.", ""),
        ("Traffic lights: green = under budget (or cheapest price) · amber = right at it (or use-first) · red = over.", ""),
        ("", ""),
        ("GOOGLE SHEETS USERS", "SEC"),
        ("Upload this file to Google Drive, then open it — it converts automatically and all formulas work.", ""),
        ("", ""),
        ("Questions or a formula acting up? Message us on Etsy any time — we answer fast and we'll fix it.", ""),
    ]
    r = 4
    for text, kind in rows:
        cell = ws.cell(row=r, column=2, value=text)
        if kind == "SEC":
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            cell.font = H2; cell.fill = SEC_FILL
            ws.row_dimensions[r].height = 20
        else:
            cell.font = B
        r += 1

    # ---------- Tab 2: Weekly Meal Plan ----------
    ws = wb.create_sheet("Weekly Meal Plan")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 21  # fits the "Meals planned (of 28):" label
    for c in "CDEFGHI": ws.column_dimensions[c].width = 17
    banner(ws, "B2:I2", "WEEKLY MEAL PLAN — 4 WEEKS")
    ws["B4"] = ("Week 1 is a full worked example — clear it and plan yours. Start with dinners (the budget-buster meal) "
                "and build them around Pantry Inventory 'use first' flags.")
    ws["B4"].font = SMALL
    for w in range(4):  # week block: header s · days s+1 · meals s+2..s+5 · counter s+6 · spacer
        s = 6 + w * 8
        ws.merge_cells(start_row=s, start_column=2, end_row=s, end_column=9)
        t = ws.cell(row=s, column=2, value=f"WEEK {w + 1}"); t.font = H2; t.fill = SEC_FILL
        ws.row_dimensions[s].height = 20
        for i, d in enumerate(DAYS):
            c = ws.cell(row=s + 1, column=3 + i, value=d); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
        ws.cell(row=s + 1, column=2, value="Meal").font = BB
        ws.cell(row=s + 1, column=2).fill = ALT_FILL; ws.cell(row=s + 1, column=2).border = BORDER
        for j, meal in enumerate(MEAL_ROWS):
            rr = s + 2 + j
            ws.cell(row=rr, column=2, value=meal).font = BB
            ws.cell(row=rr, column=2).border = BORDER
            for i in range(7):
                v = WEEK1_MEALS[meal][i] if w == 0 else None
                c = ws.cell(row=rr, column=3 + i, value=v)
                c.fill = INPUT_FILL; c.border = BORDER; c.font = B
                c.alignment = Alignment(wrap_text=True, vertical="top")
        cr = s + 6
        ws.cell(row=cr, column=2, value="Meals planned (of 28):").font = BB
        cc = ws.cell(row=cr, column=3, value=f"=COUNTA(C{s + 2}:I{s + 5})")
        cc.font = BB; cc.fill = ALT_FILL; cc.border = BORDER
    ws.freeze_panes = "A6"

    # ---------- Tab 3: Grocery List Builder ----------
    ws = wb.create_sheet("Grocery List Builder")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDE", (30, 13, 10, 28)): ws.column_dimensions[c].width = w
    banner(ws, "B2:E2", "GROCERY LIST BUILDER")
    ws["B4"] = "Weekly grocery budget:"; ws["B4"].font = BB
    ws["C4"] = WEEKLY_BUDGET; ws["C4"].fill = INPUT_FILL; ws["C4"].number_format = MONEY; ws["C4"].border = BORDER
    ws["B5"] = "Build the list from your meal plan. If 'Left vs budget' goes red, swap or cut items BEFORE the store — that's the whole trick."
    ws["B5"].font = SMALL
    ws["B6"] = "Grocery list total (est.):"; ws["B6"].font = BB
    ws["C6"] = "=C20+C30+C40+C50+C60+C70"; ws["C6"].font = BB; ws["C6"].number_format = MONEY
    ws["C6"].fill = ALT_FILL; ws["C6"].border = BORDER
    ws["B7"] = "Left vs budget:"; ws["B7"].font = BB
    ws["C7"] = "=C4-C6"; ws["C7"].font = BB; ws["C7"].number_format = MONEY; ws["C7"].border = BORDER
    ws["B8"] = "In cart so far (marked Yes):"; ws["B8"].font = BB
    ws["C8"] = '=SUMIF(D12:D69,"Yes",C12:C69)'; ws["C8"].font = BB; ws["C8"].number_format = MONEY
    ws["C8"].fill = ALT_FILL; ws["C8"].border = BORDER
    ws["B9"] = "Mark Got it? = Yes as you shop — 'In cart so far' keeps a running total. Type Yes exactly."
    ws["B9"].font = SMALL
    traffic(ws, "C7")
    hdrs = ["Item", "Est. price", "Got it?", "Notes"]
    for i, h in enumerate(hdrs):
        c = ws.cell(row=10, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
    r = 11
    for cat, items in GROCERY:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        t = ws.cell(row=r, column=2, value=cat); t.font = H2; t.fill = SEC_FILL
        ws.row_dimensions[r].height = 18
        first = r + 1
        for j in range(ROWS_PER_CAT):
            rr = first + j
            item, price, got = items[j] if j < len(items) else ("", None, "")
            ws.cell(row=rr, column=2, value=item or None).font = B
            p = ws.cell(row=rr, column=3, value=price); p.number_format = MONEY
            g = ws.cell(row=rr, column=4, value=got or None); g.font = B
            n = ws.cell(row=rr, column=5); n.font = B
            for cc in (ws.cell(row=rr, column=2), p, g, n):
                cc.fill = INPUT_FILL; cc.border = BORDER
        sub = first + ROWS_PER_CAT
        ws.cell(row=sub, column=2, value=f"{cat.title()} subtotal").font = BB
        sc = ws.cell(row=sub, column=3, value=f"=SUM(C{first}:C{sub - 1})")
        sc.font = BB; sc.number_format = MONEY
        for col in range(2, 6):
            c = ws.cell(row=sub, column=col); c.fill = ALT_FILL; c.border = BORDER
        r = sub + 1
    ws.freeze_panes = "A11"

    # ---------- Tab 4: Price Book ----------
    ws = wb.create_sheet("Price Book")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEFGH", (24, 12, 13, 13, 13, 16, 14)): ws.column_dimensions[c].width = w
    banner(ws, "B2:H2", "PRICE BOOK")
    ws["B4"] = ("The quiet money-saver: jot the price of your regular staples at up to 3 stores as you shop. "
                "In a month you'll know where each one genuinely costs less — and you'll spot a real sale instantly.")
    ws["B4"].font = SMALL
    ws["B5"] = "The store names are yellow — replace them with your real stores. The cheapest price in each row lights up green."
    ws["B5"].font = SMALL
    hdrs = ["Item", "Size", STORES[0], STORES[1], STORES[2], "Cheapest store", "Cheapest price"]
    for i, h in enumerate(hdrs):
        c = ws.cell(row=6, column=2 + i, value=h); c.font = BB; c.border = BORDER
        c.fill = INPUT_FILL if 2 <= i <= 4 else ALT_FILL
    for r in range(PB_FIRST, PB_LAST + 1):
        item, size, p1, p2, p3 = PRICE_ROWS[r - PB_FIRST] if r - PB_FIRST < len(PRICE_ROWS) else ("", "", None, None, None)
        ws.cell(row=r, column=2, value=item or None).font = B
        ws.cell(row=r, column=3, value=size or None).font = B
        for k, p in enumerate((p1, p2, p3)):
            c = ws.cell(row=r, column=4 + k, value=p); c.number_format = MONEY
        st = ws.cell(row=r, column=7, value=f'=IF(COUNT(D{r}:F{r})=0,"",INDEX($D$6:$F$6,MATCH(MIN(D{r}:F{r}),D{r}:F{r},0)))')
        st.font = B
        cp = ws.cell(row=r, column=8, value=f'=IF(COUNT(D{r}:F{r})=0,"",MIN(D{r}:F{r}))')
        cp.number_format = MONEY
        for col in range(2, 9):
            c = ws.cell(row=r, column=col); c.border = BORDER
            if col <= 6: c.fill = INPUT_FILL
    ws.conditional_formatting.add(
        f"D{PB_FIRST}:F{PB_LAST}",
        FormulaRule(formula=[f'AND(D{PB_FIRST}<>"",D{PB_FIRST}=MIN($D{PB_FIRST}:$F{PB_FIRST}))'],
                    fill=GREEN_FILL, font=GREEN_FONT))
    ws.freeze_panes = "A7"

    # ---------- Tab 5: Monthly Grocery Budget ----------
    ws = wb.create_sheet("Monthly Grocery Budget")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDE", (34, 14, 14, 14)): ws.column_dimensions[c].width = w
    banner(ws, "B2:E2", "MONTHLY GROCERY BUDGET")
    ws["B4"] = "Monthly grocery budget:"; ws["B4"].font = BB
    ws["C4"] = MONTH_BUDGET; ws["C4"].fill = INPUT_FILL; ws["C4"].number_format = MONEY; ws["C4"].border = BORDER
    ws["B5"] = "Last month's grocery total:"; ws["B5"].font = BB
    ws["C5"] = LAST_MONTH_TOTAL; ws["C5"].fill = INPUT_FILL; ws["C5"].number_format = MONEY; ws["C5"].border = BORDER
    ws["B6"] = ("Weekly budget = monthly budget ÷ 4 (calculated for you). Log each week's REAL total after your main shop — "
                "include the little top-up runs. Months with a 5th shopping week: log that shop into Week 4.")
    ws["B6"].font = SMALL
    hdrs = ["Week", "Budget", "Actual spent", "Over / under"]
    for i, h in enumerate(hdrs):
        c = ws.cell(row=8, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
    for i, actual in enumerate(WEEK_ACTUALS):
        r = 9 + i
        ws.cell(row=r, column=2, value=f"Week {i + 1}").font = B
        bc = ws.cell(row=r, column=3, value="=$C$4/4"); bc.number_format = MONEY
        ac = ws.cell(row=r, column=4, value=actual); ac.fill = INPUT_FILL; ac.number_format = MONEY
        dc = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); dc.number_format = MONEY
        for col in range(2, 6): ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=13, column=2, value="Month total").font = BB
    ws.cell(row=13, column=3, value="=SUM(C9:C12)").font = BB
    ws.cell(row=13, column=4, value="=SUM(D9:D12)").font = BB
    ws.cell(row=13, column=5, value="=C13-D13").font = BB
    for col in range(2, 6):
        c = ws.cell(row=13, column=col); c.fill = ALT_FILL; c.border = BORDER
        if col > 2: c.number_format = MONEY
    traffic(ws, "E9:E13")
    ws.merge_cells("B15:E15")
    ws["B15"] = "THE BOTTOM LINE"; ws["B15"].font = H2; ws["B15"].fill = HEAD_FILL
    lines = [
        ("Spent this month:", "=D13", MONEY),
        ("Left in budget:", "=C4-D13", MONEY),
        ("% of budget used:", '=IF(C4=0,"",D13/C4)', PCT),
        ("vs last month (+ means you spent less):", "=C5-D13", MONEY),
    ]
    for i, (lab, formula, fmt) in enumerate(lines):
        r = 16 + i
        ws.cell(row=r, column=2, value=lab).font = BB
        fc = ws.cell(row=r, column=3, value=formula); fc.font = BB; fc.number_format = fmt; fc.border = BORDER
    traffic(ws, "C17"); traffic(ws, "C19")
    ws["B21"] = ("'vs last month' just shows the difference — when it turns green, it's usually the pantry flags, "
                 "the priced list, and the Price Book doing their job.")
    ws["B21"].font = SMALL
    ws.freeze_panes = "A8"

    # ---------- Tab 6: Pantry Inventory ----------
    ws = wb.create_sheet("Pantry Inventory")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for c, w in zip("BCDEF", (26, 14, 13, 12, 34)): ws.column_dimensions[c].width = w
    banner(ws, "B2:F2", "PANTRY INVENTORY")
    ws["B4"] = "Items on hand:"; ws["B4"].font = BB
    ws["C4"] = f"=COUNTA(B{PAN_FIRST}:B{PAN_LAST})"; ws["C4"].font = BB; ws["C4"].fill = ALT_FILL; ws["C4"].border = BORDER
    ws["B5"] = "Flagged use-first:"; ws["B5"].font = BB
    ws["C5"] = f'=COUNTIF(E{PAN_FIRST}:E{PAN_LAST},"Yes")'; ws["C5"].font = BB; ws["C5"].fill = ALT_FILL; ws["C5"].border = BORDER
    ws["B6"] = ("Shop your kitchen first. Flag Use first? = Yes on anything ageing out, then plan this week's dinners "
                "around those rows — that's food you already paid for. Type Yes exactly; the counter counts that word.")
    ws["B6"].font = SMALL
    hdrs = ["Item", "Where", "Qty on hand", "Use first?", "Notes"]
    for i, h in enumerate(hdrs):
        c = ws.cell(row=8, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
    for r in range(PAN_FIRST, PAN_LAST + 1):
        item, where, qty, use, note = PANTRY[r - PAN_FIRST] if r - PAN_FIRST < len(PANTRY) else ("", "", "", "", "")
        for j, v in enumerate((item, where, qty, use, note)):
            c = ws.cell(row=r, column=2 + j, value=v or None)
            c.fill = INPUT_FILL; c.border = BORDER; c.font = B
    ws.conditional_formatting.add(
        f"E{PAN_FIRST}:E{PAN_LAST}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=YEL_FILL, font=YEL_FONT))
    ws.freeze_panes = "A9"

    wb.save(XLSX)
    print("saved", XLSX)


if __name__ == "__main__":
    build()
