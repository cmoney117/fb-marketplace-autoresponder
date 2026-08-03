#!/usr/bin/env python3
"""Build the Savings Goal Tracker workbook deterministically.

History: v1 originally shipped without a generator script. This script first
reproduced the shipped Savings-Goal-Tracker.xlsx cell-for-cell (verified with a
full style/value diff on 2026-08-03), then the P1 review fixes were applied on
top:
  - "Funded by (auto)" projected-date column on My Goals (honors the cover's
    "a number, a date, and a plan" promise with a real EDATE formula)
  - native bar chart of per-goal progress on My Goals + data-bar conditional
    formatting on the "% there" column
  - dropdown (data validation) on the Savings Log "Goal" column fed by the
    goal names on My Goals, and an "x" dropdown on the 52-Week "Done?" column
  - hero cells ("Save per month" plan cells + totals) render $0.00 instead of
    the accounting dash when a goal is fully funded
  - START HERE copy updated to match the file (no phantom "target month" input)
No macros — formulas + native features only (Excel 2010+ / Google Sheets).
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

OUT = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/savings-goal-tracker-v1"
os.makedirs(OUT, exist_ok=True)

NAVY = "1F3A5F"; TEAL = "2E7D6B"; GOLD = "F2C14E"; LIGHT = "F4F6F8"; YELLOW = "FFF2CC"
H1 = Font(name="Arial", size=18, bold=True, color="FFFFFF")
H2 = Font(name="Arial", size=12, bold=True, color="FFFFFF")
B = Font(name="Arial", size=11)
BB = Font(name="Arial", size=11, bold=True)
SMALL = Font(name="Arial", size=10, italic=True, color="666666")
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00;($#,##0.00);"-"'
MONEY0 = "$#,##0.00;($#,##0.00)"  # hero cells: 0 renders $0.00, never "-"
PCT = "0.0%"
DATEFMT = "mmm yyyy"

GOALS = [  # (name, target, saved, months-left) — 4 worked examples + 4 blank rows
    ("Emergency fund", 3000, 850, 10),
    ("Vacation", 1800, 300, 9),
    ("New tires", 600, 150, 4),
    ("Christmas fund", 900, 200, 4),
]
N_GOAL_ROWS = 8

wb = Workbook()

# ---------- Tab 1: START HERE ----------
ws = wb.active; ws.title = "START HERE"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDE", (34, 30, 30, 30)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:E2"); ws["B2"] = "SAVINGS GOAL TRACKER"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
rows = [
    (4, "Every goal gets a number, a date, and a plan. Then it actually happens.", None),
    (5, "", None),
    (6, "HOW TO USE", "SEC"),
    (7, "1. 'My Goals' tab: name each goal, target amount, months left, and what you've saved so far (yellow cells).", None),
    (8, "2. The sheet shows how much to save per month — and the month each goal is funded ('Funded by').", None),
    (9, "3. '52-Week Challenge' tab: the classic challenge — tick off a deposit each week, watch it snowball to $1,378.", None),
    (10, "4. 'Savings Log' tab: log each deposit (pick the goal from the dropdown), then update 'Saved so far' on My Goals.", None),
    (11, "", None),
    (12, "COLOR LEGEND", "SEC"),
    (13, "Yellow cells = yours to edit. Everything else calculates automatically.", None),
    (14, "Pre-filled numbers are a realistic example — replace with your own.", None),
    (15, "", None),
    (16, "GOOGLE SHEETS USERS", "SEC"),
    (17, "Upload this file to Google Drive, then open it — it converts automatically and all formulas work.", None),
    (18, "The progress chart and dropdowns convert too; chart colors may shift slightly in Sheets. The numbers are identical.", None),
    (19, "", None),
    (20, "Questions? Message us — real humans, fast replies.", None),
]
for r, text, kind in rows:
    cell = ws.cell(row=r, column=2, value=text)
    if kind == "SEC":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell.font = H2; cell.fill = SEC_FILL
        ws.row_dimensions[r].height = 20
    else:
        cell.font = B

# ---------- Tab 2: My Goals ----------
ws = wb.create_sheet("My Goals")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDEFGH", (22, 14, 14, 14, 12, 18, 14)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:H2"); ws["B2"] = "MY GOALS"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
ws["B4"] = ("'Months left' = how many months until you want that goal fully funded. "
            "'Funded by' = the month you'll get there on that plan. Yellow cells are yours to edit.")
ws["B4"].font = SMALL
hdrs = ["Goal", "Target $", "Saved so far", "Months left", "% there", "Save per month", "Funded by (auto)"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=6, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
G_FIRST, G_LAST = 7, 6 + N_GOAL_ROWS  # 7..14
for i in range(N_GOAL_ROWS):
    r = G_FIRST + i
    name, target, saved, months = GOALS[i] if i < len(GOALS) else (None, None, None, None)
    ws.cell(row=r, column=2, value=name)
    c = ws.cell(row=r, column=3, value=target); c.number_format = MONEY
    d = ws.cell(row=r, column=4, value=saved); d.number_format = MONEY
    ws.cell(row=r, column=5, value=months)
    f = ws.cell(row=r, column=6, value=f'=IF(OR(C{r}="",C{r}=0),"",D{r}/C{r})'); f.number_format = PCT
    g = ws.cell(row=r, column=7, value=f'=IF(OR(C{r}="",E{r}="",E{r}=0),"",MAX(0,(C{r}-D{r})/E{r}))'); g.number_format = MONEY0
    h = ws.cell(row=r, column=8,
                value=f'=IF(OR(C{r}="",E{r}="",E{r}=0),"",IF(D{r}>=C{r},"FUNDED!",EDATE(TODAY(),E{r})))')
    h.number_format = DATEFMT
    for col in range(2, 9):
        cc = ws.cell(row=r, column=col); cc.border = BORDER
        if col <= 5: cc.fill = INPUT_FILL
G_TOT = G_LAST + 1  # 15
ws.cell(row=G_TOT, column=2, value="Total").font = BB
ct = ws.cell(row=G_TOT, column=3, value=f"=SUM(C{G_FIRST}:C{G_LAST})"); ct.font = BB; ct.number_format = MONEY0
dt = ws.cell(row=G_TOT, column=4, value=f"=SUM(D{G_FIRST}:D{G_LAST})"); dt.font = BB; dt.number_format = MONEY0
gt = ws.cell(row=G_TOT, column=7, value=f'=SUMIF(G{G_FIRST}:G{G_LAST},">0")'); gt.font = BB; gt.number_format = MONEY0
for col in range(2, 9):
    cc = ws.cell(row=G_TOT, column=col); cc.fill = ALT_FILL; cc.border = BORDER

# data bars on "% there" — progress bars that survive Google Sheets import untouched
ws.conditional_formatting.add(
    f"F{G_FIRST}:F{G_LAST}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                color=TEAL, showValue=True, minLength=None, maxLength=None))

# native per-goal progress chart, driven by the live ranges (updates as goals are edited)
chart = BarChart()
chart.type = "bar"  # horizontal bars: one progress bar per goal
chart.style = 10
chart.title = "How close is each goal? (% there)"
data = Reference(ws, min_col=6, min_row=G_FIRST, max_row=G_LAST)  # % there
cats = Reference(ws, min_col=2, min_row=G_FIRST, max_row=G_LAST)  # goal names
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)
ser = chart.series[0]
ser.graphicalProperties = GraphicalProperties(solidFill=TEAL)
chart.legend = None
chart.y_axis.scaling.min = 0   # value axis (horizontal on a bar chart)
chart.y_axis.scaling.max = 1
chart.y_axis.numFmt = "0%"
chart.y_axis.majorGridlines = None
chart.width = 18
chart.height = 9
ws.add_chart(chart, f"B{G_TOT + 3}")

# ---------- Tab 3: 52-Week Challenge ----------
ws = wb.create_sheet("52-Week Challenge")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDE", (10, 14, 16, 12)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:E2"); ws["B2"] = "52-WEEK CHALLENGE"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
for i, h in enumerate(["Week", "Deposit", "Total saved", "Done? (x)"]):
    c = ws.cell(row=4, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
W_FIRST, W_LAST = 5, 56
for wk in range(1, 53):
    r = W_FIRST + wk - 1
    ws.cell(row=r, column=2, value=wk).font = B
    c = ws.cell(row=r, column=3, value=wk); c.number_format = MONEY
    d = ws.cell(row=r, column=4, value=f"=SUM(C{W_FIRST}:C{r})"); d.number_format = MONEY
    e = ws.cell(row=r, column=5); e.fill = INPUT_FILL
    for col in range(2, 6): ws.cell(row=r, column=col).border = BORDER
ws.cell(row=58, column=2, value="Weeks done:").font = BB
c58 = ws.cell(row=58, column=3, value=f'=COUNTIF(E{W_FIRST}:E{W_LAST},"<>"&"")'); c58.font = BB
ws.cell(row=59, column=2, value="Flip it: do week 52 first ($52) while motivation is highest.").font = SMALL
dv_x = DataValidation(type="list", formula1='"x"', allow_blank=True, showDropDown=False)
dv_x.error = "Pick x from the dropdown (or leave blank)."
dv_x.errorTitle = "Done marker"
dv_x.prompt = "Pick x when the deposit is done."
dv_x.promptTitle = "Done?"
ws.add_data_validation(dv_x)
dv_x.add(f"E{W_FIRST}:E{W_LAST}")

# ---------- Tab 4: Savings Log ----------
ws = wb.create_sheet("Savings Log")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDE", (14, 22, 14, 26)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:E2"); ws["B2"] = "SAVINGS LOG"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
for i, h in enumerate(["Date", "Goal", "Amount", "Note"]):
    c = ws.cell(row=4, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
L_FIRST, L_LAST = 5, 44
example = ("2026-08-01", "Emergency fund", 50, "example row — replace me")
for r in range(L_FIRST, L_LAST + 1):
    vals = example if r == L_FIRST else (None, None, None, None)
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=2 + j, value=v)
        c.fill = INPUT_FILL; c.border = BORDER
        if j == 2: c.number_format = MONEY
ws.cell(row=46, column=2, value="Total logged").font = BB
t = ws.cell(row=46, column=4, value=f"=SUM(D{L_FIRST}:D{L_LAST})"); t.font = BB; t.number_format = MONEY0
# goal dropdown fed by the My Goals names (defined name survives Excel 2010 + Sheets import)
wb.defined_names.add(DefinedName("GoalList", attr_text=f"'My Goals'!$B${G_FIRST}:$B${G_LAST}"))
dv_goal = DataValidation(type="list", formula1="=GoalList", allow_blank=True, showDropDown=False)
dv_goal.error = "Pick a goal from the dropdown — it keeps names matching the My Goals tab."
dv_goal.errorTitle = "Pick a goal"
dv_goal.prompt = "Pick the goal this deposit belongs to."
dv_goal.promptTitle = "Goal"
ws.add_data_validation(dv_goal)
dv_goal.add(f"C{L_FIRST}:C{L_LAST}")

# print setup: landscape, fit to one page wide — Ctrl+P gives whole tabs, not confetti
from openpyxl.worksheet.properties import PageSetupProperties
for sheet in wb.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

path = os.path.join(OUT, "Savings-Goal-Tracker.xlsx")
wb.save(path)
print("saved", path)
