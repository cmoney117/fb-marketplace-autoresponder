#!/usr/bin/env python3
"""Verify Savings-Goal-Tracker.xlsx — structure + real formula recalc.

Checks the P1 review fixes stay fixed:
  1. "Funded by (auto)" projected-date column exists and computes a real date
     (the cover's "a number, a date, and a plan" is now true in the file)
  2. dropdowns: Savings Log Goal column (fed by My Goals names via the GoalList
     defined name) and 52-Week Done? column ("x")
  3. native per-goal progress chart wired to the live My Goals ranges,
     plus data-bar conditional formatting on "% there"
  4. hero cells (Save per month + totals) render $0.00 at zero, never "-"
  5. every key formula recalculates to the expected worked-example value
     (ground truth: LibreOffice headless recalc)
"""
import datetime, os, subprocess, sys, tempfile, zipfile
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Savings-Goal-Tracker.xlsx")
MONEY0 = "$#,##0.00;($#,##0.00)"
failures = []

def check(cond, msg):
    print(("  ok  " if cond else "  FAIL") + " " + msg)
    if not cond:
        failures.append(msg)

wb = load_workbook(XLSX)
mg = wb["My Goals"]

print("[1] projected-date column (the 'a date' promise)")
check(mg["H6"].value == "Funded by (auto)", "H6 header present")
check("EDATE(TODAY()" in str(mg["H7"].value) and '"FUNDED!"' in str(mg["H7"].value),
      "H7 computes EDATE(TODAY(), months-left) with a FUNDED! state")
check(all("EDATE" in str(mg[f"H{r}"].value) for r in range(7, 15)), "formula fills all 8 goal rows")
check(mg["H7"].number_format == "mmm yyyy", "date shows as month + year")
sh = wb["START HERE"]
texts = " | ".join(str(c.value) for row in sh.iter_rows() for c in row if c.value)
check("Funded by" in texts and "target month" not in texts,
      "START HERE describes the real inputs (months left) and the new date column")

print("[2] dropdowns")
log = wb["Savings Log"]
dvs = {str(dv.sqref): dv for dv in log.data_validations.dataValidation}
check("C5:C44" in dvs and dvs.get("C5:C44") and dvs["C5:C44"].formula1 == "=GoalList",
      "Savings Log Goal column dropdown fed by GoalList")
check("GoalList" in wb.defined_names and
      wb.defined_names["GoalList"].attr_text == "'My Goals'!$B$7:$B$14",
      "GoalList defined name points at the goal names")
wk = wb["52-Week Challenge"]
dvs_w = {str(dv.sqref): dv for dv in wk.data_validations.dataValidation}
check("E5:E56" in dvs_w and dvs_w["E5:E56"].formula1 == '"x"', "52-Week Done? column x-dropdown")

print("[3] progress chart + data bars")
with zipfile.ZipFile(XLSX) as z:
    charts = [n for n in z.namelist() if n.startswith("xl/charts/chart")]
    chart_xml = z.read(charts[0]).decode() if charts else ""
check(len(charts) == 1, "exactly one chart part in the workbook")
check("'My Goals'!$F$7:$F$14" in chart_xml, "chart values = live '% there' range")
check("'My Goals'!$B$7:$B$14" in chart_xml, "chart categories = live goal names")
check('<barDir val="bar"' in chart_xml, "horizontal per-goal progress bars")
check('val="2E7D6B"' in chart_xml, "teal house-palette bars")
dbar = [r for rng in mg.conditional_formatting for r in rng.rules
        if r.type == "dataBar" and str(rng.sqref) == "F7:F14"]
check(len(dbar) == 1, "data-bar conditional formatting on % there (Sheets-safe)")

print("[4] hero-cell $0 format")
check(all(mg[f"G{r}"].number_format == MONEY0 for r in range(7, 16)),
      "Save per month cells + total render $0.00 at zero")
check(log["D46"].number_format == MONEY0, "Total logged renders $0.00 at zero")

print("[5] formula recalc (LibreOffice ground truth)")
with tempfile.TemporaryDirectory() as td:
    subprocess.run(["soffice", "--headless", "--convert-to", "xlsx", "--outdir", td, XLSX],
                   check=True, capture_output=True, timeout=180)
    calc = load_workbook(os.path.join(td, os.path.basename(XLSX)), data_only=True)
mgc, wkc, logc = calc["My Goals"], calc["52-Week Challenge"], calc["Savings Log"]
def near(a, b, eps=0.01):
    return a is not None and isinstance(a, (int, float)) and abs(a - b) < eps
check(near(mgc["F7"].value, 850 / 3000, 1e-6), "Emergency fund 28.3% there")
check(near(mgc["G7"].value, 215) and near(mgc["G8"].value, 166.67) and
      near(mgc["G9"].value, 112.5) and near(mgc["G10"].value, 175),
      "save-per-month 215 / 166.67 / 112.50 / 175 (cover pins: 215/167/175 rounded)")
check(near(mgc["C15"].value, 6300) and near(mgc["D15"].value, 1500) and near(mgc["G15"].value, 669.17),
      "goal totals 6,300 / 1,500 / 669.17")
today = datetime.date.today()
def add_months(d, n):
    y, m = divmod(d.year * 12 + (d.month - 1) + n, 12)
    return y, m + 1
for r, months in ((7, 10), (8, 9), (9, 4), (10, 4)):
    v = mgc[f"H{r}"].value
    ok = isinstance(v, (datetime.date, datetime.datetime)) and (v.year, v.month) == add_months(today, months)
    check(ok, f"H{r} projected date = today + {months} months ({v})")
check(mgc["H11"].value in ("", None), "blank goal row has no phantom date")
check(near(wkc["D56"].value, 1378), "52-week total saved $1,378")
check(wkc["C58"].value == 0, "weeks done starts at 0")
check(near(logc["D46"].value, 50), "log total = example $50 deposit")

if failures:
    print(f"\nFAILED: {len(failures)} check(s)")
    sys.exit(1)
print("\nALL CHECKS PASSED")
