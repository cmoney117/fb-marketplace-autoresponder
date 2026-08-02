#!/usr/bin/env python3
"""Build the Wedding Budget & Guest Tracker spreadsheet product (sibling listing).

Deterministic: xlsx + 2700x2025 cover PNG + 460x345 store SVG, then programmatic
verification of EVERY formula in the workbook (mini evaluator, no deps beyond
openpyxl/PIL). Exits non-zero if any formula disagrees with expected values.
"""
import os, re, shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageFont

OUT = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/wedding-budget-planner-v1"
STORE_IMG = "/home/user/fb-marketplace-autoresponder/store/site/img"
os.makedirs(OUT, exist_ok=True)

NAVY = "1F3A5F"; TEAL = "2E7D6B"; GOLD = "F2C14E"; LIGHT = "F4F6F8"; YELLOW = "FFF2CC"
H1 = Font(name="Arial", size=18, bold=True, color="FFFFFF")
H2 = Font(name="Arial", size=12, bold=True, color="FFFFFF")
B = Font(name="Arial", size=11)
BB = Font(name="Arial", size=11, bold=True)
SMALL = Font(name="Arial", size=10, italic=True, color="666666")
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00;($#,##0.00);"-"'
PCT = "0%"

# ------------------------------------------------------------------ source data
TOTAL_BUDGET = 24000
CATEGORIES = [  # (name, typical fraction of budget, planned, actual so far)
    ("Venue & catering", 0.40, 9600, 9750),
    ("Photography & video", 0.12, 2900, 2900),
    ("Attire & beauty", 0.09, 2150, 1980.50),
    ("Flowers & decor", 0.10, 2400, 2510),
    ("Music / entertainment", 0.07, 1700, 1700),
    ("Stationery & invites", 0.03, 720, 684.25),
    ("Wedding rings", 0.03, 720, 650),
    ("Transportation", 0.02, 480, 480),
    ("Cake & desserts", 0.02, 480, 495),
    ("Favors & gifts", 0.02, 480, 462),
    ("Officiant & ceremony", 0.02, 480, 450),
    ("Contingency buffer", 0.08, 1890, 0),
]
GUESTS = [  # (party, size, rsvp, meal, gift?, thank-you?)
    ("Sam & Riley Carter", 2, "Yes", "Chicken", "Yes", "Yes"),
    ("Grandma Jean", 1, "Yes", "Vegetarian", "Yes", "No"),
    ("The Ortiz family", 4, "Yes", "Chicken", "No", "No"),
    ("Priya Patel", 1, "Yes", "Beef", "No", "No"),
    ("Marcus & Dee Lewis", 2, "No", "", "No", "No"),
    ("Coach Thompson", 1, "Waiting", "", "No", "No"),
    ("Jules Moreau", 1, "Yes", "Beef", "Yes", "Yes"),
    ("The Kims", 3, "Yes", "Chicken", "No", "No"),
    ("Aunt Carol & Uncle Ray", 2, "Waiting", "", "No", "No"),
    ("Devon Brooks", 1, "No", "", "No", "No"),
    ("Hannah & Theo Ruiz", 2, "Yes", "Vegetarian", "Yes", "No"),
    ("Olivia Chen +1", 2, "Yes", "Beef", "No", "No"),
]
VENDORS = [  # (vendor, service, total, deposit, dep due, dep paid?, final due, final paid?)
    ("Willow Creek Venue", "Venue & catering", 9800, 2500, "2026-03-01", "Yes", "2026-09-15", "No"),
    ("J. Alvarez Photography", "Photo + video", 2900, 900, "2026-03-20", "Yes", "2026-09-01", "No"),
    ("Bloom & Stem", "Flowers", 2400, 600, "2026-04-10", "Yes", "2026-09-20", "No"),
    ("DJ Nightowl", "Music", 1700, 400, "2026-05-05", "Yes", "2026-10-01", "No"),
    ("Sweet Layer Bakery", "Cake", 495, 100, "2026-06-01", "Yes", "2026-09-25", "No"),
    ("Rev. Dana Whit", "Officiant", 450, 0, "", "No", "2026-10-03", "No"),
]

wb = Workbook()

# ---------- Tab 1: START HERE ----------
ws = wb.active; ws.title = "START HERE"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDE", (36, 30, 30, 30)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:E2"); ws["B2"] = "WEDDING BUDGET & GUEST TRACKER"
ws["B2"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
ws["B2"].fill = HEAD_FILL; ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 34
rows = [
    ("", ""),
    ("Welcome! Three tabs run the whole wedding: the money, the people, and the payments.", ""),
    ("", ""),
    ("HOW TO USE (3 steps)", "SEC"),
    ("1. 'Wedding Budget' tab: enter your total budget in the yellow cell. The 'Typical %' column", ""),
    ("    suggests a starting split from real weddings — then plan your own numbers in the yellow cells.", ""),
    ("2. 'Guest List' tab: add guests as you invite them. RSVPs, meal counts, headcount, and", ""),
    ("    thank-you notes are all counted for you at the top — nothing to add up, ever.", ""),
    ("3. 'Vendor Payments' tab: log every contract, deposit, and due date. The sheet shows what is", ""),
    ("    paid and exactly what is still owed, so no payment sneaks up on you.", ""),
    ("", ""),
    ("COLOR LEGEND", "SEC"),
    ("Yellow cells = yours to edit.  Everything else calculates automatically — please don't type over it.", ""),
    ("The numbers already in the sheet are a realistic worked example. Replace them with your own.", ""),
    ("", ""),
    ("GOOGLE SHEETS USERS", "SEC"),
    ("Upload this file to Google Drive, then open it — it converts automatically and all formulas work.", ""),
    ("", ""),
    ("Questions or a formula acting up? Message us on Etsy any time — we answer fast and we'll fix it.", ""),
]
r = 4
for text, kind in rows:
    cell = ws.cell(row=r, column=2, value=text)
    if kind == "SEC":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell.font = H2; cell.fill = SEC_FILL
        ws.row_dimensions[r].height = 20
    else:
        cell.font = B
    r += 1

# ---------- Tab 2: Wedding Budget ----------
ws = wb.create_sheet("Wedding Budget")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDEFG", (26, 11, 14, 14, 14, 14)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:G2"); ws["B2"] = "WEDDING BUDGET"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
ws["B4"] = "Total wedding budget:"; ws["B4"].font = BB
ws["C4"] = TOTAL_BUDGET; ws["C4"].fill = INPUT_FILL; ws["C4"].number_format = MONEY; ws["C4"].border = BORDER
ws["B5"] = "'Typical %' is guidance from real weddings, not a rule — every wedding is different. 'Typical $' updates from YOUR budget."
ws["B5"].font = SMALL
hdrs = ["Category", "Typical %", "Typical $", "Planned", "Actual", "Difference"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=7, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
CAT_FIRST, CAT_LAST = 8, 8 + len(CATEGORIES) - 1
for i, (name, pct, planned, actual) in enumerate(CATEGORIES):
    r = CAT_FIRST + i
    ws.cell(row=r, column=2, value=name).font = B
    pc = ws.cell(row=r, column=3, value=pct); pc.number_format = PCT
    tc = ws.cell(row=r, column=4, value=f"=$C$4*C{r}"); tc.number_format = MONEY
    pl = ws.cell(row=r, column=5, value=planned); pl.fill = INPUT_FILL; pl.number_format = MONEY
    ac = ws.cell(row=r, column=6, value=actual); ac.fill = INPUT_FILL; ac.number_format = MONEY
    df = ws.cell(row=r, column=7, value=f"=E{r}-F{r}"); df.number_format = MONEY
    for col in range(2, 8): ws.cell(row=r, column=col).border = BORDER
TOT = CAT_LAST + 1
ws.cell(row=TOT, column=2, value="Total").font = BB
ws.cell(row=TOT, column=3, value=f"=SUM(C{CAT_FIRST}:C{CAT_LAST})").number_format = PCT
ws.cell(row=TOT, column=4, value=f"=SUM(D{CAT_FIRST}:D{CAT_LAST})").number_format = MONEY
ws.cell(row=TOT, column=5, value=f"=SUM(E{CAT_FIRST}:E{CAT_LAST})").number_format = MONEY
ws.cell(row=TOT, column=6, value=f"=SUM(F{CAT_FIRST}:F{CAT_LAST})").number_format = MONEY
ws.cell(row=TOT, column=7, value=f"=E{TOT}-F{TOT}").number_format = MONEY
for col in range(2, 8):
    c = ws.cell(row=TOT, column=col); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
BL = TOT + 2
ws.merge_cells(start_row=BL, start_column=2, end_row=BL, end_column=7)
ws.cell(row=BL, column=2, value="THE BOTTOM LINE").font = H2
ws.cell(row=BL, column=2).fill = HEAD_FILL
for i, (lab, formula) in enumerate([
    ("Left to plan (budget − planned) — get this to $0", f"=C4-E{TOT}"),
    ("Budget remaining right now (budget − spent so far)", f"=C4-F{TOT}"),
]):
    rr = BL + 1 + i
    ws.cell(row=rr, column=2, value=lab).font = BB
    fc = ws.cell(row=rr, column=3, value=formula); fc.font = BB; fc.number_format = MONEY; fc.border = BORDER
ws.cell(row=BL + 3, column=2, value="Tip: keep the contingency line — the average wedding runs over in at least one category.").font = SMALL
ws.freeze_panes = "A8"

# ---------- Tab 3: Guest List ----------
ws = wb.create_sheet("Guest List")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDEFGH", (26, 10, 10, 13, 13, 14, 24)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:H2"); ws["B2"] = "GUEST LIST"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
G_FIRST, G_LAST = 13, 112  # 100 guest rows
stats = [
    ("Parties invited", f"=COUNTA(B{G_FIRST}:B{G_LAST})", "RSVP yes", f'=COUNTIF(D{G_FIRST}:D{G_LAST},"Yes")'),
    ("Headcount attending", f'=SUMIF(D{G_FIRST}:D{G_LAST},"Yes",C{G_FIRST}:C{G_LAST})', "RSVP no", f'=COUNTIF(D{G_FIRST}:D{G_LAST},"No")'),
    ("Chicken meals", f'=SUMIFS(C{G_FIRST}:C{G_LAST},D{G_FIRST}:D{G_LAST},"Yes",E{G_FIRST}:E{G_LAST},"Chicken")', "Awaiting reply", f'=COUNTIF(D{G_FIRST}:D{G_LAST},"Waiting")'),
    ("Beef meals", f'=SUMIFS(C{G_FIRST}:C{G_LAST},D{G_FIRST}:D{G_LAST},"Yes",E{G_FIRST}:E{G_LAST},"Beef")', "Gifts received", f'=COUNTIF(F{G_FIRST}:F{G_LAST},"Yes")'),
    ("Vegetarian meals", f'=SUMIFS(C{G_FIRST}:C{G_LAST},D{G_FIRST}:D{G_LAST},"Yes",E{G_FIRST}:E{G_LAST},"Vegetarian")', "Thank-yous still owed", f'=COUNTIF(F{G_FIRST}:F{G_LAST},"Yes")-COUNTIF(G{G_FIRST}:G{G_LAST},"Yes")'),
]
for i, (lab1, f1, lab2, f2) in enumerate(stats):
    r = 4 + i
    ws.cell(row=r, column=2, value=lab1).font = BB
    c1 = ws.cell(row=r, column=3, value=f1); c1.font = BB; c1.fill = ALT_FILL; c1.border = BORDER
    ws.cell(row=r, column=5, value=lab2).font = BB
    c2 = ws.cell(row=r, column=6, value=f2); c2.font = BB; c2.fill = ALT_FILL; c2.border = BORDER
ws.cell(row=10, column=2, value="Type RSVP exactly as Yes / No / Waiting and meals as Chicken / Beef / Vegetarian — the counters above match those words.").font = SMALL
hdrs = ["Guest / party", "# in party", "RSVP", "Meal choice", "Gift received", "Thank-you sent", "Notes"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=12, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for i in range(G_FIRST, G_LAST + 1):
    data = GUESTS[i - G_FIRST] if i - G_FIRST < len(GUESTS) else ("", None, "", "", "", "")
    vals = list(data) + [""]  # notes column
    for j, v in enumerate(vals):
        c = ws.cell(row=i, column=2 + j, value=v if v != "" else None)
        c.fill = INPUT_FILL; c.border = BORDER; c.font = B
ws.freeze_panes = "B13"

# ---------- Tab 4: Vendor Payments ----------
ws = wb.create_sheet("Vendor Payments")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDEFGHIJ", (22, 16, 13, 12, 12, 12, 13, 12, 11)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:J2"); ws["B2"] = "VENDOR PAYMENT SCHEDULE"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
ws["B4"] = "One row per vendor. Balance calculates itself. Mark Paid? = Yes as money goes out — the totals below always show what's still owed."
ws["B4"].font = SMALL
hdrs = ["Vendor", "Service", "Total contract", "Deposit", "Deposit due", "Deposit paid?", "Balance", "Final due", "Final paid?"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=6, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
V_FIRST, V_LAST = 7, 26  # 20 vendor rows
for i in range(V_FIRST, V_LAST + 1):
    data = VENDORS[i - V_FIRST] if i - V_FIRST < len(VENDORS) else ("", "", None, None, "", "", "", "")
    vendor, service, total, dep, depdue, deppaid, findue, finpaid = data
    ws.cell(row=i, column=2, value=vendor or None).font = B
    ws.cell(row=i, column=3, value=service or None).font = B
    tc = ws.cell(row=i, column=4, value=total); tc.number_format = MONEY
    dc = ws.cell(row=i, column=5, value=dep); dc.number_format = MONEY
    ws.cell(row=i, column=6, value=depdue or None).font = B
    ws.cell(row=i, column=7, value=deppaid or None).font = B
    bc = ws.cell(row=i, column=8, value=f"=D{i}-E{i}"); bc.number_format = MONEY
    ws.cell(row=i, column=9, value=findue or None).font = B
    ws.cell(row=i, column=10, value=finpaid or None).font = B
    for col in (2, 3, 4, 5, 6, 7, 9, 10):
        ws.cell(row=i, column=col).fill = INPUT_FILL
    for col in range(2, 11):
        ws.cell(row=i, column=col).border = BORDER
VT = V_LAST + 1
ws.cell(row=VT, column=2, value="Total").font = BB
ws.cell(row=VT, column=4, value=f"=SUM(D{V_FIRST}:D{V_LAST})").number_format = MONEY
ws.cell(row=VT, column=5, value=f"=SUM(E{V_FIRST}:E{V_LAST})").number_format = MONEY
ws.cell(row=VT, column=8, value=f"=SUM(H{V_FIRST}:H{V_LAST})").number_format = MONEY
for col in range(2, 11):
    c = ws.cell(row=VT, column=col); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
PAID_R, OWE_R = VT + 2, VT + 3
ws.cell(row=PAID_R, column=2, value="Paid so far").font = BB
pc = ws.cell(row=PAID_R, column=4, value=f'=SUMIF(G{V_FIRST}:G{V_LAST},"Yes",E{V_FIRST}:E{V_LAST})+SUMIF(J{V_FIRST}:J{V_LAST},"Yes",H{V_FIRST}:H{V_LAST})')
pc.font = BB; pc.number_format = MONEY; pc.border = BORDER
ws.cell(row=OWE_R, column=2, value="Still to pay").font = BB
oc = ws.cell(row=OWE_R, column=4, value=f"=D{VT}-D{PAID_R}")
oc.font = BB; oc.number_format = MONEY; oc.border = BORDER
ws.cell(row=OWE_R + 2, column=2, value="Type Yes in the Paid? columns exactly — the 'Paid so far' total counts that word.").font = SMALL
ws.freeze_panes = "B7"

XLSX = os.path.join(OUT, "Wedding-Budget-Guest-Tracker.xlsx")
wb.save(XLSX)
print("saved", XLSX)

# ================================================================== verification
def _cells(ws, ref):
    if ":" in ref:
        return [c for row in ws[ref] for c in row]
    return [ws[ref]]

def _val(ws, cell):
    """Numeric value of a cell, recursively evaluating formulas."""
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        return evaluate(ws, v)
    return v if isinstance(v, (int, float)) else 0

def _match(v, crit):
    return isinstance(v, str) and v.strip().lower() == crit.strip().lower()

def evaluate(ws, formula):
    """Evaluate the closed set of formulas this product uses: SUM, COUNTA,
    COUNTIF, SUMIF, SUMIFS + cell arithmetic (+ - *)."""
    expr = formula.lstrip("=").replace("$", "")
    def fn(m):
        name, args = m.group(1).upper(), [a.strip() for a in m.group(2).split(",")]
        if name == "SUM":
            return repr(sum(_val(ws, c) for c in _cells(ws, args[0])))
        if name == "COUNTA":
            return repr(sum(1 for c in _cells(ws, args[0]) if c.value not in (None, "")))
        if name == "COUNTIF":
            return repr(sum(1 for c in _cells(ws, args[0]) if _match(c.value, args[1].strip('"'))))
        if name == "SUMIF":
            crit_cells, sum_cells = _cells(ws, args[0]), _cells(ws, args[2])
            return repr(sum(_val(ws, s) for c, s in zip(crit_cells, sum_cells) if _match(c.value, args[1].strip('"'))))
        if name == "SUMIFS":
            sum_cells = _cells(ws, args[0])
            pairs = [( _cells(ws, args[i]), args[i + 1].strip('"')) for i in range(1, len(args), 2)]
            return repr(sum(_val(ws, s) for idx, s in enumerate(sum_cells)
                            if all(_match(rng[idx].value, crit) for rng, crit in pairs)))
        raise ValueError("unsupported function " + name)
    expr = re.sub(r"([A-Z]+)\(([^()]*)\)", fn, expr)
    expr = re.sub(r"\b([A-Z]{1,2}[0-9]{1,4})\b", lambda m: repr(_val(ws, ws[m.group(1)])), expr)
    return eval(expr)

wb2 = load_workbook(XLSX)
# expected values computed independently from the source data
pcts = [c[1] for c in CATEGORIES]; plans = [c[2] for c in CATEGORIES]; acts = [c[3] for c in CATEGORIES]
plan_total, act_total = sum(plans), round(sum(acts), 2)
expected = {}
wsb = wb2["Wedding Budget"]
for i, (name, pct, planned, actual) in enumerate(CATEGORIES):
    r = CAT_FIRST + i
    expected[("Wedding Budget", f"D{r}")] = TOTAL_BUDGET * pct
    expected[("Wedding Budget", f"G{r}")] = round(planned - actual, 2)
expected[("Wedding Budget", f"C{TOT}")] = round(sum(pcts), 10)
expected[("Wedding Budget", f"D{TOT}")] = TOTAL_BUDGET  # typical $ sums back to the full budget
expected[("Wedding Budget", f"E{TOT}")] = plan_total
expected[("Wedding Budget", f"F{TOT}")] = act_total
expected[("Wedding Budget", f"G{TOT}")] = round(plan_total - act_total, 2)
expected[("Wedding Budget", f"C{BL+1}")] = TOTAL_BUDGET - plan_total  # left to plan
expected[("Wedding Budget", f"C{BL+2}")] = round(TOTAL_BUDGET - act_total, 2)  # remaining

yes = [g for g in GUESTS if g[2] == "Yes"]
head = sum(g[1] for g in yes)
meals = {m: sum(g[1] for g in yes if g[3] == m) for m in ("Chicken", "Beef", "Vegetarian")}
gifts = sum(1 for g in GUESTS if g[4] == "Yes"); thanks = sum(1 for g in GUESTS if g[5] == "Yes")
expected[("Guest List", "C4")] = len(GUESTS)
expected[("Guest List", "C5")] = head
expected[("Guest List", "C6")] = meals["Chicken"]
expected[("Guest List", "C7")] = meals["Beef"]
expected[("Guest List", "C8")] = meals["Vegetarian"]
expected[("Guest List", "F4")] = len(yes)
expected[("Guest List", "F5")] = sum(1 for g in GUESTS if g[2] == "No")
expected[("Guest List", "F6")] = sum(1 for g in GUESTS if g[2] == "Waiting")
expected[("Guest List", "F7")] = gifts
expected[("Guest List", "F8")] = gifts - thanks

v_total = sum(v[2] for v in VENDORS); v_dep = sum(v[3] for v in VENDORS)
v_bal = v_total - v_dep
v_paid = sum(v[3] for v in VENDORS if v[5] == "Yes") + sum((v[2] - v[3]) for v in VENDORS if v[7] == "Yes")
for i, v in enumerate(VENDORS):
    expected[("Vendor Payments", f"H{V_FIRST+i}")] = v[2] - v[3]
for i in range(len(VENDORS), V_LAST - V_FIRST + 1):
    expected[("Vendor Payments", f"H{V_FIRST+i}")] = 0  # blank rows
expected[("Vendor Payments", f"D{VT}")] = v_total
expected[("Vendor Payments", f"E{VT}")] = v_dep
expected[("Vendor Payments", f"H{VT}")] = v_bal
expected[("Vendor Payments", f"D{PAID_R}")] = v_paid
expected[("Vendor Payments", f"D{OWE_R}")] = v_total - v_paid

checked = 0
for sheet in wb2.sheetnames:
    ws2 = wb2[sheet]
    for row in ws2.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                key = (sheet, cell.coordinate)
                assert key in expected, f"formula {key} has no expected value — add it"
                got = evaluate(ws2, cell.value)
                want = expected.pop(key)
                assert abs(got - want) < 1e-9, f"{key}: got {got}, expected {want}"
                checked += 1
assert not expected, f"expected values never matched a formula cell: {sorted(expected)}"

# hard pins — the worked example must read exactly like the listing says
assert plan_total == 24000 and TOTAL_BUDGET - plan_total == 0
assert round(sum(pcts), 10) == 1.0
assert act_total == 22061.75 and round(TOTAL_BUDGET - act_total, 2) == 1938.25
assert head == 16 and len(yes) == 8 and meals == {"Chicken": 9, "Beef": 4, "Vegetarian": 3}
assert gifts == 4 and gifts - thanks == 2
assert v_total == 17745 and v_dep == 4500 and v_bal == 13245 and v_paid == 4500
print(f"VERIFIED: all {checked} formulas match expected values")
print(f"  budget: planned ${plan_total:,} of ${TOTAL_BUDGET:,} -> left to plan $0; spent ${act_total:,.2f} -> remaining $1,938.25")
print(f"  guests: 12 parties, 8 yes / 2 no / 2 waiting, headcount 16, meals C9/B4/V3, thank-yous owed 2")
print(f"  vendors: contracts ${v_total:,}, deposits ${v_dep:,}, paid ${v_paid:,}, still to pay ${v_total - v_paid:,}")

# ================================================================== cover PNG (2700x2025)
W, H = 2700, 2025
NAVY_RGB = (31, 58, 95); TEAL_RGB = (46, 125, 107); GOLD_RGB = (242, 193, 78)
CREAM = (248, 246, 240); DARK = (40, 44, 52); LINE = (217, 212, 200); ALT = (240, 243, 238)
FB = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
FR = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
def F(p, s): return ImageFont.truetype(p, s)

img = Image.new("RGB", (W, H), CREAM); d = ImageDraw.Draw(img)
d.rectangle([0, 0, W, 300], fill=NAVY_RGB)
size = 128
while d.textlength("WEDDING BUDGET & GUEST TRACKER", font=F(FB, size)) > W - 260: size -= 4
d.text((W / 2, 150), "WEDDING BUDGET & GUEST TRACKER", font=F(FB, size), fill=(255, 255, 255), anchor="mm")
d.text((W / 2, 460), "The budget, the guest list, and every vendor payment —", font=F(FR, 66), fill=DARK, anchor="mm")
d.text((W / 2, 570), "one organized spreadsheet.", font=F(FB, 66), fill=DARK, anchor="mm")
cx0, cy0, cx1, cy1 = 450, 700, 2250, 1600
d.rounded_rectangle([cx0, cy0, cx1, cy1], radius=24, fill=(255, 255, 255), outline=LINE, width=3)
d.rounded_rectangle([cx0, cy0, cx1, cy0 + 110], radius=24, fill=TEAL_RGB)
d.rectangle([cx0, cy0 + 60, cx1, cy0 + 110], fill=TEAL_RGB)
d.text((cx0 + 60, cy0 + 55), "WEDDING BUDGET", font=F(FB, 56), fill=(255, 255, 255), anchor="lm")
rows = [("Venue & catering", "$9,600"), ("Photography & video", "$2,900"), ("Guests attending", "16"), ("LEFT TO PLAN", "$0")]
y = cy0 + 220
for i, (lab, val) in enumerate(rows):
    last = i == len(rows) - 1
    if last:
        d.rectangle([cx0 + 6, y - 60, cx1 - 6, y + 60], fill=ALT)
    d.text((cx0 + 70, y), lab, font=F(FB if last else FR, 60), fill=NAVY_RGB if last else DARK, anchor="lm")
    d.text((cx1 - 70, y), val, font=F(FB, 60), fill=TEAL_RGB if last else NAVY_RGB, anchor="rm")
    y += 165
d.rectangle([0, H - 190, W, H], fill=GOLD_RGB)
d.text((W / 2, H - 95), "Google Sheets + Excel  •  Instant Download  •  Budget + Guests + Vendors", font=F(FB, 60), fill=DARK, anchor="mm")
img.save(os.path.join(OUT, "01-cover.png"), optimize=True)
print("saved", os.path.join(OUT, "01-cover.png"))

# ================================================================== store SVG (460x345)
SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 460 345" font-family="Helvetica,Arial,sans-serif">
<rect width="460" height="345" fill="#f8f6f0"/>
<rect width="460" height="62" fill="#1f3a5f"/>
<text x="230" y="40" text-anchor="middle" font-size="22" font-weight="bold" fill="#ffffff">WEDDING BUDGET &amp; GUESTS</text>
<text x="230" y="92" text-anchor="middle" font-size="15" fill="#282c34">The budget, the guest list, and every</text>
<text x="230" y="112" text-anchor="middle" font-size="15" font-weight="bold" fill="#282c34">vendor payment — one spreadsheet.</text>
<rect x="100" y="132" width="290" height="164" rx="8" fill="#ffffff" stroke="#d9d4c8"/>
<rect x="100" y="132" width="290" height="30" fill="#2e7d6b" rx="8"/>
<rect x="100" y="152" width="290" height="12" fill="#2e7d6b"/>
<text x="115" y="152" font-size="13" font-weight="bold" fill="#ffffff">WEDDING BUDGET</text>
<text x="120" y="208" font-size="13" fill="#282c34">Venue &amp; catering</text><text x="360" y="208" font-size="13" font-weight="bold" fill="#1f3a5f">$9,600</text><text x="120" y="232" font-size="13" fill="#282c34">Photography &amp; video</text><text x="360" y="232" font-size="13" font-weight="bold" fill="#1f3a5f">$2,900</text><text x="120" y="256" font-size="13" fill="#282c34">Guests attending</text><text x="360" y="256" font-size="13" font-weight="bold" fill="#1f3a5f">16</text><text x="120" y="280" font-size="13" fill="#282c34">LEFT TO PLAN</text><text x="360" y="280" font-size="13" font-weight="bold" fill="#2e7d6b">$0</text>
<rect y="305" width="460" height="40" fill="#f2c14e"/>
<text x="230" y="330" text-anchor="middle" font-size="13" font-weight="bold" fill="#282c34">Google Sheets + Excel  •  Budget + Guests + Vendors</text>
</svg>"""
with open(os.path.join(OUT, "store-card.svg"), "w") as f:
    f.write(SVG)
os.makedirs(STORE_IMG, exist_ok=True)
shutil.copy(os.path.join(OUT, "store-card.svg"), os.path.join(STORE_IMG, "wedding-budget.svg"))
print("saved store-card.svg (+ copied to store/site/img/wedding-budget.svg)")
