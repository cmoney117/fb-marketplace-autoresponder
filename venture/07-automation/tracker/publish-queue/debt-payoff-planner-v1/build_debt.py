#!/usr/bin/env python3
"""Build the Debt Payoff Planner v2 spreadsheet product.

v2 replaces the v1 file (whose exact reproduction was verified cell-by-cell
before this rewrite — see listing.md QA note). What changed and why:

* REAL PAYOFF ENGINE — two "Schedule" tabs run a month-by-month amortization
  (up to 360 months) for all 10 debts under each strategy: interest compounds
  monthly (APR/12), every debt gets its minimum, and the whole extra payment
  PLUS every freed-up minimum rolls onto the top unpaid debt (true
  snowball/avalanche rollover). Formulas only — no macros.
* LINKED STRATEGY TABS — Snowball and Avalanche read My Debts and re-order
  themselves with SMALL/MATCH/INDEX. No manual re-entry, all 10 rows.
* HONEST DEBT-FREE DATE — projected debt-free MONTH + YEAR per strategy
  (EDATE off TODAY), plus total interest, months saved and interest saved by
  the extra payment (vs a minimums-only run of the same engine).
* CHARTS — native Excel charts on the Progress tab: total-balance decline
  line (Snowball vs Avalanche, first 120 months) + per-debt balance bars.

Companion verify_debt.py re-derives every formula string, re-runs the engine
in Python (plus an independent Decimal-based second simulator), and compares
against the LibreOffice-recalculated cache. Must pass before publish.

Excel-2010+/Google-Sheets-safe functions only: SUM SUMPRODUCT MIN MAX IF OR
IFERROR MATCH INDEX SMALL ROUND EDATE TODAY TEXT N T COLUMN.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

OUT = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/debt-payoff-planner-v1"
os.makedirs(OUT, exist_ok=True)

NAVY = "1F3A5F"; TEAL = "2E7D6B"; GOLD = "F2C14E"; LIGHT = "F4F6F8"; YELLOW = "FFF2CC"
H1 = Font(name="Arial", size=18, bold=True, color="FFFFFF")
H2 = Font(name="Arial", size=12, bold=True, color="FFFFFF")
B = Font(name="Arial", size=11)
BB = Font(name="Arial", size=11, bold=True)
BIG = Font(name="Arial", size=14, bold=True, color=NAVY)
SMALL = Font(name="Arial", size=10, italic=True, color="666666")
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
GOLD_FILL = PatternFill("solid", fgColor=GOLD)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00;($#,##0.00);"-"'   # dash for zero (tables)
MONEY0 = '$#,##0.00;($#,##0.00)'      # $0.00 for zero (hero/summary cells)
PLAIN2 = '#,##0.00'
PCT = "0.0%"
PCTM = "0.000%"

DEBTS = [  # (name, balance, apr, min payment) — worked example, replace with your own
    ("Store card", 640, 0.279, 35),
    ("Medical bill", 1200, 0.0, 50),
    ("Credit card A", 2850, 0.246, 75),
    ("Credit card B", 4100, 0.219, 95),
    ("Car loan", 9400, 0.069, 310),
]
EXTRA = 150

# ---- engine layout constants (verify_debt.py re-derives from these numbers) ----
N_DEBTS = 10
N_MONTHS = 360
GRID0 = 14                 # month-0 row; months 1..360 live in GRID0+1 .. GRID0+N_MONTHS
GRID_LAST = GRID0 + N_MONTHS
BAL0_COL = 2               # B..K   plan balances
TOT_COL = 12               # L      plan total balance
INT_COL = 13               # M      plan interest this month
MP0_COL = 14               # N..W   plan min-due helper
PAY0_COL = 24              # X..AG  plan payment
BBAL_COL = 35              # AI..AR minimums-only balances
BTOT_COL = 45              # AS     minimums-only total
BINT_COL = 46              # AT     minimums-only interest
BMP_COL = 47               # AU..BD minimums-only min-due helper
BPAY_COL = 57              # BE..BN minimums-only payment
CHART_MONTHS = 120         # decline chart plots months 0..120

MD = "'My Debts'"

def L(col):
    return get_column_letter(col)

wb = Workbook()

def banner(ws, last_col, text):
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    c = ws.cell(row=2, column=2, value=text)
    c.font = H1; c.fill = HEAD_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 30
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

# ================================================================ START HERE
ws = wb.active; ws.title = "START HERE"
for c, w in zip("BCDE", (34, 30, 30, 30)): ws.column_dimensions[c].width = w
banner(ws, 5, "DEBT PAYOFF PLANNER")
rows = [
    (4, "Snowball or Avalanche — see your projected debt-free date either way.", None),
    (6, "HOW TO USE", "SEC"),
    (7, "1. 'My Debts' tab: list every debt in the yellow cells (up to 10) and your extra payment at the top.", None),
    (8, "2. The Snowball and Avalanche tabs re-order your debts automatically and run the month-by-month", None),
    (9, "    math: your projected debt-free month and year, per debt and overall, plus total interest.", None),
    (10, "3. 'The Verdict' on My Debts puts both plans side by side. Pick the one you'll actually stick to.", None),
    (11, "4. Log every win on the Progress tab — the charts there update as you pay things down.", None),
    (13, "COLOR LEGEND", "SEC"),
    (14, "Yellow cells = yours to edit. Everything else calculates automatically — please don't type over it.", None),
    (15, "The strategy tabs and the two Schedule tabs have NO yellow: they are 100% calculated for you.", None),
    (16, "The pre-filled numbers are a realistic example — replace them with your own.", None),
    (18, "HOW THE MATH WORKS (the honest fine print)", "SEC"),
    (19, "• Interest compounds monthly at APR ÷ 12. Every month each debt gets its minimum payment;", None),
    (20, "   your whole extra payment — plus every freed-up minimum — attacks the top unpaid debt.", None),
    (21, "• The Schedule tabs show all 360 months of that math. Look under the hood any time.", None),
    (22, "• Dates project from today and assume steady payments; real due dates shift things a few weeks.", None),
    (23, "• '30+ years' means the payment doesn't cover the interest — raise it. This is a calculator,", None),
    (24, "   not financial advice.", None),
    (26, "CHARTS & COMPATIBILITY", "SEC"),
    (27, "Native Excel charts (Progress tab): full support in Excel 2010+. Google Sheets converts them on", None),
    (28, "upload — colors can shift slightly there, the data is identical. Formulas work everywhere; no macros.", None),
    (30, "Snowball = smallest balance first (fastest wins, best motivation).", None),
    (31, "Avalanche = highest interest rate first (mathematically cheapest).", None),
    (33, "Questions? Message us — real humans, fast replies, we'll fix anything.", None),
]
rowmap = {r: (t, k) for r, t, k in rows}
for r in range(4, 34):
    text, kind = rowmap.get(r, ("", None))
    cell = ws.cell(row=r, column=2, value=text or None)
    if kind == "SEC":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell.font = H2; cell.fill = SEC_FILL
        ws.row_dimensions[r].height = 20
    else:
        cell.font = B

# ================================================================ My Debts
ws = wb.create_sheet("My Debts")
for c, w in zip("BCDEF", (24, 14, 10, 14, 22)): ws.column_dimensions[c].width = w
banner(ws, 6, "MY DEBTS")
ws["B4"] = "Monthly EXTRA payment (beyond minimums):"; ws["B4"].font = BB
ws["E4"] = EXTRA; ws["E4"].fill = INPUT_FILL; ws["E4"].number_format = MONEY; ws["E4"].border = BORDER
for i, h in enumerate(["Debt name", "Balance", "APR", "Min payment", "Notes"]):
    c = ws.cell(row=6, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for r in range(7, 17):
    data = DEBTS[r - 7] if r - 7 < len(DEBTS) else (None, None, None, None)
    name, bal, apr, mn = data
    ws.cell(row=r, column=2, value=name)
    bc = ws.cell(row=r, column=3, value=bal); bc.number_format = MONEY
    pc = ws.cell(row=r, column=4, value=apr); pc.number_format = PCT
    mc = ws.cell(row=r, column=5, value=mn); mc.number_format = MONEY
    for col in range(2, 7):
        cc = ws.cell(row=r, column=col); cc.fill = INPUT_FILL; cc.border = BORDER
ws.cell(row=17, column=2, value="Total").font = BB
c = ws.cell(row=17, column=3, value="=SUM(C7:C16)"); c.font = BB; c.number_format = MONEY0
c = ws.cell(row=17, column=5, value="=SUM(E7:E16)"); c.font = BB; c.number_format = MONEY
for col in range(2, 7):
    cc = ws.cell(row=17, column=col); cc.fill = ALT_FILL; cc.border = BORDER
ws.cell(row=19, column=2, value="Total monthly attack (minimums + extra):").font = BB
c = ws.cell(row=19, column=5, value="=E17+E4"); c.font = BB; c.number_format = MONEY

# ---- THE VERDICT (linked to both engines) ----
ws.merge_cells("B21:F21")
c = ws["B21"]; c.value = "THE VERDICT — PICK YOUR PLAN"; c.font = H2; c.fill = SEC_FILL
ws.row_dimensions[21].height = 20
for i, h in enumerate(["Strategy", "Debt-free date", "Months", "Total interest"]):
    c = ws.cell(row=22, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for r, (label, sched) in ((23, ("Snowball", "'Snowball Schedule'")), (24, ("Avalanche", "'Avalanche Schedule'"))):
    ws.cell(row=r, column=2, value=label).font = BB
    dc = ws.cell(row=r, column=3, value=(f'=IF({sched}!$L$14<=0,"—",IF({sched}!$O$2=999,"30+ years",'
                                         f'TEXT(EDATE(TODAY(),{sched}!$O$2),"MMM YYYY")))'))
    mc = ws.cell(row=r, column=4, value=f'=IF({sched}!$L$14<=0,"—",IF({sched}!$O$2=999,"360+",{sched}!$O$2))')
    mc.number_format = "0"
    ic = ws.cell(row=r, column=5, value=f'=IF({sched}!$L$14<=0,"—",ROUND({sched}!$O$3,2))')
    ic.number_format = MONEY0
    for col in range(2, 6):
        cc = ws.cell(row=r, column=col); cc.border = BORDER
        if col > 2: cc.font = BB
ws.cell(row=25, column=2, value=("Same monthly attack either way. Snowball = faster first wins; Avalanche usually pays "
                                 "less interest. The Schedule tabs show every month of the math.")).font = SMALL

# ================================================================ Schedule engine sheets
def schedule_sheet(title, strategy):
    """strategy: 'snowball' (smallest balance first) or 'avalanche' (highest APR first)."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for col in range(2, 67):
        ws.column_dimensions[L(col)].width = 11
    c = ws["A1"]
    c.value = (f"{('SNOWBALL' if strategy == 'snowball' else 'AVALANCHE')} ENGINE — month-by-month payoff math. "
               "Nothing to edit here; it recalculates from My Debts. Column banks: balances, then min-due and "
               "payment helpers, then the same engine run again with minimums only (for the 'saved' lines).")
    c.font = SMALL
    hdr_labels = {
        2: "Slot (attack order)", 3: "Sort key (internal)", 4: "Feeds from debt #",
        5: "Start balance", 6: "Debt", 7: "APR", 8: "Monthly rate", 9: "Min payment",
        10: "Paid off in month #", 11: "Monthly budget (minimums + extra)", 12: "Budget if minimums only",
    }
    for r, lab in hdr_labels.items():
        ws.cell(row=r, column=1, value=lab).font = BB
    for j in range(N_DEBTS):
        col = BAL0_COL + j
        cl = L(col)
        md_row = 7 + j  # the My Debts row this KEY (not slot) describes
        ws.cell(row=2, column=col, value=j + 1).font = BB
        if strategy == "snowball":
            key = f"=IF(N({MD}!C{md_row})<=0,9000000000+COLUMN(),N({MD}!C{md_row})+COLUMN()/1000)"
        else:
            key = f"=IF(N({MD}!C{md_row})<=0,9000000000+COLUMN(),10-N({MD}!D{md_row})+COLUMN()/100000)"
        ws.cell(row=3, column=col, value=key)
        ws.cell(row=4, column=col, value=f"=MATCH(SMALL($B$3:$K$3,{cl}$2),$B$3:$K$3,0)")
        c = ws.cell(row=5, column=col, value=f"=ROUND(MAX(0,N(INDEX({MD}!$C$7:$C$16,{cl}$4))),2)")
        c.number_format = PLAIN2
        ws.cell(row=6, column=col, value=f'=IF({cl}$5<=0,"",T(INDEX({MD}!$B$7:$B$16,{cl}$4)))')
        c = ws.cell(row=7, column=col, value=f"=IF({cl}$5<=0,0,MAX(0,N(INDEX({MD}!$D$7:$D$16,{cl}$4))))")
        c.number_format = PCT
        c = ws.cell(row=8, column=col, value=f"={cl}$7/12"); c.number_format = PCTM
        c = ws.cell(row=9, column=col, value=f"=IF({cl}$5<=0,0,MAX(0,N(INDEX({MD}!$E$7:$E$16,{cl}$4))))")
        c.number_format = PLAIN2
        ws.cell(row=10, column=col, value=f'=IF({cl}$5<=0,"",IFERROR(MATCH(0,{cl}${GRID0 + 1}:{cl}${GRID_LAST},0),999))')
    c = ws.cell(row=11, column=2, value=f"=MAX(0,N({MD}!$E$4))+SUM($B$9:$K$9)"); c.font = BB; c.number_format = PLAIN2
    c = ws.cell(row=12, column=2, value="=SUM($B$9:$K$9)"); c.font = BB; c.number_format = PLAIN2
    # summary block
    for r, lab, f in (
        (2, "Months to debt-free (this plan)", f"=IF($L${GRID0}<=0,0,IFERROR(MATCH(0,$L${GRID0 + 1}:$L${GRID_LAST},0),999))"),
        (3, "Total interest (this plan)", f"=SUM($M${GRID0 + 1}:$M${GRID_LAST})"),
        (4, "Months to debt-free (minimums only)", f"=IF($L${GRID0}<=0,0,IFERROR(MATCH(0,$AS${GRID0 + 1}:$AS${GRID_LAST},0),999))"),
        (5, "Total interest (minimums only)", f"=SUM($AT${GRID0 + 1}:$AT${GRID_LAST})"),
    ):
        ws.cell(row=r, column=14, value=lab).font = BB
        c = ws.cell(row=r, column=15, value=f); c.font = BB; c.fill = ALT_FILL
        c.number_format = PLAIN2
    # grid headers
    hdr = {1: "Month", TOT_COL: "Total balance", INT_COL: "Interest", BTOT_COL: "(min-only) Total", BINT_COL: "(min-only) Interest"}
    for j in range(N_DEBTS):
        hdr[BAL0_COL + j] = f"Bal {j + 1}"
        hdr[MP0_COL + j] = f"Min due {j + 1}"
        hdr[PAY0_COL + j] = f"Pay {j + 1}"
        hdr[BBAL_COL + j] = f"(min-only) Bal {j + 1}"
        hdr[BMP_COL + j] = f"(min-only) Min due {j + 1}"
        hdr[BPAY_COL + j] = f"(min-only) Pay {j + 1}"
    for col, lab in hdr.items():
        c = ws.cell(row=13, column=col, value=lab); c.font = BB; c.fill = ALT_FILL
    # month 0
    ws.cell(row=GRID0, column=1, value=0).font = B
    for j in range(N_DEBTS):
        c = ws.cell(row=GRID0, column=BAL0_COL + j, value=f"={L(BAL0_COL + j)}$5"); c.number_format = PLAIN2
        c = ws.cell(row=GRID0, column=BBAL_COL + j, value=f"={L(BAL0_COL + j)}$5"); c.number_format = PLAIN2
    c = ws.cell(row=GRID0, column=TOT_COL, value=f"=SUM(B{GRID0}:K{GRID0})"); c.number_format = PLAIN2; c.font = BB
    c = ws.cell(row=GRID0, column=BTOT_COL, value=f"=SUM(AI{GRID0}:AR{GRID0})"); c.number_format = PLAIN2
    # months 1..N_MONTHS: the waterfall. For each slot j (attack order):
    #   min-due  MP_j  = MIN(prev*(1+rate), min)
    #   payment PAY_j = MIN(prev*(1+rate), MAX(MP_j, budget - payments to earlier slots - min-dues of later slots))
    #   balance BAL_j = ROUND(MAX(0, prev*(1+rate) - PAY_j), 2)
    # Every active debt gets at least its minimum; the whole surplus (extra + freed minimums)
    # hits the top unpaid debt and spills to the next when it finishes mid-month.
    for m in range(1, N_MONTHS + 1):
        r = GRID0 + m
        p = r - 1
        ws.cell(row=r, column=1, value=m)
        for base_bal, base_mp, base_pay, budget in ((BAL0_COL, MP0_COL, PAY0_COL, "$B$11"),
                                                    (BBAL_COL, BMP_COL, BPAY_COL, "$B$12")):
            for j in range(N_DEBTS):
                bal = L(base_bal + j); mp = L(base_mp + j); pay = L(base_pay + j)
                rate = L(BAL0_COL + j)  # rate/min headers live over the plan balance bank
                owed = f"{bal}{p}*(1+{rate}$8)"
                c = ws.cell(row=r, column=base_mp + j, value=f"=MIN({owed},{rate}$9)")
                c.number_format = PLAIN2
                terms = [budget]
                if j > 0:
                    terms.append(f"-SUM({L(base_pay)}{r}:{L(base_pay + j - 1)}{r})")
                if j < N_DEBTS - 1:
                    terms.append(f"-SUM({L(base_mp + j + 1)}{r}:{L(base_mp + N_DEBTS - 1)}{r})")
                c = ws.cell(row=r, column=base_pay + j,
                            value=f"=MIN({owed},MAX({mp}{r},{''.join(terms)}))")
                c.number_format = PLAIN2
                c = ws.cell(row=r, column=base_bal + j,
                            value=f"=ROUND(MAX(0,{owed}-{pay}{r}),2)")
                c.number_format = PLAIN2
        c = ws.cell(row=r, column=TOT_COL, value=f"=SUM(B{r}:K{r})"); c.number_format = PLAIN2; c.font = BB
        c = ws.cell(row=r, column=INT_COL, value=f"=SUMPRODUCT(B{p}:K{p},$B$8:$K$8)"); c.number_format = PLAIN2
        c = ws.cell(row=r, column=BTOT_COL, value=f"=SUM(AI{r}:AR{r})"); c.number_format = PLAIN2
        c = ws.cell(row=r, column=BINT_COL, value=f"=SUMPRODUCT(AI{p}:AR{p},$B$8:$K$8)"); c.number_format = PLAIN2
    ws.freeze_panes = f"B{GRID0 + 1}"
    # keep Ctrl+P sane: print only the summary + first 5 years of the engine
    ws.print_area = f"A1:O{GRID0 + 60}"
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    return ws

# ================================================================ Strategy tabs (buyer-facing)
def strategy_tab(title, sched, blurb):
    ws = wb.create_sheet(title)
    for c, w in zip("BCDEFG", (24, 14, 10, 14, 16, 16)): ws.column_dimensions[c].width = w
    banner(ws, 7, title.upper() + " PLAN")
    ws["B4"] = blurb; ws["B4"].font = SMALL
    S = f"'{sched}'"
    stats = [
        (6, "PROJECTED DEBT-FREE DATE", (f'=IF({S}!$L$14<=0,"add your debts first",IF({S}!$O$2=999,'
                                         f'"30+ years — raise your payment",TEXT(EDATE(TODAY(),{S}!$O$2),"MMMM YYYY")))'), None),
        (7, "Months from today", f'=IF({S}!$L$14<=0,"—",IF({S}!$O$2=999,"360+",{S}!$O$2))', "0"),
        (8, "Total interest on this plan", f'=IF({S}!$L$14<=0,"—",ROUND({S}!$O$3,2))', MONEY0),
        (9, "Months saved by your extra payment", f'=IF({S}!$L$14<=0,"—",IF(OR({S}!$O$2=999,{S}!$O$4=999),"—",{S}!$O$4-{S}!$O$2))', "0"),
        (10, "Interest saved by your extra payment", f'=IF({S}!$L$14<=0,"—",IF(OR({S}!$O$2=999,{S}!$O$4=999),"—",ROUND({S}!$O$5-{S}!$O$3,2)))', MONEY0),
    ]
    for r, lab, f, fmt in stats:
        ws.cell(row=r, column=2, value=lab).font = BB
        c = ws.cell(row=r, column=4, value=f)
        c.border = BORDER
        if fmt: c.number_format = fmt
        if r == 6:
            ws.merge_cells(start_row=6, start_column=4, end_row=6, end_column=6)
            c.font = BIG; c.fill = GOLD_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[6].height = 24
            ws.cell(row=6, column=2).font = Font(name="Arial", size=12, bold=True)
        else:
            c.font = BB; c.fill = ALT_FILL
    for i, h in enumerate(["Debt (attack order)", "Balance", "APR", "Min payment", "Paid off (month #)", "Debt-free date"]):
        c = ws.cell(row=12, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
        c.alignment = Alignment(wrap_text=True)
    for i in range(1, N_DEBTS + 1):
        r = 12 + i
        ws.cell(row=r, column=2, value=f"=INDEX({S}!$B$6:$K$6,{i})").font = B
        c = ws.cell(row=r, column=3, value=f'=IF(INDEX({S}!$B$5:$K$5,{i})<=0,"",INDEX({S}!$B$5:$K$5,{i}))')
        c.number_format = MONEY
        c = ws.cell(row=r, column=4, value=f'=IF(C{r}="","",INDEX({S}!$B$7:$K$7,{i}))'); c.number_format = PCT
        c = ws.cell(row=r, column=5, value=f'=IF(C{r}="","",INDEX({S}!$B$9:$K$9,{i}))'); c.number_format = MONEY
        ws.cell(row=r, column=6, value=(f'=IF(C{r}="","",IF(INDEX({S}!$B$10:$K$10,{i})=999,"360+",'
                                        f'INDEX({S}!$B$10:$K$10,{i})))'))
        ws.cell(row=r, column=7, value=(f'=IF(C{r}="","",IF(INDEX({S}!$B$10:$K$10,{i})=999,"—",'
                                        f'TEXT(EDATE(TODAY(),INDEX({S}!$B$10:$K$10,{i})),"MMM YYYY")))'))
        for col in range(2, 8):
            ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=24, column=2, value=("How the math works: interest compounds monthly (APR ÷ 12); every debt gets its minimum; "
                                     "your whole extra payment plus every freed-up minimum attacks the top unpaid debt.")).font = SMALL
    ws.cell(row=25, column=2, value=("Dates project from today. Nothing on this tab is typed by you — it all follows "
                                     "My Debts. Full month-by-month math: see the " + sched + " tab.")).font = SMALL
    return ws

strategy_tab("Snowball", "Snowball Schedule",
             "Smallest balance first for quick wins. This tab reads My Debts, re-orders itself, and runs the month-by-month math — nothing to type here.")
strategy_tab("Avalanche", "Avalanche Schedule",
             "Highest interest rate first — the mathematically cheapest path. Reads My Debts and re-orders itself; nothing to type here.")

# ================================================================ Progress (+ charts)
ws = wb.create_sheet("Progress")
for c, w in zip("BCDE", (24, 16, 16, 30)): ws.column_dimensions[c].width = w
banner(ws, 5, "PROGRESS — DEBTS DESTROYED")
for i, h in enumerate(["Debt", "Paid off (date)", "Amount killed", "How it felt"]):
    c = ws.cell(row=4, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for r in range(5, 15):
    for col in range(2, 6):
        c = ws.cell(row=r, column=col); c.fill = INPUT_FILL; c.border = BORDER
        if col == 4: c.number_format = MONEY
ws.cell(row=16, column=2, value="Total destroyed").font = BB
c = ws.cell(row=16, column=4, value="=SUM(D5:D14)"); c.font = BB; c.number_format = MONEY0
for col in range(2, 6):
    cc = ws.cell(row=16, column=col); cc.fill = ALT_FILL; cc.border = BORDER
ws.cell(row=18, column=2, value="Charts update from My Debts and the Schedule engines — nothing to type.").font = SMALL

# engine sheets must exist before charts reference them
snow = schedule_sheet("Snowball Schedule", "snowball")
aval = schedule_sheet("Avalanche Schedule", "avalanche")

line = LineChart()
line.title = "Total balance, month by month (first 10 years)"
line.style = None
for sheet, tcol, color, name in ((snow, TOT_COL, NAVY, "Snowball"),
                                 (aval, TOT_COL, TEAL, "Avalanche"),
                                 (snow, BTOT_COL, GOLD, "Minimums only")):
    ref = Reference(sheet, min_col=tcol, min_row=GRID0, max_row=GRID0 + CHART_MONTHS)
    line.add_data(ref, titles_from_data=False)
    s = line.series[-1]
    s.tx = SeriesLabel(v=name)
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width = 28000
    s.smooth = False
cats = Reference(snow, min_col=1, min_row=GRID0, max_row=GRID0 + CHART_MONTHS)
line.set_categories(cats)
line.x_axis.title = "Months from today"
line.y_axis.numFmt = "$#,##0"
line.x_axis.tickLblSkip = 12
line.width = 22; line.height = 11
ws.add_chart(line, "G4")

bar = BarChart(); bar.type = "col"
bar.title = "Your debts today"
data = Reference(wb["My Debts"], min_col=3, min_row=7, max_row=16)
bar.add_data(data, titles_from_data=False)
bar.series[0].tx = SeriesLabel(v="Balance")
bar.series[0].graphicalProperties.solidFill = GOLD
bar.series[0].graphicalProperties.line.solidFill = NAVY
bar.set_categories(Reference(wb["My Debts"], min_col=2, min_row=7, max_row=16))
bar.y_axis.numFmt = "$#,##0"
bar.legend = None
bar.width = 22; bar.height = 10
ws.add_chart(bar, "G27")

XLSX = os.path.join(OUT, "Debt-Payoff-Planner.xlsx")
wb.save(XLSX)
print("saved", XLSX)
print("next: run the xlsx skill's recalc.py on it, then verify_debt.py")
