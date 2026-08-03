#!/usr/bin/env python3
"""Verify Debt-Payoff-Planner.xlsx (v2) — must pass before publish.

Three independent proofs that the payoff engine is right:

1. FORMULA TEXT — re-derives the exact formula string of every formula cell
   (all ~46k of them, both Schedule engines included) from the same layout
   constants as build_debt.py and asserts total two-way coverage.
2. PYTHON ENGINE vs LIBREOFFICE CACHE — re-implements the amortization
   waterfall in Python (float, same operation order) and compares EVERY grid
   cell of both strategies, plan AND minimums-only baseline: balances,
   min-dues, payments, totals, interest, payoff months, summary lines.
   The build workflow runs the xlsx skill's recalc.py first so every formula
   has a LibreOffice-computed cached value.
3. SECOND OPINION SIMULATOR — an independently coded allocator ("pay every
   minimum, then pour the surplus down the attack order") must agree with the
   engine on every month's balances, every payoff month, and total interest.
   Plus model invariants: minimums always honored, budget never exceeded,
   surplus only ever flows to a prefix of the attack order, and money
   conservation (payments = principal + interest, within rounding).

Also asserts: no macros, Excel-2010+/Sheets-safe functions only, the
yellow-cell invariant, chart presence/series/colors on Progress, and the
worked example's pinned headline numbers.

TODAY()-dependent cells (debt-free dates) are checked against EDATE(today, n)
allowing a one-day-old cache; mismatches there are warnings, not failures,
because Excel/Sheets recalculate on open. Exits non-zero on any failure.
"""
import math
import re
import sys
from calendar import monthrange
from datetime import date, timedelta

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart
from openpyxl.utils import get_column_letter

XLSX = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/debt-payoff-planner-v1/Debt-Payoff-Planner.xlsx"

# ---- same layout constants as build_debt.py ----
N_DEBTS = 10
N_MONTHS = 360
GRID0 = 14
GRID_LAST = GRID0 + N_MONTHS
BAL0_COL, TOT_COL, INT_COL, MP0_COL, PAY0_COL = 2, 12, 13, 14, 24
BBAL_COL, BTOT_COL, BINT_COL, BMP_COL, BPAY_COL = 35, 45, 46, 47, 57
CHART_MONTHS = 120
MD = "'My Debts'"
NAVY, TEAL, GOLD = "1F3A5F", "2E7D6B", "F2C14E"

ALLOWED_FUNCS = {"SUM", "SUMPRODUCT", "MIN", "MAX", "IF", "OR", "IFERROR",
                 "MATCH", "INDEX", "SMALL", "ROUND", "EDATE", "TODAY", "TEXT",
                 "N", "T", "COLUMN"}

fails, warns = [], []
L = get_column_letter


def xround(x, n=2):
    """Excel/LibreOffice ROUND: half away from zero."""
    p = 10 ** n
    return math.copysign(math.floor(abs(x) * p + 0.5) / p, x)


def edate(d, months):
    y, m = divmod(d.month - 1 + months, 12)
    y += d.year
    m += 1
    return date(y, m, min(d.day, monthrange(y, m)[1]))


wf = load_workbook(XLSX, data_only=False)
wv = load_workbook(XLSX, data_only=True)
TODAY = date.today()

assert wf.sheetnames == ["START HERE", "My Debts", "Snowball", "Avalanche", "Progress",
                         "Snowball Schedule", "Avalanche Schedule"], wf.sheetnames
assert wf.vba_archive is None and XLSX.endswith(".xlsx"), "workbook must be macro-free"
HAS_CACHE = wv["Snowball Schedule"]["B5"].value is not None
assert HAS_CACHE, "no cached values — run the xlsx skill's recalc.py first"

# ------------------------------------------------- read worked example back out
md = wf["My Debts"]
EXTRA = float(md["E4"].value or 0)
DEBTS = []
for r in range(7, 17):
    name, bal, apr, mn = (md.cell(r, c).value for c in (2, 3, 4, 5))
    DEBTS.append((name, float(bal or 0), float(apr or 0), float(mn or 0)))

# ------------------------------------------------- expected formula text, cell by cell
F = {}
F[("My Debts", "C17")] = "=SUM(C7:C16)"
F[("My Debts", "E17")] = "=SUM(E7:E16)"
F[("My Debts", "E19")] = "=E17+E4"
for r, sched in ((23, "'Snowball Schedule'"), (24, "'Avalanche Schedule'")):
    F[("My Debts", f"C{r}")] = (f'=IF({sched}!$L$14<=0,"—",IF({sched}!$O$2=999,"30+ years",'
                                f'TEXT(EDATE(TODAY(),{sched}!$O$2),"MMM YYYY")))')
    F[("My Debts", f"D{r}")] = f'=IF({sched}!$L$14<=0,"—",IF({sched}!$O$2=999,"360+",{sched}!$O$2))'
    F[("My Debts", f"E{r}")] = f'=IF({sched}!$L$14<=0,"—",ROUND({sched}!$O$3,2))'
F[("Progress", "D16")] = "=SUM(D5:D14)"

for tab, sched in (("Snowball", "Snowball Schedule"), ("Avalanche", "Avalanche Schedule")):
    S = f"'{sched}'"
    F[(tab, "D6")] = (f'=IF({S}!$L$14<=0,"add your debts first",IF({S}!$O$2=999,'
                      f'"30+ years — raise your payment",TEXT(EDATE(TODAY(),{S}!$O$2),"MMMM YYYY")))')
    F[(tab, "D7")] = f'=IF({S}!$L$14<=0,"—",IF({S}!$O$2=999,"360+",{S}!$O$2))'
    F[(tab, "D8")] = f'=IF({S}!$L$14<=0,"—",ROUND({S}!$O$3,2))'
    F[(tab, "D9")] = f'=IF({S}!$L$14<=0,"—",IF(OR({S}!$O$2=999,{S}!$O$4=999),"—",{S}!$O$4-{S}!$O$2))'
    F[(tab, "D10")] = f'=IF({S}!$L$14<=0,"—",IF(OR({S}!$O$2=999,{S}!$O$4=999),"—",ROUND({S}!$O$5-{S}!$O$3,2)))'
    for i in range(1, N_DEBTS + 1):
        r = 12 + i
        F[(tab, f"B{r}")] = f"=INDEX({S}!$B$6:$K$6,{i})"
        F[(tab, f"C{r}")] = f'=IF(INDEX({S}!$B$5:$K$5,{i})<=0,"",INDEX({S}!$B$5:$K$5,{i}))'
        F[(tab, f"D{r}")] = f'=IF(C{r}="","",INDEX({S}!$B$7:$K$7,{i}))'
        F[(tab, f"E{r}")] = f'=IF(C{r}="","",INDEX({S}!$B$9:$K$9,{i}))'
        F[(tab, f"F{r}")] = (f'=IF(C{r}="","",IF(INDEX({S}!$B$10:$K$10,{i})=999,"360+",'
                             f'INDEX({S}!$B$10:$K$10,{i})))')
        F[(tab, f"G{r}")] = (f'=IF(C{r}="","",IF(INDEX({S}!$B$10:$K$10,{i})=999,"—",'
                             f'TEXT(EDATE(TODAY(),INDEX({S}!$B$10:$K$10,{i})),"MMM YYYY")))')

for sheet, strategy in (("Snowball Schedule", "snowball"), ("Avalanche Schedule", "avalanche")):
    for j in range(N_DEBTS):
        col = BAL0_COL + j
        cl = L(col)
        md_row = 7 + j
        if strategy == "snowball":
            key = f"=IF(N({MD}!C{md_row})<=0,9000000000+COLUMN(),N({MD}!C{md_row})+COLUMN()/1000)"
        else:
            key = f"=IF(N({MD}!C{md_row})<=0,9000000000+COLUMN(),10-N({MD}!D{md_row})+COLUMN()/100000)"
        F[(sheet, f"{cl}3")] = key
        F[(sheet, f"{cl}4")] = f"=MATCH(SMALL($B$3:$K$3,{cl}$2),$B$3:$K$3,0)"
        F[(sheet, f"{cl}5")] = f"=ROUND(MAX(0,N(INDEX({MD}!$C$7:$C$16,{cl}$4))),2)"
        F[(sheet, f"{cl}6")] = f'=IF({cl}$5<=0,"",T(INDEX({MD}!$B$7:$B$16,{cl}$4)))'
        F[(sheet, f"{cl}7")] = f"=IF({cl}$5<=0,0,MAX(0,N(INDEX({MD}!$D$7:$D$16,{cl}$4))))"
        F[(sheet, f"{cl}8")] = f"={cl}$7/12"
        F[(sheet, f"{cl}9")] = f"=IF({cl}$5<=0,0,MAX(0,N(INDEX({MD}!$E$7:$E$16,{cl}$4))))"
        F[(sheet, f"{cl}10")] = f'=IF({cl}$5<=0,"",IFERROR(MATCH(0,{cl}${GRID0 + 1}:{cl}${GRID_LAST},0),999))'
        F[(sheet, f"{L(BAL0_COL + j)}{GRID0}")] = f"={L(BAL0_COL + j)}$5"
        F[(sheet, f"{L(BBAL_COL + j)}{GRID0}")] = f"={L(BAL0_COL + j)}$5"
    F[(sheet, "B11")] = f"=MAX(0,N({MD}!$E$4))+SUM($B$9:$K$9)"
    F[(sheet, "B12")] = "=SUM($B$9:$K$9)"
    F[(sheet, "O2")] = f"=IF($L${GRID0}<=0,0,IFERROR(MATCH(0,$L${GRID0 + 1}:$L${GRID_LAST},0),999))"
    F[(sheet, "O3")] = f"=SUM($M${GRID0 + 1}:$M${GRID_LAST})"
    F[(sheet, "O4")] = f"=IF($L${GRID0}<=0,0,IFERROR(MATCH(0,$AS${GRID0 + 1}:$AS${GRID_LAST},0),999))"
    F[(sheet, "O5")] = f"=SUM($AT${GRID0 + 1}:$AT${GRID_LAST})"
    F[(sheet, f"L{GRID0}")] = f"=SUM(B{GRID0}:K{GRID0})"
    F[(sheet, f"AS{GRID0}")] = f"=SUM(AI{GRID0}:AR{GRID0})"
    for m in range(1, N_MONTHS + 1):
        r = GRID0 + m
        p = r - 1
        for base_bal, base_mp, base_pay, budget in ((BAL0_COL, MP0_COL, PAY0_COL, "$B$11"),
                                                    (BBAL_COL, BMP_COL, BPAY_COL, "$B$12")):
            for j in range(N_DEBTS):
                bal = L(base_bal + j); mp = L(base_mp + j); pay = L(base_pay + j)
                rate = L(BAL0_COL + j)
                owed = f"{bal}{p}*(1+{rate}$8)"
                F[(sheet, f"{mp}{r}")] = f"=MIN({owed},{rate}$9)"
                terms = [budget]
                if j > 0:
                    terms.append(f"-SUM({L(base_pay)}{r}:{L(base_pay + j - 1)}{r})")
                if j < N_DEBTS - 1:
                    terms.append(f"-SUM({L(base_mp + j + 1)}{r}:{L(base_mp + N_DEBTS - 1)}{r})")
                F[(sheet, f"{pay}{r}")] = f"=MIN({owed},MAX({mp}{r},{''.join(terms)}))"
                F[(sheet, f"{bal}{r}")] = f"=ROUND(MAX(0,{owed}-{pay}{r}),2)"
        F[(sheet, f"L{r}")] = f"=SUM(B{r}:K{r})"
        F[(sheet, f"M{r}")] = f"=SUMPRODUCT(B{p}:K{p},$B$8:$K$8)"
        F[(sheet, f"AS{r}")] = f"=SUM(AI{r}:AR{r})"
        F[(sheet, f"AT{r}")] = f"=SUMPRODUCT(AI{p}:AR{p},$B$8:$K$8)"

# ---------------------------------------------- pass 1: every formula string matches
func_re = re.compile(r"([A-Z][A-Z0-9]*)\(")
n_formulas = 0
pending = dict(F)
for sheet in wf.sheetnames:
    ws = wf[sheet]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            n_formulas += 1
            key = (sheet, cell.coordinate)
            if key not in F:
                fails.append(f"unexpected formula at {key}: {v}")
                continue
            if v != F[key]:
                fails.append(f"formula mismatch at {key}:\n  found    {v}\n  expected {F[key]}")
            pending.pop(key, None)
            bad = set(func_re.findall(v)) - ALLOWED_FUNCS
            if bad:
                fails.append(f"{key} uses non-whitelisted function(s) {bad}")
for key in sorted(pending):
    fails.append(f"expected formula never found at {key}")

# ---------------------------------------------- python engine (mirrors the formulas)
def order_debts(strategy):
    keys = []
    for j, (name, bal, apr, mn) in enumerate(DEBTS):
        col = BAL0_COL + j  # COLUMN() of the key cell
        if bal <= 0:
            keys.append(9000000000 + col)
        elif strategy == "snowball":
            keys.append(bal + col / 1000)
        else:
            keys.append(10 - apr + col / 100000)
    order = sorted(range(N_DEBTS), key=lambda j: keys[j])
    return order  # order[slot] = debt index (0-based)


def run_engine(order, budget):
    """Float engine, same operation order as the spreadsheet formulas."""
    start = [xround(max(0.0, DEBTS[i][1])) for i in order]
    rate = [(DEBTS[i][2] / 12 if start[s] > 0 else 0.0) for s, i in enumerate(order)]
    mn = [(max(0.0, DEBTS[i][3]) if start[s] > 0 else 0.0) for s, i in enumerate(order)]
    bal = start[:]
    months = {"bal": [bal[:]], "mp": [], "pay": [], "tot": [sum(bal)], "int": []}
    for m in range(1, N_MONTHS + 1):
        prev = bal[:]
        intm = sum(prev[j] * rate[j] for j in range(N_DEBTS))
        owed = [prev[j] * (1 + rate[j]) for j in range(N_DEBTS)]
        mp = [min(owed[j], mn[j]) for j in range(N_DEBTS)]
        pays = []
        for j in range(N_DEBTS):
            avail = max(mp[j], budget - sum(pays) - sum(mp[j + 1:]))
            pays.append(min(owed[j], avail))
        bal = [xround(max(0.0, owed[j] - pays[j])) for j in range(N_DEBTS)]
        months["bal"].append(bal[:])
        months["mp"].append(mp)
        months["pay"].append(pays)
        months["tot"].append(sum(bal))
        months["int"].append(intm)
    payoff = []
    for j in range(N_DEBTS):
        if start[j] <= 0:
            payoff.append(None)
        else:
            po = next((m for m in range(1, N_MONTHS + 1) if months["bal"][m][j] == 0), 999)
            payoff.append(po)
    if sum(start) <= 0:
        free = 0
    else:
        free = next((m for m in range(1, N_MONTHS + 1) if months["tot"][m] == 0), 999)
    total_int = sum(months["int"])
    return {"start": start, "rate": rate, "mn": mn, "months": months,
            "payoff": payoff, "free": free, "total_int": total_int}


def second_opinion(order, budget):
    """Independently coded allocator: every active debt pays its minimum, then
    the whole surplus pours down the attack order (spilling past payoffs)."""
    bal = [xround(max(0.0, DEBTS[i][1])) for i in order]
    rate = [(DEBTS[i][2] / 12 if bal[s] > 0 else 0.0) for s, i in enumerate(order)]
    mn = [(max(0.0, DEBTS[i][3]) if bal[s] > 0 else 0.0) for s, i in enumerate(order)]
    total_int = 0.0
    payoff = [None if bal[j] <= 0 else 999 for j in range(N_DEBTS)]
    free = 0 if sum(bal) <= 0 else 999
    hist = [bal[:]]
    for m in range(1, N_MONTHS + 1):
        total_int += sum(bal[j] * rate[j] for j in range(N_DEBTS))
        owed = [bal[j] * (1 + rate[j]) for j in range(N_DEBTS)]
        pays = [min(owed[j], mn[j]) for j in range(N_DEBTS)]
        surplus = budget - sum(pays)
        for j in range(N_DEBTS):
            if surplus <= 0:
                break
            add = min(owed[j] - pays[j], surplus)
            if add > 0:
                pays[j] += add
                surplus -= add
        newbal = []
        for j in range(N_DEBTS):
            b = xround(max(0.0, owed[j] - pays[j]))
            newbal.append(b)
            if b == 0 and bal[j] > 0 and payoff[j] == 999:
                payoff[j] = m
        bal = newbal
        hist.append(bal[:])
        if free == 999 and all(b == 0 for b in bal):
            free = m
    return {"payoff": payoff, "free": free, "total_int": total_int, "hist": hist}


# ---------------------------------------------- pass 2: engine vs LibreOffice cache
n_values = 0


def check(sheet, coord, want, tol=1e-6, volatile=False):
    global n_values
    got = wv[sheet][coord].value
    ok = ((want is None and got in (None, "")) or
          (isinstance(want, str) and got == want) or
          (isinstance(want, (int, float)) and isinstance(got, (int, float)) and
           abs(got - want) <= max(tol, abs(want) * 1e-9)))
    n_values += 1
    if not ok:
        msg = f"{sheet}!{coord}: expected {want!r}, cached {got!r}"
        (warns if volatile else fails).append(
            msg + (" [TODAY()-dependent; Excel/Sheets recalc on open]" if volatile else ""))


def date_text(nmonths, fmt_long):
    """Accept a cache computed today or yesterday (recalc may predate verify)."""
    outs = []
    for d in (TODAY, TODAY - timedelta(days=1)):
        e = edate(d, nmonths)
        outs.append(e.strftime("%B %Y") if fmt_long else e.strftime("%b %Y"))
    return outs


results = {}
for sheet, strategy in (("Snowball Schedule", "snowball"), ("Avalanche Schedule", "avalanche")):
    order = order_debts(strategy)
    mins_active = sum(max(0.0, DEBTS[i][3]) for i in order if xround(max(0.0, DEBTS[i][1])) > 0)
    plan = run_engine(order, max(0.0, EXTRA) + mins_active)
    base = run_engine(order, mins_active)
    results[strategy] = {"order": order, "plan": plan, "base": base, "mins_active": mins_active}

    # header block
    check(sheet, "B11", max(0.0, EXTRA) + mins_active)
    check(sheet, "B12", mins_active)
    for s in range(N_DEBTS):
        cl = L(BAL0_COL + s)
        di = order[s]
        check(sheet, f"{cl}4", di + 1)
        check(sheet, f"{cl}5", plan["start"][s])
        name = DEBTS[di][0]
        check(sheet, f"{cl}6", name if (plan["start"][s] > 0 and isinstance(name, str)) else None)
        check(sheet, f"{cl}7", DEBTS[di][2] if plan["start"][s] > 0 else 0)
        check(sheet, f"{cl}8", (DEBTS[di][2] if plan["start"][s] > 0 else 0) / 12)
        check(sheet, f"{cl}9", plan["mn"][s])
        check(sheet, f"{cl}10", plan["payoff"][s])
    # summary
    check(sheet, "O2", plan["free"])
    check(sheet, "O3", plan["total_int"], tol=1e-4)
    check(sheet, "O4", base["free"])
    check(sheet, "O5", base["total_int"], tol=1e-4)
    # every grid cell, plan and baseline
    for run, cols in ((plan, (BAL0_COL, MP0_COL, PAY0_COL, TOT_COL, INT_COL)),
                      (base, (BBAL_COL, BMP_COL, BPAY_COL, BTOT_COL, BINT_COL))):
        cb, cm, cp, ct, ci = cols
        for s in range(N_DEBTS):
            check(sheet, f"{L(cb + s)}{GRID0}", run["months"]["bal"][0][s])
        check(sheet, f"{L(ct)}{GRID0}", run["months"]["tot"][0])
        for m in range(1, N_MONTHS + 1):
            r = GRID0 + m
            for s in range(N_DEBTS):
                check(sheet, f"{L(cb + s)}{r}", run["months"]["bal"][m][s])
                check(sheet, f"{L(cm + s)}{r}", run["months"]["mp"][m - 1][s])
                check(sheet, f"{L(cp + s)}{r}", run["months"]["pay"][m - 1][s])
            check(sheet, f"{L(ct)}{r}", run["months"]["tot"][m])
            check(sheet, f"{L(ci)}{r}", run["months"]["int"][m - 1], tol=1e-5)

# buyer-facing tabs
for tab, strategy in (("Snowball", "snowball"), ("Avalanche", "avalanche")):
    res = results[strategy]
    plan, base, order = res["plan"], res["base"], res["order"]
    if plan["free"] in (0, 999):
        pass  # degenerate cases surface via the schedule checks
    else:
        got = wv[tab]["D6"].value
        if got not in date_text(plan["free"], True):
            warns.append(f"{tab}!D6 debt-free date {got!r} not in {date_text(plan['free'], True)} [TODAY()-dependent]")
        n_values += 1
        check(tab, "D7", plan["free"])
        check(tab, "D8", xround(plan["total_int"]))
        check(tab, "D9", base["free"] - plan["free"])
        check(tab, "D10", xround(base["total_int"] - plan["total_int"]), tol=0.011)
    for s in range(N_DEBTS):
        r = 13 + s
        di = order[s]
        active = plan["start"][s] > 0
        name = DEBTS[di][0]
        check(tab, f"B{r}", name if (active and isinstance(name, str)) else None)
        check(tab, f"C{r}", plan["start"][s] if active else None)
        check(tab, f"D{r}", DEBTS[di][2] if active else None)
        check(tab, f"E{r}", plan["mn"][s] if active else None)
        check(tab, f"F{r}", (plan["payoff"][s] if plan["payoff"][s] != 999 else "360+") if active else None)
        if active and plan["payoff"][s] != 999:
            got = wv[tab][f"G{r}"].value
            if got not in date_text(plan["payoff"][s], False):
                warns.append(f"{tab}!G{r} payoff date {got!r} not in {date_text(plan['payoff'][s], False)} [TODAY()-dependent]")
            n_values += 1
        else:
            check(tab, f"G{r}", None if not active else "—")

# My Debts totals + verdict
tot_bal = sum(max(0.0, d[1]) for d in DEBTS)
tot_min = sum(max(0.0, d[3]) for d in DEBTS)
check("My Debts", "C17", tot_bal)
check("My Debts", "E17", tot_min)
check("My Debts", "E19", tot_min + EXTRA)
for r, strategy in ((23, "snowball"), (24, "avalanche")):
    plan = results[strategy]["plan"]
    got = wv["My Debts"][f"C{r}"].value
    if plan["free"] not in (0, 999):
        if got not in date_text(plan["free"], False):
            warns.append(f"My Debts!C{r} {got!r} not in {date_text(plan['free'], False)} [TODAY()-dependent]")
        n_values += 1
        check("My Debts", f"D{r}", plan["free"])
        check("My Debts", f"E{r}", xround(plan["total_int"]))

# ---------------------------------------------- pass 3: second opinion + invariants
n_sim2 = 0
for strategy in ("snowball", "avalanche"):
    res = results[strategy]
    order = res["order"]
    for run, budget in ((res["plan"], max(0.0, EXTRA) + res["mins_active"]),
                        (res["base"], res["mins_active"])):
        alt = second_opinion(order, budget)
        assert alt["free"] == run["free"], f"{strategy}: sim2 free {alt['free']} vs engine {run['free']}"
        assert alt["payoff"] == run["payoff"], f"{strategy}: sim2 payoffs {alt['payoff']} vs {run['payoff']}"
        assert abs(alt["total_int"] - run["total_int"]) < 0.05, \
            f"{strategy}: sim2 interest {alt['total_int']} vs {run['total_int']}"
        for m in range(N_MONTHS + 1):
            for s in range(N_DEBTS):
                assert abs(alt["hist"][m][s] - run["months"]["bal"][m][s]) <= 0.02, \
                    f"{strategy} m{m} slot{s}: sim2 {alt['hist'][m][s]} vs {run['months']['bal'][m][s]}"
                n_sim2 += 1
        # invariants on the engine run
        principal = sum(run["start"])
        paid = 0.0
        for m in range(1, N_MONTHS + 1):
            mp = run["months"]["mp"][m - 1]
            pays = run["months"]["pay"][m - 1]
            prevbal = run["months"]["bal"][m - 1]
            newbal = run["months"]["bal"][m]
            assert sum(pays) <= budget + 1e-6, f"{strategy} m{m}: budget exceeded"
            beyond = [j for j in range(N_DEBTS) if pays[j] > mp[j] + 1e-9]
            for j in beyond:
                assert all(newbal[k] == 0 for k in range(j)), \
                    f"{strategy} m{m}: surplus reached slot {j} while an earlier debt is unpaid"
            for j in range(N_DEBTS):
                assert pays[j] >= mp[j] - 1e-9, f"{strategy} m{m} slot{j}: minimum not honored"
                assert newbal[j] <= prevbal[j] * (1 + run["rate"][j]) + 1e-6, f"{strategy} m{m} slot{j}: balance grew past interest"
            paid += sum(pays)
        assert abs(paid - (principal + run["total_int"])) <= 5.0, \
            f"{strategy}: conservation off — paid {paid:.2f} vs principal {principal:.2f} + interest {run['total_int']:.2f}"
    # avalanche never pays more interest than snowball on the same budget
assert results["avalanche"]["plan"]["total_int"] <= results["snowball"]["plan"]["total_int"] + 0.01, \
    "avalanche paid MORE interest than snowball — engine ordering is broken"

# ---------------------------------------------- pinned worked example (regression lock)
sp, ap = results["snowball"]["plan"], results["avalanche"]["plan"]
sb, ab = results["snowball"]["base"], results["avalanche"]["base"]
assert results["snowball"]["order"][:5] == [0, 1, 2, 3, 4]
assert results["avalanche"]["order"][:5] == [0, 2, 3, 4, 1]
assert sp["free"] == 31 and ap["free"] == 30, (sp["free"], ap["free"])
assert sp["payoff"][:5] == [4, 8, 18, 28, 31], sp["payoff"][:5]
assert ap["payoff"][:5] == [4, 16, 28, 30, 24], ap["payoff"][:5]
assert abs(sp["total_int"] - 3387.3029925) < 1e-4, sp["total_int"]
assert abs(ap["total_int"] - 3163.8753425) < 1e-4, ap["total_int"]
assert sb["free"] == 42 and ab["free"] == 42
assert abs(sb["total_int"] - 5459.5686975) < 1e-4
assert sp["months"]["pay"][0][0] == 185.0  # month 1: store card gets min 35 + extra 150

# ---------------------------------------------- yellow-cell invariant
n_yellow = 0
for sheet in wf.sheetnames:
    ws = wf[sheet]
    for row in ws.iter_rows():
        for cell in row:
            rgb = getattr(cell.fill.fgColor, "rgb", None)
            yellow = cell.fill.patternType == "solid" and isinstance(rgb, str) and rgb.endswith("FFF2CC")
            if yellow:
                n_yellow += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    fails.append(f"yellow input cell {sheet}!{cell.coordinate} contains a formula")
assert n_yellow >= 90, f"suspiciously few yellow input cells ({n_yellow})"
for sheet in ("Snowball", "Avalanche", "Snowball Schedule", "Avalanche Schedule"):
    ws = wf[sheet]
    for row in ws.iter_rows():
        for cell in row:
            rgb = getattr(cell.fill.fgColor, "rgb", None)
            if cell.fill.patternType == "solid" and isinstance(rgb, str) and rgb.endswith("FFF2CC"):
                fails.append(f"{sheet} must have no yellow cells (fully calculated) — found {cell.coordinate}")

# ---------------------------------------------- charts on Progress
def color_hex(colorish):
    if colorish is None:
        return None
    if isinstance(colorish, str):
        return colorish.lower()
    return (getattr(colorish, "srgbClr", None) or "").lower() or None

prog = wf["Progress"]
charts = prog._charts
if len(charts) != 2:
    fails.append(f"Progress must hold 2 charts, found {len(charts)}")
lines = [c for c in charts if isinstance(c, LineChart)]
bars = [c for c in charts if isinstance(c, BarChart)]
if not lines:
    fails.append("no LineChart (balance decline) on Progress")
else:
    ch = lines[0]
    refs = [s.val.numRef.f if s.val and s.val.numRef else "" for s in ch.series]
    want = [f"'Snowball Schedule'!$L${GRID0}:$L${GRID0 + CHART_MONTHS}",
            f"'Avalanche Schedule'!$L${GRID0}:$L${GRID0 + CHART_MONTHS}",
            f"'Snowball Schedule'!$AS${GRID0}:$AS${GRID0 + CHART_MONTHS}"]
    for w in want:
        if not any(w == r for r in refs):
            fails.append(f"decline chart missing series {w}; has {refs}")
    cols = [color_hex(s.graphicalProperties.line.solidFill) if s.graphicalProperties and s.graphicalProperties.line else None
            for s in ch.series]
    for w in (NAVY.lower(), TEAL.lower(), GOLD.lower()):
        if w not in cols:
            fails.append(f"decline chart series colors {cols} missing {w}")
if not bars:
    fails.append("no BarChart (per-debt balances) on Progress")
else:
    ch = bars[0]
    refs = [s.val.numRef.f if s.val and s.val.numRef else "" for s in ch.series]
    if not any("'My Debts'!$C$7:$C$16" == r for r in refs):
        fails.append(f"per-debt bar chart wrong series refs: {refs}")
    col = color_hex(ch.series[0].graphicalProperties.solidFill) if ch.series and ch.series[0].graphicalProperties else None
    if col != GOLD.lower():
        fails.append(f"per-debt bars color {col} != gold")

# ---------------------------------------------------------------- report
print(f"formulas checked : {n_formulas} (every string matches; two-way coverage; whitelist ok)")
print(f"values checked   : {n_values} (python engine vs LibreOffice cache — all {N_MONTHS} months x {N_DEBTS} debts x plan+baseline x both strategies)")
print(f"second opinion   : {n_sim2} balance cells agree with the independently coded allocator; invariants + conservation hold")
print(f"yellow inputs    : {n_yellow}; strategy/schedule tabs contain none (fully calculated)")
print(f"charts           : decline line (Snowball navy / Avalanche teal, months 0-{CHART_MONTHS}) + per-debt gold bars")
print(f"macros           : none (vba_archive is None)")
print("worked example (engine, verified 3 ways):")
print(f"  budget ${max(0.0, EXTRA) + results['snowball']['mins_active']:,.2f}/mo "
      f"(minimums ${results['snowball']['mins_active']:,.2f} + extra ${EXTRA:,.2f})")
print(f"  Snowball : debt-free month {sp['free']} ({edate(TODAY, sp['free']):%B %Y}), interest ${sp['total_int']:,.2f}, "
      f"payoffs {[p for p in sp['payoff'] if p]}")
print(f"  Avalanche: debt-free month {ap['free']} ({edate(TODAY, ap['free']):%B %Y}), interest ${ap['total_int']:,.2f}, "
      f"payoffs {[p for p in ap['payoff'] if p]}")
print(f"  minimums only: {sb['free']} months, interest ${sb['total_int']:,.2f} -> extra payment saves "
      f"{sb['free'] - sp['free']} months and ${sb['total_int'] - sp['total_int']:,.2f} (snowball)")

if warns:
    print(f"\nWARNINGS ({len(warns)}) — TODAY()-dependent caches; Excel/Sheets refresh on open:")
    for w in warns:
        print("  " + w)
if fails:
    print(f"\nFAILURES ({len(fails)}):")
    for f in fails[:60]:
        print("  " + f)
    if len(fails) > 60:
        print(f"  ... and {len(fails) - 60} more")
    sys.exit(1)
print("\nVERIFIED: all checks passed")
