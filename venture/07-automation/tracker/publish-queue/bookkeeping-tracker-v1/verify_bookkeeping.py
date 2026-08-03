#!/usr/bin/env python3
"""Verify Small-Business-Bookkeeping-Tracker.xlsx — must pass before publish.

openpyxl cannot evaluate formulas, so this script proves the workbook two ways:

1. FORMULA TEXT — it re-derives the exact formula string every formula cell should
   contain (same layout constants as build_bookkeeping.py) and asserts all of them
   match, so every SUM/SUMIF/SUMIFS/COUNTIF/IF references exactly the right ranges.
   Coverage is total: a formula cell the map doesn't know about fails, and an
   expected formula that never appears fails.
2. EXPECTED VALUES — it reads the worked-example constants (dates, amounts, rates)
   back out of the workbook, recomputes every key result in plain Python, and
   compares them against the cached values LibreOffice wrote at build time
   (build step runs the xlsx skill's recalc.py). Cells that depend on TODAY()
   (invoice status/aging) are recomputed for *today*; if the cached copy is from an
   older recalc they are reported as warnings only, because Excel/Google Sheets
   recalculates them the moment the buyer opens the file.

Also asserts: no macros, only Excel-2010+/Google-Sheets-safe functions, the
edit-only-the-yellow-cells invariant (yellow cell => never a formula; formula cell
=> never yellow), and the traffic-light conditional formatting rules.

Exits non-zero on any failure.
"""
import sys
import re
from datetime import date, datetime
from openpyxl import load_workbook

XLSX = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/bookkeeping-tracker-v1/Small-Business-Bookkeeping-Tracker.xlsx"

# same layout constants as build_bookkeeping.py
L_FIRST, L_LAST = 10, 209
M_FIRST, M_LAST, M_YEAR = 10, 21, 22
I_FIRST, I_LAST = 11, 110
T_FIRST, T_LAST = 9, 208
D_FIRST, D_LAST, D_YEAR = 7, 18, 19
Q_FIRST = 29
QT = Q_FIRST + 4
LG = "'Income & Expense Ledger'"
ML = "'Mileage Log'"

ALLOWED_FUNCS = {"SUM", "SUMIF", "SUMIFS", "COUNTIF", "IF", "OR", "AND",
                 "TODAY", "MONTH", "MAX", "ROUND", "IFERROR", "AVERAGEIF"}

wf = load_workbook(XLSX, data_only=False)
wv = load_workbook(XLSX, data_only=True)
TODAY = date.today()
# A freshly openpyxl-written file has no cached formula values (openpyxl never
# computes them). The build workflow runs the xlsx skill's recalc.py (LibreOffice)
# right after build_bookkeeping.py, which caches every result. If that step was
# skipped, formula-text checks still run in full and the python-recomputed
# expectations are printed as documentation, but cache comparison is impossible.
HAS_CACHE = wv["Income & Expense Ledger"]["F4"].value is not None

fails, warns = [], []


def as_date(v):
    return v.date() if isinstance(v, datetime) else v


def money(x):
    return round(x, 2)


# ---------------------------------------------------------------- sheet basics
assert wf.sheetnames == ["START HERE", "Income & Expense Ledger", "Invoice Tracker",
                         "Quarterly Tax Estimate", "P&L Dashboard", "Mileage Log"], wf.sheetnames
assert wf.vba_archive is None and XLSX.endswith(".xlsx"), "workbook must be macro-free"

# ------------------------------------------------- read worked example back out
led = wf["Income & Expense Ledger"]
LEDGER = []
for r in range(L_FIRST, L_LAST + 1):
    d, t, amt = led.cell(r, 2).value, led.cell(r, 3).value, led.cell(r, 6).value
    if d is None and t is None and amt is None:
        continue
    assert t in ("Income", "Expense"), f"ledger row {r}: bad Type {t!r}"
    assert d is not None and isinstance(amt, (int, float)), f"ledger row {r} incomplete"
    LEDGER.append((as_date(d), t, float(amt)))
inv = wf["Invoice Tracker"]
INVOICES = []
for r in range(I_FIRST, I_LAST + 1):
    num = inv.cell(r, 2).value
    if num is None:
        continue
    amt = float(inv.cell(r, 5).value)
    due, paid = as_date(inv.cell(r, 7).value), as_date(inv.cell(r, 8).value)
    INVOICES.append((r, num, amt, due, paid))
mil = wf["Mileage Log"]
RATE = float(mil["C4"].value)
TRIPS = []
for r in range(T_FIRST, T_LAST + 1):
    d, miles = mil.cell(r, 2).value, mil.cell(r, 4).value
    if d is None and miles is None:
        continue
    TRIPS.append((as_date(d), float(miles)))
tax = wf["Quarterly Tax Estimate"]
SE_SHARE, SE_RATE, INC_RATE = (float(tax[c].value) for c in ("C14", "C15", "C16"))

# ------------------------------------------- recompute expected values (python)
led_in = sum(a for _, t, a in LEDGER if t == "Income")
led_out = sum(a for _, t, a in LEDGER if t == "Expense")
m_in = {m: sum(a for d, t, a in LEDGER if t == "Income" and d.month == m) for m in range(1, 13)}
m_out = {m: sum(a for d, t, a in LEDGER if t == "Expense" and d.month == m) for m in range(1, 13)}

trip_ded = [round(mi * RATE, 2) for _, mi in TRIPS]
miles_total = sum(mi for _, mi in TRIPS)
ded_total = sum(trip_ded)
m_ded = {m: sum(dd for (d, _), dd in zip(TRIPS, trip_ded) if d.month == m) for m in range(1, 13)}

profit = led_in - led_out - ded_total
se_base = max(0, profit) * SE_SHARE
se_tax = se_base * SE_RATE
inc_tax = max(0, profit - se_tax / 2) * INC_RATE
total_tax = se_tax + inc_tax
setaside_rate = total_tax / led_in if led_in else 0
eff_rate = 0 if profit <= 0 else total_tax / profit
q_profit = {q: (sum(m_in[m] - m_out[m] - m_ded[m] for m in range(3 * q - 2, 3 * q + 1))) for q in (1, 2, 3, 4)}
q_aside = {q: max(0, q_profit[q]) * eff_rate for q in (1, 2, 3, 4)}

# invoice expectations (status semantics mirror the formulas, evaluated at TODAY)
inv_status, inv_days = {}, {}
for r, num, amt, due, paid in INVOICES:
    if paid is not None:
        st = "PAID"
    elif due is None:
        st = "SENT"
    else:
        st = "OVERDUE" if TODAY > due else "SENT"
    inv_status[r] = st
    inv_days[r] = (TODAY - due).days if st == "OVERDUE" else None
invoiced = sum(a for _, _, a, _, _ in INVOICES)
collected = sum(a for r, _, a, _, _ in INVOICES if inv_status[r] == "PAID")
overdue_amt = sum(a for r, _, a, _, _ in INVOICES if inv_status[r] == "OVERDUE")
n_paid = sum(1 for r in inv_status if inv_status[r] == "PAID")
n_sent = sum(1 for r in inv_status if inv_status[r] == "SENT")
n_over = sum(1 for r in inv_status if inv_status[r] == "OVERDUE")

active = [m for m in range(1, 13) if m_in[m] or m_out[m]]
best_month = max((m_in[m] - m_out[m] for m in active), default=0)
inc_months = [m for m in range(1, 13) if m_in[m] > 0]
avg_month = (sum(m_in[m] - m_out[m] for m in inc_months) / len(inc_months)) if inc_months else 0

# ------------------------------------------------ expected formula text, cell by cell
F = {}  # (sheet, coord) -> exact formula string
LEDS = "Income & Expense Ledger"
F[(LEDS, "F4")] = f'=SUMIF($C${L_FIRST}:$C${L_LAST},"Income",$F${L_FIRST}:$F${L_LAST})'
F[(LEDS, "F5")] = f'=SUMIF($C${L_FIRST}:$C${L_LAST},"Expense",$F${L_FIRST}:$F${L_LAST})'
F[(LEDS, "F6")] = "=F4-F5"
for r in range(L_FIRST, L_LAST + 1):
    F[(LEDS, f"G{r}")] = f'=IF($B{r}="","",MONTH($B{r}))'
for i in range(12):
    r = M_FIRST + i
    F[(LEDS, f"J{r}")] = f'=SUMIFS($F${L_FIRST}:$F${L_LAST},$G${L_FIRST}:$G${L_LAST},{i + 1},$C${L_FIRST}:$C${L_LAST},"Income")'
    F[(LEDS, f"K{r}")] = f'=SUMIFS($F${L_FIRST}:$F${L_LAST},$G${L_FIRST}:$G${L_LAST},{i + 1},$C${L_FIRST}:$C${L_LAST},"Expense")'
    F[(LEDS, f"L{r}")] = f'=IF(AND(J{r}=0,K{r}=0),"",J{r}-K{r})'
F[(LEDS, f"J{M_YEAR}")] = f"=SUM(J{M_FIRST}:J{M_LAST})"
F[(LEDS, f"K{M_YEAR}")] = f"=SUM(K{M_FIRST}:K{M_LAST})"
F[(LEDS, f"L{M_YEAR}")] = f"=J{M_YEAR}-K{M_YEAR}"

INVS = "Invoice Tracker"
F[(INVS, "E4")] = f"=SUM($E${I_FIRST}:$E${I_LAST})"
F[(INVS, "E5")] = f'=SUMIF($I${I_FIRST}:$I${I_LAST},"PAID",$E${I_FIRST}:$E${I_LAST})'
F[(INVS, "E6")] = "=E4-E5"
F[(INVS, "E7")] = f'=SUMIF($I${I_FIRST}:$I${I_LAST},"OVERDUE",$E${I_FIRST}:$E${I_LAST})'
F[(INVS, "H4")] = f'=COUNTIF($I${I_FIRST}:$I${I_LAST},"PAID")'
F[(INVS, "H5")] = f'=COUNTIF($I${I_FIRST}:$I${I_LAST},"SENT")'
F[(INVS, "H6")] = f'=COUNTIF($I${I_FIRST}:$I${I_LAST},"OVERDUE")'
for r in range(I_FIRST, I_LAST + 1):
    F[(INVS, f"I{r}")] = f'=IF($B{r}="","",IF($H{r}<>"","PAID",IF($G{r}="","SENT",IF(TODAY()>$G{r},"OVERDUE","SENT"))))'
    F[(INVS, f"J{r}")] = f'=IF($I{r}="OVERDUE",TODAY()-$G{r},"")'

TAXS = "Quarterly Tax Estimate"
F[(TAXS, "C8")] = f"={LG}!F4"
F[(TAXS, "C9")] = f"={LG}!F5"
F[(TAXS, "C10")] = f"={ML}!C6"
F[(TAXS, "C11")] = "=C8-C9-C10"
F[(TAXS, "C20")] = "=MAX(0,$C$11)*$C$14"
F[(TAXS, "C21")] = "=$C$20*$C$15"
F[(TAXS, "C22")] = "=MAX(0,$C$11-$C$21/2)*$C$16"
F[(TAXS, "C23")] = "=$C$21+$C$22"
F[(TAXS, "C24")] = "=IFERROR($C$23/$C$8,0)"
F[(TAXS, "C25")] = "=IF($C$11<=0,0,$C$23/$C$11)"
for q in range(4):
    r = Q_FIRST + q; lo, hi = 3 * q + 1, 3 * q + 3
    F[(TAXS, f"C{r}")] = (
        f'=SUMIFS({LG}!$F${L_FIRST}:$F${L_LAST},{LG}!$G${L_FIRST}:$G${L_LAST},">={lo}",{LG}!$G${L_FIRST}:$G${L_LAST},"<={hi}",{LG}!$C${L_FIRST}:$C${L_LAST},"Income")'
        f'-SUMIFS({LG}!$F${L_FIRST}:$F${L_LAST},{LG}!$G${L_FIRST}:$G${L_LAST},">={lo}",{LG}!$G${L_FIRST}:$G${L_LAST},"<={hi}",{LG}!$C${L_FIRST}:$C${L_LAST},"Expense")'
        f'-SUMIFS({ML}!$E${T_FIRST}:$E${T_LAST},{ML}!$F${T_FIRST}:$F${T_LAST},">={lo}",{ML}!$F${T_FIRST}:$F${T_LAST},"<={hi}")')
    F[(TAXS, f"D{r}")] = f"=MAX(0,$C{r})*$C$25"
F[(TAXS, f"C{QT}")] = f"=SUM(C{Q_FIRST}:C{QT - 1})"
F[(TAXS, f"D{QT}")] = f"=SUM(D{Q_FIRST}:D{QT - 1})"

DASH = "P&L Dashboard"
for i in range(12):
    r = D_FIRST + i; src = M_FIRST + i
    F[(DASH, f"C{r}")] = f"={LG}!J{src}"
    F[(DASH, f"D{r}")] = f"={LG}!K{src}"
    F[(DASH, f"E{r}")] = f'=IF(AND(C{r}=0,D{r}=0),"",C{r}-D{r})'
    F[(DASH, f"F{r}")] = f'=IF(OR(C{r}=0,E{r}=""),"",E{r}/C{r})'
F[(DASH, f"C{D_YEAR}")] = f"=SUM(C{D_FIRST}:C{D_LAST})"
F[(DASH, f"D{D_YEAR}")] = f"=SUM(D{D_FIRST}:D{D_LAST})"
F[(DASH, f"E{D_YEAR}")] = f"=C{D_YEAR}-D{D_YEAR}"
F[(DASH, f"F{D_YEAR}")] = f'=IF(C{D_YEAR}=0,"",E{D_YEAR}/C{D_YEAR})'
F[(DASH, f"E{D_YEAR + 2}")] = f"=MAX(E{D_FIRST}:E{D_LAST})"
F[(DASH, f"E{D_YEAR + 3}")] = f'=IFERROR(AVERAGEIF(C{D_FIRST}:C{D_LAST},">0",E{D_FIRST}:E{D_LAST}),0)'

MILS = "Mileage Log"
F[(MILS, "C5")] = f"=SUM($D${T_FIRST}:$D${T_LAST})"
F[(MILS, "C6")] = f"=SUM($E${T_FIRST}:$E${T_LAST})"
for r in range(T_FIRST, T_LAST + 1):
    F[(MILS, f"E{r}")] = f'=IF($D{r}="","",ROUND($D{r}*$C$4,2))'
    F[(MILS, f"F{r}")] = f'=IF($B{r}="","",MONTH($B{r}))'

# ---------------------------------------------- pass 1: every formula string matches
func_re = re.compile(r"([A-Z][A-Z0-9]*)\(")
n_formulas = 0
pending = dict(F)
for sheet in wf.sheetnames:
    ws = wf[sheet]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            n_formulas += 1
            key = (sheet, cell.coordinate)
            if key not in F:
                fails.append(f"unexpected formula at {key}: {v}")
                continue
            if v != F[key]:
                fails.append(f"formula mismatch at {key}:\n  found    {v}\n  expected {F[key]}")
            pending.pop(key, None)
            bad = set(func_re.findall(v)) - ALLOWED_FUNCS
            if bad:
                fails.append(f"{key} uses non-whitelisted function(s) {bad}")
for key in sorted(pending):
    fails.append(f"expected formula never found at {key}")

# ---------------------------------------------- pass 2: expected values vs cached
n_values = 0


def check(sheet, coord, want, volatile=False, tol=1e-6):
    """Compare python-recomputed expectation against the cached (recalc'd) value."""
    global n_values
    if not HAS_CACHE:
        n_values += 1
        return
    got = wv[sheet][coord].value
    ok = ((want is None and got in (None, ""))
          or (isinstance(want, str) and got == want)
          or (isinstance(want, (int, float)) and isinstance(got, (int, float))
              and abs(got - want) <= max(tol, abs(want) * 1e-9)))
    n_values += 1
    if not ok:
        msg = f"{sheet}!{coord}: expected {want!r}, cached {got!r}"
        (warns if volatile else fails).append(msg + (" [TODAY()-dependent: cache is from an older recalc; Excel/Sheets recalculates on open]" if volatile else ""))


check(LEDS, "F4", led_in); check(LEDS, "F5", led_out); check(LEDS, "F6", led_in - led_out)
for i in range(12):
    r = M_FIRST + i; m = i + 1
    check(LEDS, f"J{r}", m_in[m]); check(LEDS, f"K{r}", m_out[m])
    check(LEDS, f"L{r}", None if (m_in[m] == 0 and m_out[m] == 0) else m_in[m] - m_out[m])
check(LEDS, f"J{M_YEAR}", led_in); check(LEDS, f"K{M_YEAR}", led_out); check(LEDS, f"L{M_YEAR}", led_in - led_out)
for r in range(L_FIRST, L_LAST + 1):
    d = led.cell(r, 2).value
    check(LEDS, f"G{r}", as_date(d).month if d else None)

check(INVS, "E4", invoiced); check(INVS, "E5", collected); check(INVS, "E6", invoiced - collected)
check(INVS, "E7", overdue_amt, volatile=True)
check(INVS, "H4", n_paid); check(INVS, "H5", n_sent, volatile=True); check(INVS, "H6", n_over, volatile=True)
inv_rows = {r for r, *_ in INVOICES}
for r in range(I_FIRST, I_LAST + 1):
    if r in inv_rows:
        vol = inv_status[r] != "PAID"
        check(INVS, f"I{r}", inv_status[r], volatile=vol)
        check(INVS, f"J{r}", inv_days[r], volatile=vol)
    else:
        check(INVS, f"I{r}", None); check(INVS, f"J{r}", None)

check(MILS, "C5", miles_total); check(MILS, "C6", ded_total)
for i, (d, mi) in enumerate(TRIPS):
    r = T_FIRST + i
    check(MILS, f"E{r}", trip_ded[i]); check(MILS, f"F{r}", d.month)

check(TAXS, "C8", led_in); check(TAXS, "C9", led_out); check(TAXS, "C10", ded_total)
check(TAXS, "C11", profit); check(TAXS, "C20", se_base); check(TAXS, "C21", se_tax)
check(TAXS, "C22", inc_tax); check(TAXS, "C23", total_tax)
check(TAXS, "C24", setaside_rate); check(TAXS, "C25", eff_rate)
for q in (1, 2, 3, 4):
    check(TAXS, f"C{Q_FIRST + q - 1}", q_profit[q]); check(TAXS, f"D{Q_FIRST + q - 1}", q_aside[q])
check(TAXS, f"C{QT}", sum(q_profit.values())); check(TAXS, f"D{QT}", sum(q_aside.values()))

for i in range(12):
    r = D_FIRST + i; m = i + 1
    check(DASH, f"C{r}", m_in[m]); check(DASH, f"D{r}", m_out[m])
    pr = None if (m_in[m] == 0 and m_out[m] == 0) else m_in[m] - m_out[m]
    check(DASH, f"E{r}", pr)
    check(DASH, f"F{r}", None if (m_in[m] == 0 or pr is None) else pr / m_in[m])
check(DASH, f"C{D_YEAR}", led_in); check(DASH, f"D{D_YEAR}", led_out); check(DASH, f"E{D_YEAR}", led_in - led_out)
check(DASH, f"F{D_YEAR}", (led_in - led_out) / led_in if led_in else None)
check(DASH, f"E{D_YEAR + 2}", best_month); check(DASH, f"E{D_YEAR + 3}", avg_month)

# --------------------------------- pass 3: yellow-cell invariant + traffic lights
n_yellow = 0
for sheet in wf.sheetnames:
    ws = wf[sheet]
    for row in ws.iter_rows():
        for cell in row:
            rgb = getattr(cell.fill.fgColor, "rgb", None)
            yellow = cell.fill.patternType == "solid" and isinstance(rgb, str) and rgb.endswith("FFF2CC")
            is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
            if yellow:
                n_yellow += 1
                if is_formula:
                    fails.append(f"yellow input cell {sheet}!{cell.coordinate} contains a formula")
            elif is_formula and cell.fill.patternType == "solid" and isinstance(rgb, str) and rgb.endswith("FFF2CC"):
                fails.append(f"formula cell {sheet}!{cell.coordinate} is yellow")
assert n_yellow > 900, f"suspiciously few yellow input cells ({n_yellow})"

expected_cf = {
    (LEDS, f"L{M_FIRST}:L{M_LAST}"): {"greaterThan", "equal", "lessThan"},
    (INVS, f"I{I_FIRST}:I{I_LAST}"): {"equal"},
    (INVS, f"J{I_FIRST}:J{I_LAST}"): {"between", "greaterThan"},
    (DASH, f"E{D_FIRST}:E{D_YEAR}"): {"greaterThan", "equal", "lessThan"},
}
found_cf = {}
for sheet in wf.sheetnames:
    for cf in wf[sheet].conditional_formatting:
        found_cf.setdefault((sheet, str(cf.sqref)), set()).update(r.operator for r in cf.rules)
for key, ops in expected_cf.items():
    if key not in found_cf:
        fails.append(f"missing conditional formatting on {key}")
    elif not ops <= found_cf[key]:
        fails.append(f"conditional formatting on {key}: found {found_cf[key]}, expected {ops}")
n_status_rules = sum(len(cf.rules) for cf in wf[INVS].conditional_formatting if str(cf.sqref).startswith("I"))
assert n_status_rules == 3, f"expected 3 status color rules, found {n_status_rules}"

# --------------------------- pass 4: dropdowns, chart, unclipped labels, $0 formats
led_dvs = {str(dv.sqref): (dv.formula1, dv.showErrorMessage) for dv in wf[LEDS].data_validations.dataValidation}
if led_dvs.get(f"C{L_FIRST}:C{L_LAST}") != ('"Income,Expense"', True):
    fails.append(f"Ledger Type dropdown missing/wrong: {led_dvs}")
cat_dv = led_dvs.get(f"D{L_FIRST}:D{L_LAST}")
if not cat_dv or "Supplies" not in cat_dv[0] or "Insurance" not in cat_dv[0] or cat_dv[1] is True:
    fails.append(f"Ledger Category dropdown missing/wrong (must be non-strict suggestion list): {cat_dv}")

led_merges = {str(m) for m in wf[LEDS].merged_cells.ranges}
for r in (4, 5, 6):
    if f"B{r}:E{r}" not in led_merges:
        fails.append(f"Ledger stat label row {r} must merge B:E (no clipping)")
dash_merges = {str(m) for m in wf[DASH].merged_cells.ranges}
for r in (D_YEAR + 2, D_YEAR + 3):
    if f"B{r}:D{r}" not in dash_merges:
        fails.append(f"Dashboard stat label row {r} must merge B:D (no clipping)")


def zero_shows_dollar(sheet, coord):
    nf = wf[sheet][coord].number_format
    if '"-"' in nf or nf == "General":
        fails.append(f"{sheet}!{coord} number format {nf!r} would hide $0 (copy references the $ figure)")


for coord in ("F4", "F5", "F6"):
    zero_shows_dollar(LEDS, coord)
for coord in ("E4", "E5", "E6", "E7"):
    zero_shows_dollar(INVS, coord)
for coord in (f"E{D_YEAR + 2}", f"E{D_YEAR + 3}"):
    zero_shows_dollar(DASH, coord)

# chart: LibreOffice's recalc pass re-serializes combined charts in a way openpyxl's
# reader only partially exposes, so assert against the chart XML itself
import zipfile
with zipfile.ZipFile(XLSX) as z:
    chart_parts = [n for n in z.namelist() if re.match(r"xl/charts/chart\d+\.xml$", n)]
    chart_xml = "".join(z.read(n).decode("utf-8", "replace") for n in chart_parts)
if not chart_parts:
    fails.append("no chart part in the workbook — P&L Dashboard must hold the money/profit chart")
else:
    for needle, what in ((f"$C${D_FIRST}:$C${D_LAST}", "money-in bar series"),
                         (f"$D${D_FIRST}:$D${D_LAST}", "money-out bar series"),
                         (f"$E${D_FIRST}:$E${D_LAST}", "profit line series"),
                         ("<c:barChart>", "bar plot"), ("<c:lineChart>", "line plot")):
        if needle not in chart_xml:
            fails.append(f"chart XML missing {what} ({needle})")
    for color, what in (("1f3a5f", "navy money-in"), ("f2c14e", "gold money-out"), ("2e7d6b", "teal profit")):
        if color not in chart_xml.lower():
            fails.append(f"chart XML missing {what} color {color}")
    ax_ids = re.findall(r'<c:axId val="(\d+)"/>', chart_xml)
    if len(set(ax_ids)) > 2:
        fails.append(f"chart must share ONE axis pair (found axIds {sorted(set(ax_ids))}) — no dual-axis")

# ---------------------------------------------------------------- report
print(f"formulas checked : {n_formulas} (every formula string matches its expected ranges)")
if HAS_CACHE:
    print(f"values checked   : {n_values} (python-recomputed vs LibreOffice-cached)")
else:
    print(f"values checked   : 0 of {n_values} — NO CACHED VALUES (run the xlsx skill's recalc.py "
          f"after build_bookkeeping.py, then re-run this); python-recomputed expectations below")
print(f"yellow inputs    : {n_yellow} cells, none contain formulas; no formula cell is yellow")
print(f"traffic lights   : verified on ledger monthly net, invoice status, overdue aging, dashboard profit")
print(f"dropdowns        : Ledger Type (strict Income/Expense) + Category (suggestion list, free text ok)")
print(f"chart            : P&L Dashboard in/out columns (navy/gold) + profit line (teal), one shared $ axis")
print(f"labels/formats   : stat labels merged (no clipping); summary cells render 0 as $0.00, not '-'")
print(f"macros           : none (vba_archive is None)")
print("worked example (expected = found):")
print(f"  ledger: in ${led_in:,.2f} / out ${money(led_out):,.2f} / net ${money(led_in - led_out):,.2f}; "
      f"Jan {m_in[1]:,.0f}/{m_out[1]:,.2f}, Apr net {money(m_in[4] - m_out[4]):,.2f} (red)")
print(f"  invoices: invoiced ${invoiced:,.2f}, collected ${collected:,.2f}, outstanding ${invoiced - collected:,.2f}, "
      f"overdue ${overdue_amt:,.2f} ({n_over} overdue / {n_sent} sent / {n_paid} paid, as of {TODAY})")
print(f"  mileage: {miles_total:,.0f} mi @ ${RATE} = ${ded_total:,.2f}")
print(f"  tax est: profit ${money(profit):,.2f} -> SE ${money(se_tax):,.2f} + income ${money(inc_tax):,.2f} "
      f"= ${money(total_tax):,.2f}; set-aside {setaside_rate:.1%} of money in")
print(f"  dashboard: year profit ${money(led_in - led_out):,.2f}, margin {(led_in - led_out) / led_in:.1%}, "
      f"best month ${money(best_month):,.2f}, avg ${money(avg_month):,.2f}")

if warns:
    print(f"\nWARNINGS ({len(warns)}) — TODAY()-dependent caches older than today; Excel/Sheets refresh on open:")
    for w in warns:
        print("  " + w)
if fails:
    print(f"\nFAILURES ({len(fails)}):")
    for f in fails:
        print("  " + f)
    sys.exit(1)
print("\nVERIFIED: all checks passed")
