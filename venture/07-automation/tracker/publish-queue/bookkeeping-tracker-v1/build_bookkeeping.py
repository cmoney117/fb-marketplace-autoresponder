#!/usr/bin/env python3
"""Build the Small Business Bookkeeping & Invoice Tracker spreadsheet product.

Deterministic: writes Small-Business-Bookkeeping-Tracker.xlsx + store-card.svg
(1200x1200) into this folder and NOTHING outside it. Formulas only — no macros —
chosen to work in Excel 2010+ and free Google Sheets (SUM/SUMIF/SUMIFS/COUNTIF/
IF/OR/AND/TODAY/MONTH/MAX/ROUND/IFERROR/AVERAGEIF).

Companion script verify_bookkeeping.py re-derives every formula string and every
expected value from the worked-example data and must pass before publish.
"""
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference

OUT = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/bookkeeping-tracker-v1"
os.makedirs(OUT, exist_ok=True)

NAVY = "1F3A5F"; TEAL = "2E7D6B"; GOLD = "F2C14E"; LIGHT = "F4F6F8"; YELLOW = "FFF2CC"
H1 = Font(name="Arial", size=18, bold=True, color="FFFFFF")
H2 = Font(name="Arial", size=12, bold=True, color="FFFFFF")
B = Font(name="Arial", size=11)
BB = Font(name="Arial", size=11, bold=True)
SMALL = Font(name="Arial", size=10, italic=True, color="666666")
WARN = Font(name="Arial", size=11, bold=True, color="9C0006")
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00;($#,##0.00);"-"'   # dash for zero (table cells)
MONEY0 = '$#,##0.00;($#,##0.00)'      # $0.00 for zero (summary/hero cells)
PCT1 = "0.0%"
PCT2 = "0.00%"
DATEFMT = "yyyy-mm-dd"
RATEFMT = "$0.00##"
# traffic-light fills (classic Excel good/neutral/bad)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE"); GREEN_FONT = Font(name="Arial", color="006100")
YEL_FILL = PatternFill("solid", fgColor="FFEB9C"); YEL_FONT = Font(name="Arial", color="9C6500")
RED_FILL = PatternFill("solid", fgColor="FFC7CE"); RED_FONT = Font(name="Arial", color="9C0006")

# ------------------------------------------------------------------ worked example
# One realistic example business: a part-time solo residential/office cleaner.
# (Fictional clients, generic names only — no real brands.)
LEDGER = [  # (date, type, category, description, amount)
    (date(2026, 1, 6),  "Income",  "Client payment", "INV-1001 — Harbor Dental, weekly office cleaning (Jan)", 480.00),
    (date(2026, 1, 9),  "Income",  "Client payment", "INV-1002 — Reyes family, one-time deep clean", 260.00),
    (date(2026, 1, 12), "Expense", "Supplies", "Cleaning supplies restock", 84.63),
    (date(2026, 1, 15), "Expense", "Insurance", "Liability insurance (monthly)", 42.00),
    (date(2026, 1, 20), "Income",  "Client payment", "INV-1003 — Birchwood Apartments, move-out cleans x2", 520.00),
    (date(2026, 1, 22), "Expense", "Advertising", "Neighborhood flyers + local ad", 60.00),
    (date(2026, 1, 28), "Expense", "Software / subscriptions", "Scheduling app (monthly)", 19.00),
    (date(2026, 1, 30), "Income",  "Client payment", "INV-1004 — Harbor Dental, carpet + window add-on", 240.00),
    (date(2026, 2, 3),  "Income",  "Client payment", "INV-1005 — Harbor Dental, weekly office cleaning (Feb)", 480.00),
    (date(2026, 2, 6),  "Expense", "Supplies", "Vacuum bags + microfiber restock", 47.25),
    (date(2026, 2, 10), "Income",  "Client payment", "INV-1006 — Nguyen rental, turnover cleans x4", 640.00),
    (date(2026, 2, 14), "Expense", "Insurance", "Liability insurance (monthly)", 42.00),
    (date(2026, 2, 18), "Income",  "Client payment", "INV-1007 — Reyes family, standard clean", 140.00),
    (date(2026, 2, 24), "Expense", "Equipment", "Replacement mop system", 129.99),
    (date(2026, 2, 26), "Expense", "Software / subscriptions", "Scheduling app (monthly)", 19.00),
    (date(2026, 2, 27), "Income",  "Client payment", "INV-1008 — Birchwood Apartments, move-out clean", 260.00),
    (date(2026, 3, 2),  "Income",  "Client payment", "INV-1009 — Harbor Dental, weekly office cleaning (Mar)", 480.00),
    (date(2026, 3, 5),  "Expense", "Supplies", "Cleaning supplies restock", 96.12),
    (date(2026, 3, 9),  "Income",  "Client payment", "INV-1010 — Nguyen rental, turnover cleans x5", 800.00),
    (date(2026, 3, 13), "Expense", "Insurance", "Liability insurance (monthly)", 42.00),
    (date(2026, 3, 17), "Expense", "Phone", "Business share of phone plan", 35.00),
    (date(2026, 3, 20), "Income",  "Client payment", "INV-1011 — Ferris & Co., quarterly deep clean", 350.00),
    (date(2026, 3, 25), "Expense", "Software / subscriptions", "Scheduling app (monthly)", 19.00),
    (date(2026, 3, 27), "Expense", "Bank / card fees", "Card reader fees (Q1)", 28.47),
    (date(2026, 4, 13), "Expense", "Insurance", "Liability insurance (monthly)", 42.00),
    (date(2026, 4, 27), "Expense", "Software / subscriptions", "Scheduling app (monthly)", 19.00),
]
INVOICES = [  # (number, client, what for, amount, sent, due, paid-or-None)
    ("INV-1001", "Harbor Dental", "Weekly office cleaning — January", 480.00, date(2026, 1, 2), date(2026, 1, 16), date(2026, 1, 6)),
    ("INV-1002", "Reyes family", "One-time deep clean", 260.00, date(2026, 1, 5), date(2026, 1, 19), date(2026, 1, 9)),
    ("INV-1003", "Birchwood Apartments", "Move-out cleans (2 units)", 520.00, date(2026, 1, 13), date(2026, 1, 27), date(2026, 1, 20)),
    ("INV-1004", "Harbor Dental", "Carpet + window add-on", 240.00, date(2026, 1, 26), date(2026, 2, 9), date(2026, 1, 30)),
    ("INV-1005", "Harbor Dental", "Weekly office cleaning — February", 480.00, date(2026, 2, 1), date(2026, 2, 15), date(2026, 2, 3)),
    ("INV-1006", "Nguyen rental", "Turnover cleans (4 stays)", 640.00, date(2026, 2, 4), date(2026, 2, 18), date(2026, 2, 10)),
    ("INV-1007", "Reyes family", "Standard clean", 140.00, date(2026, 2, 16), date(2026, 3, 2), date(2026, 2, 18)),
    ("INV-1008", "Birchwood Apartments", "Move-out clean (1 unit)", 260.00, date(2026, 2, 23), date(2026, 3, 9), date(2026, 2, 27)),
    ("INV-1009", "Harbor Dental", "Weekly office cleaning — March", 480.00, date(2026, 3, 1), date(2026, 3, 15), date(2026, 3, 2)),
    ("INV-1010", "Nguyen rental", "Turnover cleans (5 stays)", 800.00, date(2026, 3, 6), date(2026, 3, 20), date(2026, 3, 9)),
    ("INV-1011", "Ferris & Co.", "Quarterly deep clean", 350.00, date(2026, 3, 16), date(2026, 3, 30), date(2026, 3, 20)),
    ("INV-1012", "Birchwood Apartments", "Move-out clean (1 unit)", 260.00, date(2026, 3, 30), date(2026, 4, 14), None),
    ("INV-1013", "Ferris & Co.", "Window + break-room add-on", 350.00, date(2026, 6, 30), date(2026, 7, 15), None),
    ("INV-1014", "Nguyen rental", "Turnover cleans (3 stays)", 480.00, date(2026, 7, 24), date(2026, 8, 23), None),
]
TRIPS = [  # (date, where/why, miles)
    (date(2026, 1, 6),  "Harbor Dental — weekly clean", 18),
    (date(2026, 1, 9),  "Reyes family — deep clean", 24),
    (date(2026, 1, 20), "Birchwood Apartments — move-outs", 31),
    (date(2026, 2, 3),  "Harbor Dental — weekly clean", 18),
    (date(2026, 2, 10), "Nguyen rental — turnovers", 42),
    (date(2026, 2, 27), "Birchwood Apartments — move-out", 31),
    (date(2026, 3, 9),  "Nguyen rental — turnovers", 42),
    (date(2026, 3, 20), "Ferris & Co. — deep clean", 26),
]
MILEAGE_RATE = 0.70   # example only — user enters the current IRS standard rate themselves
SE_SHARE = 0.9235     # share of profit subject to SE tax (user-editable assumption cell)
SE_RATE = 0.153       # combined Social Security + Medicare (user-editable assumption cell)
INC_RATE = 0.12       # rough income-tax rate placeholder (user-editable assumption cell)

# layout constants (verify_bookkeeping.py re-derives formulas from these same numbers)
L_FIRST, L_LAST = 10, 209          # ledger entry rows (200)
M_FIRST, M_LAST, M_YEAR = 10, 21, 22   # ledger monthly-subtotal block
I_FIRST, I_LAST = 11, 110          # invoice rows (100)
T_FIRST, T_LAST = 9, 208           # mileage rows (200)
D_FIRST, D_LAST, D_YEAR = 7, 18, 19    # dashboard month rows

wb = Workbook()

def banner(ws, last_col, text, size=18, height=30):
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    c = ws.cell(row=2, column=2, value=text)
    c.font = Font(name="Arial", size=size, bold=True, color="FFFFFF")
    c.fill = HEAD_FILL; c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = height
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3

# ---------- Tab 1: START HERE ----------
ws = wb.active; ws.title = "START HERE"
for c, w in zip("BCDE", (36, 30, 30, 30)): ws.column_dimensions[c].width = w
banner(ws, 5, "SMALL BUSINESS BOOKKEEPING & INVOICE TRACKER", size=20, height=34)
rows = [
    ("", ""),
    ("Welcome! One workbook runs the money side of a solo service business: what comes in, what goes", ""),
    ("out, who still owes you, and roughly what to set aside for taxes.", ""),
    ("", ""),
    ("HOW TO USE (3 steps)", "SEC"),
    ("1. 'Income & Expense Ledger' tab: log every payment in and every business expense out — date, type,", ""),
    ("    category, amount in the yellow cells. Monthly subtotals and the P&L fill themselves.", ""),
    ("2. 'Invoice Tracker' tab: one row per invoice with sent / due / paid dates. The sheet stamps each one", ""),
    ("    PAID, SENT, or OVERDUE, shows days overdue, and totals what's still owed to you.", ""),
    ("3. Once a month, glance at 'P&L Dashboard'. Once a quarter, open 'Quarterly Tax Estimate' and move", ""),
    ("    the suggested set-aside into savings. Log business drives in 'Mileage Log' as you go.", ""),
    ("", ""),
    ("THE WORKED EXAMPLE — WALK THROUGH IT", "SEC"),
    ("The numbers already in the sheets are one realistic example business (a part-time solo cleaner):", ""),
    ("• Ledger: 26 entries, Jan–Apr. Money in $4,650.00, money out $725.46, profit so far $3,924.54.", ""),
    ("   The month table on the right shows Jan–Mar profitable (green) and April slightly negative (red).", ""),
    ("• Invoice Tracker: 14 invoices. INV-1012 and INV-1013 are OVERDUE (that's $610.00 to chase);", ""),
    ("   INV-1014 is sent and not yet due. The totals at the top count all of it for you.", ""),
    ("• Mileage Log: 232 business miles at the example $0.70 rate = $162.40 deduction.", ""),
    ("• Quarterly Tax Estimate: the example profit produces a suggested set-aside of roughly 20%.", ""),
    ("Replace the yellow cells with your own business — every total recalculates instantly.", ""),
    ("", ""),
    ("COLOR LEGEND", "SEC"),
    ("Yellow cells = yours to edit. Everything else calculates automatically — please don't type over it.", ""),
    ("Green / yellow / red = fine / watch / needs attention (profit months, invoice status, overdue days).", ""),
    ("Enter dates as real dates (e.g. 1/6/2026) — statuses and monthly totals read them.", ""),
    ("Type and Category on the Ledger are dropdowns — pick, don't type (custom categories are fine).", ""),
    ("", ""),
    ("A FEW HONEST NOTES", "SEC"),
    ("• The Quarterly Tax Estimate is a simplified ESTIMATE to help you set money aside. It is not tax,", ""),
    ("   legal, or accounting advice, and it doesn't know your full situation. Confirm with a tax professional.", ""),
    ("• Mileage: enter the current standard mileage rate yourself in the yellow rate cell (check irs.gov —", ""),
    ("   it changes). If you use the standard-rate deduction, don't also log fuel/repairs for the same", ""),
    ("   vehicle as expenses — it's one or the other. Ask your tax pro.", ""),
    ("", ""),
    ("GOOGLE SHEETS USERS", "SEC"),
    ("Upload this file to Google Drive, then open it — it converts automatically: formulas, dropdowns, and", ""),
    ("traffic-light colors work as-is. The P&L chart converts to a Sheets chart (colors can shift slightly;", ""),
    ("data identical). Excel 2010+ shows everything exactly as built. No macros anywhere.", ""),
    ("", ""),
    ("Questions or a formula acting up? Message us on Etsy any time — we answer fast and we'll fix it.", ""),
]
r = 4
for text, kind in rows:
    cell = ws.cell(row=r, column=2, value=text)
    if kind == "SEC":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell.font = H2; cell.fill = SEC_FILL
        ws.row_dimensions[r].height = 20
    else:
        cell.font = B
    r += 1

# ---------- Tab 2: Income & Expense Ledger ----------
ws = wb.create_sheet("Income & Expense Ledger")
for c, w in zip("BCDEFGHIJKL", (12, 13, 22, 46, 13, 6, 2, 10, 13, 13, 13)):
    ws.column_dimensions[c].width = w
banner(ws, 12, "INCOME & EXPENSE LEDGER")
stats = [  # labels merged B:E so they can never clip; values in F ($0.00 for zero)
    ("Money in (year so far)", f'=SUMIF($C${L_FIRST}:$C${L_LAST},"Income",$F${L_FIRST}:$F${L_LAST})'),
    ("Money out (year so far)", f'=SUMIF($C${L_FIRST}:$C${L_LAST},"Expense",$F${L_FIRST}:$F${L_LAST})'),
    ("Net profit (year so far)", "=F4-F5"),
]
for i, (lab, f) in enumerate(stats):
    r = 4 + i
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(row=r, column=2, value=lab).font = BB
    c = ws.cell(row=r, column=6, value=f)
    c.font = BB; c.fill = ALT_FILL; c.border = BORDER; c.number_format = MONEY0
ws["B8"] = ("Log one row per transaction. Type and Category are dropdowns — pick instead of typing "
            "(custom categories are fine too). 'Mo.' fills itself from the date.")
ws["B8"].font = SMALL
hdrs = ["Date", "Type", "Category", "Description / client", "Amount", "Mo."]
for i, h in enumerate(hdrs):
    c = ws.cell(row=9, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for i, h in enumerate(["Month", "Money in", "Money out", "Net"]):
    c = ws.cell(row=9, column=9 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for r in range(L_FIRST, L_LAST + 1):
    data = LEDGER[r - L_FIRST] if r - L_FIRST < len(LEDGER) else (None, None, None, None, None)
    d, t, cat, desc, amt = data
    dc = ws.cell(row=r, column=2, value=d); dc.number_format = DATEFMT
    ws.cell(row=r, column=3, value=t)
    ws.cell(row=r, column=4, value=cat)
    ws.cell(row=r, column=5, value=desc)
    ac = ws.cell(row=r, column=6, value=amt); ac.number_format = MONEY
    mo = ws.cell(row=r, column=7, value=f'=IF($B{r}="","",MONTH($B{r}))')
    for col in range(2, 7):
        cc = ws.cell(row=r, column=col); cc.fill = INPUT_FILL; cc.border = BORDER; cc.font = B
    mo.border = BORDER; mo.font = B
MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
for i, m in enumerate(MONTH_NAMES):
    r = M_FIRST + i
    ws.cell(row=r, column=9, value=m).font = B
    jc = ws.cell(row=r, column=10, value=f'=SUMIFS($F${L_FIRST}:$F${L_LAST},$G${L_FIRST}:$G${L_LAST},{i + 1},$C${L_FIRST}:$C${L_LAST},"Income")')
    kc = ws.cell(row=r, column=11, value=f'=SUMIFS($F${L_FIRST}:$F${L_LAST},$G${L_FIRST}:$G${L_LAST},{i + 1},$C${L_FIRST}:$C${L_LAST},"Expense")')
    lc = ws.cell(row=r, column=12, value=f'=IF(AND(J{r}=0,K{r}=0),"",J{r}-K{r})')
    for cc in (jc, kc, lc): cc.number_format = MONEY; cc.font = B
    for col in range(9, 13): ws.cell(row=r, column=col).border = BORDER
ws.cell(row=M_YEAR, column=9, value="Year").font = BB
ws.cell(row=M_YEAR, column=10, value=f"=SUM(J{M_FIRST}:J{M_LAST})")
ws.cell(row=M_YEAR, column=11, value=f"=SUM(K{M_FIRST}:K{M_LAST})")
ws.cell(row=M_YEAR, column=12, value=f"=J{M_YEAR}-K{M_YEAR}")
for col in range(9, 13):
    c = ws.cell(row=M_YEAR, column=col); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
    if col > 9: c.number_format = MONEY
ws.conditional_formatting.add(f"L{M_FIRST}:L{M_LAST}", CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL, font=GREEN_FONT))
ws.conditional_formatting.add(f"L{M_FIRST}:L{M_LAST}", CellIsRule(operator="equal", formula=["0"], fill=YEL_FILL, font=YEL_FONT))
ws.conditional_formatting.add(f"L{M_FIRST}:L{M_LAST}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT))
ws.cell(row=L_LAST + 2, column=2, value=("The totals read rows 10–209 (200 entries). If you ever fill them all, "
                                         "insert new rows ABOVE the last one and every formula stretches automatically.")).font = SMALL
# dropdowns: Type is strict (the totals count those exact words); Category offers
# a Schedule-C-shaped list but accepts anything you type
dv_type = DataValidation(type="list", formula1='"Income,Expense"', allow_blank=True,
                         showErrorMessage=True, errorTitle="Income or Expense",
                         error="Pick Income or Expense — every total reads these words.")
ws.add_data_validation(dv_type); dv_type.add(f"C{L_FIRST}:C{L_LAST}")
CATEGORIES = ("Client payment,Product sales,Other income,Supplies,Equipment,Insurance,"
              "Advertising,Software / subscriptions,Phone,Internet,Rent,Utilities,"
              "Contract labor,Professional services,Bank / card fees,Travel,Other expense")
dv_cat = DataValidation(type="list", formula1=f'"{CATEGORIES}"', allow_blank=True,
                        showErrorMessage=False)
ws.add_data_validation(dv_cat); dv_cat.add(f"D{L_FIRST}:D{L_LAST}")
ws.freeze_panes = "B10"

# ---------- Tab 3: Invoice Tracker ----------
ws = wb.create_sheet("Invoice Tracker")
for c, w in zip("BCDEFGHIJ", (11, 22, 32, 12, 12, 12, 12, 11, 13)):
    ws.column_dimensions[c].width = w
banner(ws, 10, "INVOICE TRACKER")
money_stats = [
    ("Invoiced (all invoices)", f"=SUM($E${I_FIRST}:$E${I_LAST})"),
    ("Collected (paid)", f'=SUMIF($I${I_FIRST}:$I${I_LAST},"PAID",$E${I_FIRST}:$E${I_LAST})'),
    ("Outstanding (unpaid)", "=E4-E5"),
    ("Overdue right now", f'=SUMIF($I${I_FIRST}:$I${I_LAST},"OVERDUE",$E${I_FIRST}:$E${I_LAST})'),
]
for i, (lab, f) in enumerate(money_stats):
    ws.cell(row=4 + i, column=2, value=lab).font = BB
    c = ws.cell(row=4 + i, column=5, value=f)
    c.font = BB; c.fill = ALT_FILL; c.border = BORDER; c.number_format = MONEY0
count_stats = [
    ("# paid", f'=COUNTIF($I${I_FIRST}:$I${I_LAST},"PAID")'),
    ("# awaiting", f'=COUNTIF($I${I_FIRST}:$I${I_LAST},"SENT")'),
    ("# overdue", f'=COUNTIF($I${I_FIRST}:$I${I_LAST},"OVERDUE")'),
]
for i, (lab, f) in enumerate(count_stats):
    ws.cell(row=4 + i, column=7, value=lab).font = BB
    c = ws.cell(row=4 + i, column=8, value=f)
    c.font = BB; c.fill = ALT_FILL; c.border = BORDER
ws["B9"] = ("One row per invoice — yellow cells only, dates as real dates. Status and Days overdue calculate themselves: "
            "PAID once a paid date exists, OVERDUE when unpaid past the due date.")
ws["B9"].font = SMALL
hdrs = ["Invoice #", "Client", "What for", "Amount", "Date sent", "Due date", "Date paid", "Status", "Days overdue"]
for i, h in enumerate(hdrs):
    c = ws.cell(row=10, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for r in range(I_FIRST, I_LAST + 1):
    data = INVOICES[r - I_FIRST] if r - I_FIRST < len(INVOICES) else (None, None, None, None, None, None, None)
    num, client, what, amt, sent, due, paid = data
    ws.cell(row=r, column=2, value=num)
    ws.cell(row=r, column=3, value=client)
    ws.cell(row=r, column=4, value=what)
    ac = ws.cell(row=r, column=5, value=amt); ac.number_format = MONEY
    for col, v in ((6, sent), (7, due), (8, paid)):
        dc = ws.cell(row=r, column=col, value=v); dc.number_format = DATEFMT
    st = ws.cell(row=r, column=9, value=f'=IF($B{r}="","",IF($H{r}<>"","PAID",IF($G{r}="","SENT",IF(TODAY()>$G{r},"OVERDUE","SENT"))))')
    dv = ws.cell(row=r, column=10, value=f'=IF($I{r}="OVERDUE",TODAY()-$G{r},"")')
    dv.number_format = "0"
    for col in range(2, 9):
        cc = ws.cell(row=r, column=col); cc.fill = INPUT_FILL; cc.border = BORDER; cc.font = B
    st.border = BORDER; st.font = BB; dv.border = BORDER; dv.font = B
    st.alignment = Alignment(horizontal="center")
rng = f"I{I_FIRST}:I{I_LAST}"
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"PAID"'], fill=GREEN_FILL, font=GREEN_FONT))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"SENT"'], fill=YEL_FILL, font=YEL_FONT))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"OVERDUE"'], fill=RED_FILL, font=RED_FONT))
rng = f"J{I_FIRST}:J{I_LAST}"
ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["1", "30"], fill=YEL_FILL, font=YEL_FONT))
ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["30"], fill=RED_FILL, font=RED_FONT))
ws.cell(row=I_LAST + 2, column=2, value=("Tip: send the invoice, then set Due date (e.g. sent + 14 days). "
                                         "The totals above always show exactly who owes you money.")).font = SMALL
ws.freeze_panes = "B11"

# ---------- Tab 4: Quarterly Tax Estimate ----------
ws = wb.create_sheet("Quarterly Tax Estimate")
for c, w in zip("BCDE", (46, 15, 18, 24)): ws.column_dimensions[c].width = w
banner(ws, 5, "QUARTERLY TAX ESTIMATE")
ws["B4"] = "ESTIMATE ONLY — NOT TAX ADVICE. A simplified math helper for setting money aside."
ws["B4"].font = WARN
ws["B5"] = "Rates change and your situation is unique — confirm with a tax professional or irs.gov before filing anything."
ws["B5"].font = SMALL

def sec(ws, row, text, last_col=5):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=2, value=text); c.font = H2; c.fill = SEC_FILL
    ws.row_dimensions[row].height = 20

sec(ws, 7, "YOUR YEAR SO FAR (auto — from the other tabs)")
LG = "'Income & Expense Ledger'"
auto_rows = [
    (8, "Money in (Ledger)", f"={LG}!F4", MONEY),
    (9, "Business expenses (Ledger)", f"={LG}!F5", MONEY),
    (10, "Mileage deduction (Mileage Log)", "='Mileage Log'!C6", MONEY),
    (11, "Estimated profit (in − expenses − mileage)", "=C8-C9-C10", MONEY),
]
for r, lab, f, fmt in auto_rows:
    ws.cell(row=r, column=2, value=lab).font = BB
    c = ws.cell(row=r, column=3, value=f); c.font = BB; c.fill = ALT_FILL; c.border = BORDER; c.number_format = fmt
sec(ws, 13, "ASSUMPTIONS (yellow — adjust to your situation)")
assum = [
    (14, "Share of profit subject to SE tax", SE_SHARE, PCT2),
    (15, "Self-employment tax rate (Social Security + Medicare)", SE_RATE, PCT2),
    (16, "Estimated income tax rate (your rough bracket)", INC_RATE, PCT2),
]
for r, lab, v, fmt in assum:
    ws.cell(row=r, column=2, value=lab).font = BB
    c = ws.cell(row=r, column=3, value=v); c.fill = INPUT_FILL; c.border = BORDER; c.number_format = fmt
ws["B17"] = ("Defaults are the widely used federal self-employment figures plus a rough income-tax rate. Rates change — "
             "check irs.gov. State tax and the Social Security wage cap are NOT included.")
ws["B17"].font = SMALL
sec(ws, 19, "THE ESTIMATE (year so far)")
est_rows = [
    (20, "Profit subject to SE tax", "=MAX(0,$C$11)*$C$14", MONEY, False),
    (21, "Estimated self-employment tax", "=$C$20*$C$15", MONEY, False),
    (22, "Estimated income tax (after half-SE deduction)", "=MAX(0,$C$11-$C$21/2)*$C$16", MONEY, False),
    (23, "Total estimated tax (year so far)", "=$C$21+$C$22", MONEY, True),
    (24, "Suggested set-aside (share of every dollar in)", "=IFERROR($C$23/$C$8,0)", PCT1, True),
    (25, "Effective rate on profit (used by the table below)", "=IF($C$11<=0,0,$C$23/$C$11)", PCT1, False),
]
for r, lab, f, fmt, big in est_rows:
    ws.cell(row=r, column=2, value=lab).font = BB
    c = ws.cell(row=r, column=3, value=f); c.font = BB; c.border = BORDER; c.number_format = fmt
    if big: c.fill = ALT_FILL
sec(ws, 27, "QUARTER BY QUARTER")
for i, h in enumerate(["Quarter", "Profit", "Set aside"]):
    c = ws.cell(row=28, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
ML = "'Mileage Log'"
Q_FIRST = 29
qlabels = ["Q1 (Jan–Mar)", "Q2 (Apr–Jun)", "Q3 (Jul–Sep)", "Q4 (Oct–Dec)"]
for q in range(4):
    r = Q_FIRST + q; lo, hi = 3 * q + 1, 3 * q + 3
    ws.cell(row=r, column=2, value=qlabels[q]).font = B
    pf = (f'=SUMIFS({LG}!$F${L_FIRST}:$F${L_LAST},{LG}!$G${L_FIRST}:$G${L_LAST},">={lo}",{LG}!$G${L_FIRST}:$G${L_LAST},"<={hi}",{LG}!$C${L_FIRST}:$C${L_LAST},"Income")'
          f'-SUMIFS({LG}!$F${L_FIRST}:$F${L_LAST},{LG}!$G${L_FIRST}:$G${L_LAST},">={lo}",{LG}!$G${L_FIRST}:$G${L_LAST},"<={hi}",{LG}!$C${L_FIRST}:$C${L_LAST},"Expense")'
          f'-SUMIFS({ML}!$E${T_FIRST}:$E${T_LAST},{ML}!$F${T_FIRST}:$F${T_LAST},">={lo}",{ML}!$F${T_FIRST}:$F${T_LAST},"<={hi}")')
    pc = ws.cell(row=r, column=3, value=pf); pc.number_format = MONEY; pc.font = B
    sc = ws.cell(row=r, column=4, value=f"=MAX(0,$C{r})*$C$25"); sc.number_format = MONEY; sc.font = B
    for col in range(2, 5): ws.cell(row=r, column=col).border = BORDER
QT = Q_FIRST + 4
ws.cell(row=QT, column=2, value="Year").font = BB
ws.cell(row=QT, column=3, value=f"=SUM(C{Q_FIRST}:C{QT - 1})").number_format = MONEY
ws.cell(row=QT, column=4, value=f"=SUM(D{Q_FIRST}:D{QT - 1})").number_format = MONEY
for col in range(2, 5):
    c = ws.cell(row=QT, column=col); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
ws.cell(row=QT + 2, column=2, value=("Actual estimated-tax due dates and safe-harbor rules live on IRS Form 1040-ES — "
                                     "this table just splits the year's estimate by when you earned it.")).font = SMALL
ws.cell(row=QT + 3, column=2, value=("Set-asides only count profitable quarters, so their total can run a little above the year "
                                     "estimate when a quarter lost money — better to over-save than under.")).font = SMALL
ws.cell(row=QT + 5, column=2, value="Again: this whole tab is an estimate for setting money aside — not tax advice.").font = WARN

# ---------- Tab 5: P&L Dashboard ----------
ws = wb.create_sheet("P&L Dashboard")
for c, w in zip("BCDEF", (10, 14, 14, 14, 11)): ws.column_dimensions[c].width = w
banner(ws, 6, "P&L DASHBOARD")
ws["B4"] = "Nothing to type here — every number flows from the Ledger. Green months made money, red months lost it."
ws["B4"].font = SMALL
for i, h in enumerate(["Month", "Money in", "Money out", "Profit", "Margin"]):
    c = ws.cell(row=6, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for i, m in enumerate(MONTH_NAMES):
    r = D_FIRST + i; src = M_FIRST + i
    ws.cell(row=r, column=2, value=m).font = B
    cc = ws.cell(row=r, column=3, value=f"={LG}!J{src}"); cc.number_format = MONEY
    dc = ws.cell(row=r, column=4, value=f"={LG}!K{src}"); dc.number_format = MONEY
    ec = ws.cell(row=r, column=5, value=f'=IF(AND(C{r}=0,D{r}=0),"",C{r}-D{r})'); ec.number_format = MONEY
    fc = ws.cell(row=r, column=6, value=f'=IF(OR(C{r}=0,E{r}=""),"",E{r}/C{r})'); fc.number_format = PCT1
    for col in range(2, 7):
        c2 = ws.cell(row=r, column=col); c2.border = BORDER
        if col > 2: c2.font = B
ws.cell(row=D_YEAR, column=2, value="Year").font = BB
ws.cell(row=D_YEAR, column=3, value=f"=SUM(C{D_FIRST}:C{D_LAST})").number_format = MONEY
ws.cell(row=D_YEAR, column=4, value=f"=SUM(D{D_FIRST}:D{D_LAST})").number_format = MONEY
ws.cell(row=D_YEAR, column=5, value=f"=C{D_YEAR}-D{D_YEAR}").number_format = MONEY
ws.cell(row=D_YEAR, column=6, value=f'=IF(C{D_YEAR}=0,"",E{D_YEAR}/C{D_YEAR})').number_format = PCT1
for col in range(2, 7):
    c = ws.cell(row=D_YEAR, column=col); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
ws.conditional_formatting.add(f"E{D_FIRST}:E{D_YEAR}", CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL, font=GREEN_FONT))
ws.conditional_formatting.add(f"E{D_FIRST}:E{D_YEAR}", CellIsRule(operator="equal", formula=["0"], fill=YEL_FILL, font=YEL_FONT))
ws.conditional_formatting.add(f"E{D_FIRST}:E{D_YEAR}", CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT))
# bottom stats: labels merged B:D so they can never clip; values in E ($0.00 for zero)
ws.merge_cells(start_row=D_YEAR + 2, start_column=2, end_row=D_YEAR + 2, end_column=4)
ws.cell(row=D_YEAR + 2, column=2, value="Best month (profit)").font = BB
c = ws.cell(row=D_YEAR + 2, column=5, value=f"=MAX(E{D_FIRST}:E{D_LAST})")
c.font = BB; c.fill = ALT_FILL; c.border = BORDER; c.number_format = MONEY0
ws.merge_cells(start_row=D_YEAR + 3, start_column=2, end_row=D_YEAR + 3, end_column=4)
ws.cell(row=D_YEAR + 3, column=2, value="Average month (months with income)").font = BB
c = ws.cell(row=D_YEAR + 3, column=5, value=f'=IFERROR(AVERAGEIF(C{D_FIRST}:C{D_LAST},">0",E{D_FIRST}:E{D_LAST}),0)')
c.font = BB; c.fill = ALT_FILL; c.border = BORDER; c.number_format = MONEY0
ws.cell(row=D_YEAR + 5, column=2, value=("Profit here = money in − money out from the Ledger (a cash view). The mileage "
                                         "deduction is a tax-time figure — see the Quarterly Tax Estimate tab.")).font = SMALL
# the dashboard chart: monthly money in/out columns + profit line, one shared $ axis
bar = BarChart(); bar.type = "col"
bar.title = "Money in, money out, profit — by month"
bar.add_data(Reference(ws, min_col=3, max_col=4, min_row=6, max_row=D_LAST), titles_from_data=True)
bar.set_categories(Reference(ws, min_col=2, min_row=D_FIRST, max_row=D_LAST))
bar.series[0].graphicalProperties.solidFill = NAVY
bar.series[1].graphicalProperties.solidFill = GOLD
line = LineChart()
line.add_data(Reference(ws, min_col=5, min_row=6, max_row=D_LAST), titles_from_data=True)
line.series[0].graphicalProperties.line.solidFill = TEAL
line.series[0].graphicalProperties.line.width = 28000
line.series[0].smooth = False
line.y_axis.axId = bar.y_axis.axId  # same axis ids -> single shared $ axis
line.x_axis.axId = bar.x_axis.axId
bar += line
bar.y_axis.numFmt = "$#,##0"
bar.width = 22; bar.height = 11
ws.add_chart(bar, "H6")

# ---------- Tab 6: Mileage Log ----------
ws = wb.create_sheet("Mileage Log")
for c, w in zip("BCDEF", (12, 36, 10, 13, 6)): ws.column_dimensions[c].width = w
banner(ws, 6, "MILEAGE LOG")
ws["B4"] = "Rate per mile:"; ws["B4"].font = BB
c = ws["C4"]; c.value = MILEAGE_RATE; c.fill = INPUT_FILL; c.border = BORDER; c.number_format = RATEFMT
ws["D4"] = "← enter the CURRENT standard mileage rate (check irs.gov — it changes most years)"
ws["D4"].font = SMALL
ws["B5"] = "Total business miles"; ws["B5"].font = BB
c = ws["C5"]; c.value = f"=SUM($D${T_FIRST}:$D${T_LAST})"; c.font = BB; c.fill = ALT_FILL; c.border = BORDER; c.number_format = "#,##0"
ws["B6"] = "Total mileage deduction"; ws["B6"].font = BB
c = ws["C6"]; c.value = f"=SUM($E${T_FIRST}:$E${T_LAST})"; c.font = BB; c.fill = ALT_FILL; c.border = BORDER; c.number_format = MONEY
for i, h in enumerate(["Date", "Where / why (client, job)", "Miles", "Deduction", "Mo."]):
    c = ws.cell(row=8, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
for r in range(T_FIRST, T_LAST + 1):
    data = TRIPS[r - T_FIRST] if r - T_FIRST < len(TRIPS) else (None, None, None)
    d, why, miles = data
    dc = ws.cell(row=r, column=2, value=d); dc.number_format = DATEFMT
    ws.cell(row=r, column=3, value=why)
    mc = ws.cell(row=r, column=4, value=miles); mc.number_format = "#,##0"
    ded = ws.cell(row=r, column=5, value=f'=IF($D{r}="","",ROUND($D{r}*$C$4,2))'); ded.number_format = MONEY
    mo = ws.cell(row=r, column=6, value=f'=IF($B{r}="","",MONTH($B{r}))')
    for col in range(2, 5):
        cc = ws.cell(row=r, column=col); cc.fill = INPUT_FILL; cc.border = BORDER; cc.font = B
    ded.border = BORDER; ded.font = B; mo.border = BORDER; mo.font = B
ws.cell(row=T_LAST + 2, column=2, value=("Standard mileage OR actual vehicle costs — not both for the same vehicle. If you take "
                                         "this standard-rate deduction, don't also log fuel/repairs on the Ledger. Ask your tax pro.")).font = SMALL
ws.freeze_panes = "B9"

XLSX = os.path.join(OUT, "Small-Business-Bookkeeping-Tracker.xlsx")
wb.save(XLSX)
print("saved", XLSX)

# ================================================================== store card SVG (1200x1200)
SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 1200 1200" font-family="Helvetica,Arial,sans-serif">
<rect width="1200" height="1200" fill="#f8f6f0"/>
<rect width="1200" height="200" fill="#1f3a5f"/>
<text x="600" y="92" text-anchor="middle" font-size="56" font-weight="bold" fill="#ffffff">SMALL BUSINESS BOOKKEEPING</text>
<text x="600" y="160" text-anchor="middle" font-size="56" font-weight="bold" fill="#f2c14e">&amp; INVOICE TRACKER</text>
<text x="600" y="262" text-anchor="middle" font-size="33" fill="#282c34">Every dollar in, every dollar out, every invoice chased —</text>
<text x="600" y="306" text-anchor="middle" font-size="33" font-weight="bold" fill="#282c34">one spreadsheet runs the money side of your business.</text>
<rect x="170" y="350" width="860" height="416" rx="16" fill="#ffffff" stroke="#d9d4c8" stroke-width="2"/>
<rect x="170" y="350" width="860" height="64" fill="#2e7d6b" rx="16"/>
<rect x="170" y="384" width="860" height="30" fill="#2e7d6b"/>
<text x="205" y="393" font-size="30" font-weight="bold" fill="#ffffff">THIS YEAR AT A GLANCE</text>
<text x="215" y="472" font-size="32" fill="#282c34">Money in</text><text x="985" y="472" text-anchor="end" font-size="32" font-weight="bold" fill="#1f3a5f">$4,650</text>
<text x="215" y="548" font-size="32" fill="#282c34">Overdue invoices</text><text x="985" y="548" text-anchor="end" font-size="32" font-weight="bold" fill="#1f3a5f">2 — $610</text>
<text x="215" y="624" font-size="32" fill="#282c34">Profit so far</text><text x="985" y="624" text-anchor="end" font-size="32" font-weight="bold" fill="#1f3a5f">$3,925</text>
<rect x="178" y="662" width="844" height="68" fill="#f0f3ee"/>
<text x="215" y="708" font-size="32" font-weight="bold" fill="#1f3a5f">SUGGESTED TAX SET-ASIDE*</text><text x="985" y="708" text-anchor="end" font-size="32" font-weight="bold" fill="#2e7d6b">20.5%</text>
<text x="600" y="800" text-anchor="middle" font-size="24" font-style="italic" fill="#5a5f69">*Example figures. The tax line is an estimate only — not tax advice.</text>
<text x="180" y="880" font-size="34" font-weight="bold" fill="#2e7d6b">&#10003;</text><text x="230" y="880" font-size="32" fill="#282c34">Income &amp; expense ledger with monthly subtotals</text>
<text x="180" y="944" font-size="34" font-weight="bold" fill="#2e7d6b">&#10003;</text><text x="230" y="944" font-size="32" fill="#282c34">Invoice tracker flags OVERDUE + totals who owes you</text>
<text x="180" y="1008" font-size="34" font-weight="bold" fill="#2e7d6b">&#10003;</text><text x="230" y="1008" font-size="32" fill="#282c34">Quarterly tax set-aside estimate + mileage log</text>
<text x="180" y="1072" font-size="34" font-weight="bold" fill="#2e7d6b">&#10003;</text><text x="230" y="1072" font-size="32" fill="#282c34">P&amp;L dashboard — profit by month, year totals</text>
<rect y="1110" width="1200" height="90" fill="#f2c14e"/>
<text x="600" y="1167" text-anchor="middle" font-size="32" font-weight="bold" fill="#282c34">Google Sheets + Excel&#160;&#160;&#8226;&#160;&#160;No macros&#160;&#160;&#8226;&#160;&#160;Instant download</text>
</svg>"""
with open(os.path.join(OUT, "store-card.svg"), "w") as f:
    f.write(SVG)
print("saved", os.path.join(OUT, "store-card.svg"))
