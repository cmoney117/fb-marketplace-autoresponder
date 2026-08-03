#!/usr/bin/env python3
"""Build The 10-Minute Simple Budget spreadsheet product (deliberately minimal,
ADHD-friendly sibling listing).

Deterministic: xlsx + 2700x2025 cover PNG + 460x345 store SVG, then programmatic
verification of EVERY formula in the workbook (mini evaluator, no deps beyond
openpyxl/PIL). Exits non-zero if any formula disagrees with expected values.

2026-08-03 P1 review fixes baked in:
  - Bill Calendar "type an x" convention now backed by a real dropdown
    (data validation list) on the whole paid grid — no more silent COUNTIF misses
  - LEFT OVER hero cells render $0.00 (not the accounting dash) at exactly zero
  - mini bar chart (Planned vs Spent per line) beside the Monthly Budget table,
    driven by the live ranges so it updates as buyers type
No macros — formulas + native features only (Excel 2010+ / Google Sheets).
"""
import os, re, filecmp, shutil, zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageDraw, ImageFont

OUT = "/home/user/fb-marketplace-autoresponder/venture/07-automation/tracker/publish-queue/simple-budget-starter-v1"
STORE_IMG = "/home/user/fb-marketplace-autoresponder/store/site/img"
os.makedirs(OUT, exist_ok=True)

NAVY = "1F3A5F"; TEAL = "2E7D6B"; GOLD = "F2C14E"; LIGHT = "F4F6F8"; YELLOW = "FFF2CC"
H1 = Font(name="Arial", size=18, bold=True, color="FFFFFF")
H2 = Font(name="Arial", size=12, bold=True, color="FFFFFF")
B = Font(name="Arial", size=11)
BB = Font(name="Arial", size=11, bold=True)
BIG = Font(name="Arial", size=14, bold=True)
SMALL = Font(name="Arial", size=10, italic=True, color="666666")
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
SEC_FILL = PatternFill("solid", fgColor=TEAL)
ALT_FILL = PatternFill("solid", fgColor=LIGHT)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00;($#,##0.00);"-"'
MONEY0 = "$#,##0.00;($#,##0.00)"  # hero cells: 0 renders $0.00, never "-"
# traffic-light fills (classic Excel good/neutral/bad)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE"); GREEN_FONT = Font(name="Arial", color="006100")
YEL_FILL = PatternFill("solid", fgColor="FFEB9C"); YEL_FONT = Font(name="Arial", color="9C6500")
RED_FILL = PatternFill("solid", fgColor="FFC7CE"); RED_FONT = Font(name="Arial", color="9C0006")

# ------------------------------------------------------------------ source data
INCOME = [("Paychecks", 2600), ("Other money", 150)]
SPEND = [  # exactly 10 rows — that's the whole point (planned, spent so far)
    ("Rent / housing", 1100, 1100),
    ("Groceries", 400, 372.40),
    ("Utilities", 140, 151.62),
    ("Phone + internet", 95, 95),
    ("Car / transport", 220, 198.75),
    ("Debt payment", 200, 200),
    ("Subscriptions", 45, 52.99),
    ("Fun money", 120, 84.30),
    ("Savings", 150, 150),
    ("Everything else", 100, 61.15),
]
BILLS = [  # (bill, due day, amount, paid-in-Jan?) — Jan is the worked example
    ("Rent", 1, 1100, "x"),
    ("Electric", 12, 120, "x"),
    ("Water & trash", 15, 55, "x"),
    ("Internet", 18, 60, "x"),
    ("Phone", 20, 45, "x"),
    ("Car payment", 5, 265, "x"),
    ("Car insurance", 8, 110, "x"),
    ("Streaming", 22, 27.99, ""),
    ("Gym", 3, 24, "x"),
    ("Credit card min", 25, 75, ""),
    ("Student loan", 28, 180, ""),
    ("Renters insurance", 10, 15, ""),
]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

wb = Workbook()

# ---------- Tab 1: START HERE ----------
ws = wb.active; ws.title = "START HERE"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDE", (36, 30, 30, 30)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:E2"); ws["B2"] = "THE 10-MINUTE SIMPLE BUDGET"
ws["B2"].font = Font(name="Arial", size=20, bold=True, color="FFFFFF")
ws["B2"].fill = HEAD_FILL; ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 34
rows = [
    ("", ""),
    ("This budget is small on purpose. Ten spending lines. One page. No 12-tab guilt trip.", ""),
    ("If big budget spreadsheets have never stuck for you, this one is built for your brain.", ""),
    ("", ""),
    ("HOW TO USE (3 steps, ~10 minutes once)", "SEC"),
    ("1. 'Monthly Budget' tab: enter your money in, then a planned amount for each of the 10 lines.", ""),
    ("2. When you spend, update the Spent column. Colors do the judging: green = fine,", ""),
    ("    yellow = right at the line, red = over. The LEFT OVER box is the only number that matters.", ""),
    ("3. 'Bill Calendar' tab: pick the x from the dropdown when you pay a bill — the box turns green. That's it.", ""),
    ("", ""),
    ("COLOR LEGEND", "SEC"),
    ("Yellow cells = yours to edit.  Everything else calculates automatically — please don't type over it.", ""),
    ("The numbers already in the sheet are a realistic example. Replace them with your own.", ""),
    ("", ""),
    ("GOOGLE SHEETS USERS", "SEC"),
    ("Upload this file to Google Drive, then open it — it converts automatically, colors and all.", ""),
    ("The dropdowns and the little Planned-vs-Spent chart convert too; chart colors may shift slightly. Numbers are identical.", ""),
    ("", ""),
    ("Questions or a formula acting up? Message us on Etsy any time — we answer fast and we'll fix it.", ""),
]
r = 4
for text, kind in rows:
    cell = ws.cell(row=r, column=2, value=text)
    if kind == "SEC":
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        cell.font = H2; cell.fill = SEC_FILL
        ws.row_dimensions[r].height = 20
    else:
        cell.font = B
    r += 1

# ---------- Tab 2: Monthly Budget ----------
ws = wb.create_sheet("Monthly Budget")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for c, w in zip("BCDE", (26, 14, 14, 14)): ws.column_dimensions[c].width = w
ws.merge_cells("B2:E2"); ws["B2"] = "MONTHLY BUDGET"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
ws["B4"] = "Fill the yellow cells. That's the whole system."
ws["B4"].font = SMALL

ws.merge_cells("B6:E6"); ws["B6"] = "MONEY IN"; ws["B6"].font = H2; ws["B6"].fill = SEC_FILL
IN_FIRST = 7
for i, (name, amt) in enumerate(INCOME):
    r = IN_FIRST + i
    ws.cell(row=r, column=2, value=name).font = B
    c = ws.cell(row=r, column=3, value=amt); c.fill = INPUT_FILL; c.number_format = MONEY
    ws.cell(row=r, column=2).border = BORDER; c.border = BORDER
IN_TOT = IN_FIRST + len(INCOME)
ws.cell(row=IN_TOT, column=2, value="Total in").font = BB
tc = ws.cell(row=IN_TOT, column=3, value=f"=SUM(C{IN_FIRST}:C{IN_TOT-1})")
tc.font = BB; tc.number_format = MONEY
for col in (2, 3): ws.cell(row=IN_TOT, column=col).fill = ALT_FILL; ws.cell(row=IN_TOT, column=col).border = BORDER

SEC_R = IN_TOT + 2
ws.merge_cells(start_row=SEC_R, start_column=2, end_row=SEC_R, end_column=5)
ws.cell(row=SEC_R, column=2, value="MONEY OUT — 10 lines, no more").font = H2
ws.cell(row=SEC_R, column=2).fill = SEC_FILL
hdr_r = SEC_R + 1
for i, h in enumerate(["Spending", "Planned", "Spent", "Difference"]):
    c = ws.cell(row=hdr_r, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
S_FIRST = hdr_r + 1
for i, (name, planned, spent) in enumerate(SPEND):
    r = S_FIRST + i
    ws.cell(row=r, column=2, value=name).font = B
    p = ws.cell(row=r, column=3, value=planned); p.fill = INPUT_FILL
    s = ws.cell(row=r, column=4, value=spent); s.fill = INPUT_FILL
    df = ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
    for col in range(2, 6): ws.cell(row=r, column=col).border = BORDER
    for cc in (p, s, df): cc.number_format = MONEY
S_LAST = S_FIRST + len(SPEND) - 1
S_TOT = S_LAST + 1
ws.cell(row=S_TOT, column=2, value="Total out").font = BB
ws.cell(row=S_TOT, column=3, value=f"=SUM(C{S_FIRST}:C{S_LAST})").font = BB
ws.cell(row=S_TOT, column=4, value=f"=SUM(D{S_FIRST}:D{S_LAST})").font = BB
ws.cell(row=S_TOT, column=5, value=f"=C{S_TOT}-D{S_TOT}").font = BB
for col in range(2, 6):
    c = ws.cell(row=S_TOT, column=col); c.fill = ALT_FILL; c.border = BORDER
    if col > 2: c.number_format = MONEY

LO_HDR = S_TOT + 2
ws.merge_cells(start_row=LO_HDR, start_column=2, end_row=LO_HDR, end_column=5)
ws.cell(row=LO_HDR, column=2, value="LEFT OVER").font = H2
ws.cell(row=LO_HDR, column=2).fill = HEAD_FILL
LO_PLAN, LO_NOW = LO_HDR + 1, LO_HDR + 2
ws.cell(row=LO_PLAN, column=2, value="If you stick to the plan").font = BB
pc = ws.cell(row=LO_PLAN, column=3, value=f"=C{IN_TOT}-C{S_TOT}"); pc.font = BB; pc.number_format = MONEY; pc.border = BORDER
ws.cell(row=LO_NOW, column=2, value="Right now (in − spent)").font = BB
nc = ws.cell(row=LO_NOW, column=3, value=f"=C{IN_TOT}-D{S_TOT}"); nc.font = BIG; nc.number_format = MONEY; nc.border = BORDER
ws.cell(row=LO_NOW + 2, column=2, value="Green = money left. Red = you spent more than came in. No shame either way — just information.").font = SMALL
ws.freeze_panes = "A6"
# hero cells show $0.00 at exactly zero (never the accounting dash)
pc.number_format = MONEY0
nc.number_format = MONEY0

# traffic-light conditional formatting
diff_range = f"E{S_FIRST}:E{S_LAST}"
lo_range = f"C{LO_PLAN}:C{LO_NOW}"
for rng in (diff_range, lo_range):
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL, font=GREEN_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=["0"], fill=YEL_FILL, font=YEL_FONT))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL, font=RED_FONT))

# mini chart: Planned vs Spent per line, beside the table (updates as buyers type)
chart = BarChart()
chart.type = "bar"  # horizontal so the 10 line names read naturally
chart.style = 10
chart.title = "Planned vs Spent — your 10 lines"
chart.add_data(Reference(ws, min_col=3, max_col=4, min_row=hdr_r, max_row=S_LAST), titles_from_data=True)
chart.set_categories(Reference(ws, min_col=2, min_row=S_FIRST, max_row=S_LAST))
chart.series[0].graphicalProperties = GraphicalProperties(solidFill=NAVY)  # Planned
chart.series[1].graphicalProperties = GraphicalProperties(solidFill=GOLD)  # Spent
chart.y_axis.numFmt = "$#,##0"
chart.y_axis.majorGridlines = None
chart.y_axis.scaling.min = 0  # bars grow from $0
chart.legend.position = "b"
chart.width = 16
chart.height = 12
ws.add_chart(chart, "G6")

# ---------- Tab 3: Bill Calendar ----------
ws = wb.create_sheet("Bill Calendar")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 20; ws.column_dimensions["C"].width = 9; ws.column_dimensions["D"].width = 11
for i in range(len(MONTHS)):
    ws.column_dimensions[chr(ord("E") + i)].width = 6
ws.merge_cells("B2:P2"); ws["B2"] = "BILL CALENDAR"
ws["B2"].font = H1; ws["B2"].fill = HEAD_FILL
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 30
ws["B4"] = "Pick the x from the dropdown when a bill is paid — the box turns green. January is a worked example; clear it and start with your month."
ws["B4"].font = SMALL
hdrs = ["Bill", "Due day", "Amount"] + MONTHS
for i, h in enumerate(hdrs):
    c = ws.cell(row=6, column=2 + i, value=h); c.font = BB; c.fill = ALT_FILL; c.border = BORDER
B_FIRST, B_LAST = 7, 26  # 12 example bills + 8 blank rows
for i in range(B_FIRST, B_LAST + 1):
    if i - B_FIRST < len(BILLS):
        name, day, amt, jan = BILLS[i - B_FIRST]
    else:
        name, day, amt, jan = None, None, None, ""
    ws.cell(row=i, column=2, value=name).font = B
    ws.cell(row=i, column=3, value=day).font = B
    ac = ws.cell(row=i, column=4, value=amt); ac.number_format = MONEY
    ws.cell(row=i, column=5, value=jan or None).font = B
    for col in range(2, 17):
        c = ws.cell(row=i, column=col)
        c.fill = INPUT_FILL; c.border = BORDER
        if col >= 5: c.alignment = Alignment(horizontal="center")
B_TOT = B_LAST + 1
ws.cell(row=B_TOT, column=2, value="Monthly bills total").font = BB
mt = ws.cell(row=B_TOT, column=4, value=f"=SUM(D{B_FIRST}:D{B_LAST})"); mt.font = BB; mt.number_format = MONEY
PAID_R = B_TOT + 1
ws.cell(row=PAID_R, column=2, value="Paid this month").font = BB
for i in range(len(MONTHS)):
    col_letter = chr(ord("E") + i)
    c = ws.cell(row=PAID_R, column=5 + i, value=f'=COUNTIF({col_letter}{B_FIRST}:{col_letter}{B_LAST},"x")')
    c.font = BB; c.alignment = Alignment(horizontal="center")
for row in (B_TOT, PAID_R):
    for col in range(2, 17): ws.cell(row=row, column=col).fill = ALT_FILL; ws.cell(row=row, column=col).border = BORDER
ws.cell(row=PAID_R + 2, column=2, value="That little green grid filling up is the whole reward system. It works.").font = SMALL
# green when paid
ws.conditional_formatting.add(f"E{B_FIRST}:P{B_LAST}", CellIsRule(operator="equal", formula=['"x"'], fill=GREEN_FILL, font=GREEN_FONT))
# real dropdown on the whole paid grid — the "type an x" convention can no
# longer silently miss the COUNTIF (review P1-3)
dv_x = DataValidation(type="list", formula1='"x"', allow_blank=True, showDropDown=False)
dv_x.error = "Pick x from the dropdown (or leave the box blank)."
dv_x.errorTitle = "Paid marker"
dv_x.prompt = "Pick x when this bill is paid."
dv_x.promptTitle = "Paid?"
ws.add_data_validation(dv_x)
dv_x.add(f"E{B_FIRST}:P{B_LAST}")
ws.freeze_panes = "E7"

# print setup: landscape, fit to one page wide — Ctrl+P gives whole tabs, not confetti
from openpyxl.worksheet.properties import PageSetupProperties
for sheet in wb.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

XLSX = os.path.join(OUT, "10-Minute-Simple-Budget.xlsx")
wb.save(XLSX)
print("saved", XLSX)

# ================================================================== verification
def _cells(ws, ref):
    if ":" in ref:
        return [c for row in ws[ref] for c in row]
    return [ws[ref]]

def _val(ws, cell):
    """Numeric value of a cell, recursively evaluating formulas."""
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        return evaluate(ws, v)
    return v if isinstance(v, (int, float)) else 0

def _match(v, crit):
    return isinstance(v, str) and v.strip().lower() == crit.strip().lower()

def evaluate(ws, formula):
    """Evaluate the closed set of formulas this product uses: SUM, COUNTIF +
    cell arithmetic (+ -)."""
    expr = formula.lstrip("=").replace("$", "")
    def fn(m):
        name, args = m.group(1).upper(), [a.strip() for a in m.group(2).split(",")]
        if name == "SUM":
            return repr(sum(_val(ws, c) for c in _cells(ws, args[0])))
        if name == "COUNTIF":
            return repr(sum(1 for c in _cells(ws, args[0]) if _match(c.value, args[1].strip('"'))))
        raise ValueError("unsupported function " + name)
    expr = re.sub(r"([A-Z]+)\(([^()]*)\)", fn, expr)
    expr = re.sub(r"\b([A-Z]{1,2}[0-9]{1,4})\b", lambda m: repr(_val(ws, ws[m.group(1)])), expr)
    return eval(expr)

wb2 = load_workbook(XLSX)
income_total = sum(a for _, a in INCOME)
plan_total = round(sum(p for _, p, _ in SPEND), 2)
spent_total = round(sum(s for _, _, s in SPEND), 2)
bills_total = round(sum(b[2] for b in BILLS), 2)
jan_paid = sum(1 for b in BILLS if b[3] == "x")

expected = {}
expected[("Monthly Budget", f"C{IN_TOT}")] = income_total
for i, (name, planned, spent) in enumerate(SPEND):
    expected[("Monthly Budget", f"E{S_FIRST+i}")] = round(planned - spent, 2)
expected[("Monthly Budget", f"C{S_TOT}")] = plan_total
expected[("Monthly Budget", f"D{S_TOT}")] = spent_total
expected[("Monthly Budget", f"E{S_TOT}")] = round(plan_total - spent_total, 2)
expected[("Monthly Budget", f"C{LO_PLAN}")] = round(income_total - plan_total, 2)
expected[("Monthly Budget", f"C{LO_NOW}")] = round(income_total - spent_total, 2)
expected[("Bill Calendar", f"D{B_TOT}")] = bills_total
for i in range(len(MONTHS)):
    expected[("Bill Calendar", f"{chr(ord('E') + i)}{PAID_R}")] = jan_paid if i == 0 else 0

checked = 0
for sheet in wb2.sheetnames:
    ws2 = wb2[sheet]
    for row in ws2.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                key = (sheet, cell.coordinate)
                assert key in expected, f"formula {key} has no expected value — add it"
                got = evaluate(ws2, cell.value)
                want = expected.pop(key)
                assert abs(got - want) < 1e-9, f"{key}: got {got}, expected {want}"
                checked += 1
assert not expected, f"expected values never matched a formula cell: {sorted(expected)}"

# hard pins — the worked example must read exactly like the listing says
assert income_total == 2750
assert plan_total == 2570 and spent_total == 2466.21
assert round(income_total - plan_total, 2) == 180.00
assert round(income_total - spent_total, 2) == 283.79
assert bills_total == 2076.99 and jan_paid == 8
# traffic-light sanity: example rows hit all three colors
diffs = [round(p - s, 2) for _, p, s in SPEND]
assert any(d > 0 for d in diffs) and any(d == 0 for d in diffs) and any(d < 0 for d in diffs)

# --- P1 review fix assertions -------------------------------------------------
# hero cells: LEFT OVER boxes must render $0.00 at zero, never the dash
mb = wb2["Monthly Budget"]
for co in (f"C{LO_PLAN}", f"C{LO_NOW}"):
    assert mb[co].number_format == MONEY0, f"hero cell {co} lost the $0-format"
# dropdown on the whole Bill Calendar paid grid
bc = wb2["Bill Calendar"]
dvs = bc.data_validations.dataValidation
assert len(dvs) == 1 and dvs[0].type == "list" and dvs[0].formula1 == '"x"', "paid-grid dropdown missing"
assert str(dvs[0].sqref) == f"E{B_FIRST}:P{B_LAST}", f"dropdown range wrong: {dvs[0].sqref}"
# native chart present and driven by the live Monthly Budget ranges
with zipfile.ZipFile(XLSX) as z:
    chart_parts = [n for n in z.namelist() if n.startswith("xl/charts/chart")]
    assert len(chart_parts) == 1, f"expected 1 chart part, found {chart_parts}"
    chart_xml = z.read(chart_parts[0]).decode()
assert f"'Monthly Budget'!$C${S_FIRST}:$C${S_LAST}" in chart_xml, "chart Planned series not wired to live range"
assert f"'Monthly Budget'!$D${S_FIRST}:$D${S_LAST}" in chart_xml, "chart Spent series not wired to live range"
assert f"'Monthly Budget'!$B${S_FIRST}:$B${S_LAST}" in chart_xml, "chart categories not wired to line names"
assert f"'Monthly Budget'!C{hdr_r}" in chart_xml and f"'Monthly Budget'!D{hdr_r}" in chart_xml, "series titles not from header row"
print("VERIFIED: $0-format hero cells, paid-grid dropdown, live-range chart")
print(f"VERIFIED: all {checked} formulas match expected values")
print(f"  money in $2,750; planned out $2,570 -> left over $180.00 planned; spent $2,466.21 -> LEFT OVER $283.79")
print(f"  traffic lights exercised: {sum(1 for d in diffs if d > 0)} green / {sum(1 for d in diffs if d == 0)} yellow / {sum(1 for d in diffs if d < 0)} red rows")
print(f"  bill calendar: 12 bills, ${bills_total:,.2f}/month, Jan example 8 of 12 paid; Feb-Dec 0")

# ================================================================== cover PNG (2700x2025)
W, H = 2700, 2025
NAVY_RGB = (31, 58, 95); TEAL_RGB = (46, 125, 107); GOLD_RGB = (242, 193, 78)
CREAM = (248, 246, 240); DARK = (40, 44, 52); LINE = (217, 212, 200); ALT = (240, 243, 238)
FB = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
FR = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
def F(p, s): return ImageFont.truetype(p, s)

img = Image.new("RGB", (W, H), CREAM); d = ImageDraw.Draw(img)
d.rectangle([0, 0, W, 300], fill=NAVY_RGB)
size = 132
while d.textlength("THE 10-MINUTE SIMPLE BUDGET", font=F(FB, size)) > W - 260: size -= 4
d.text((W / 2, 150), "THE 10-MINUTE SIMPLE BUDGET", font=F(FB, size), fill=(255, 255, 255), anchor="mm")
d.text((W / 2, 460), "One page. Ten lines. Traffic-light colors", font=F(FR, 66), fill=DARK, anchor="mm")
d.text((W / 2, 570), "tell you exactly where you stand.", font=F(FB, 66), fill=DARK, anchor="mm")
cx0, cy0, cx1, cy1 = 450, 700, 2250, 1600
d.rounded_rectangle([cx0, cy0, cx1, cy1], radius=24, fill=(255, 255, 255), outline=LINE, width=3)
d.rounded_rectangle([cx0, cy0, cx1, cy0 + 110], radius=24, fill=NAVY_RGB)
d.rectangle([cx0, cy0 + 60, cx1, cy0 + 110], fill=NAVY_RGB)
d.text((cx0 + 60, cy0 + 55), "MONTHLY BUDGET", font=F(FB, 56), fill=(255, 255, 255), anchor="lm")
rows = [("Money in", "$2,750", DARK), ("Money out (10 lines)", "$2,466", DARK), ("Bills paid this month", "8 of 12", DARK), ("LEFT OVER", "$284", TEAL_RGB)]
y = cy0 + 220
for i, (lab, val, vcol) in enumerate(rows):
    last = i == len(rows) - 1
    if last:
        d.rectangle([cx0 + 6, y - 60, cx1 - 6, y + 60], fill=(230, 244, 233))
    d.text((cx0 + 70, y), lab, font=F(FB if last else FR, 60), fill=NAVY_RGB if last else DARK, anchor="lm")
    d.text((cx1 - 70, y), val, font=F(FB, 60), fill=TEAL_RGB if last else NAVY_RGB, anchor="rm")
    y += 165
d.rectangle([0, H - 190, W, H], fill=GOLD_RGB)
d.text((W / 2, H - 95), "ADHD-friendly by design  •  Google Sheets + Excel  •  Instant Download", font=F(FB, 60), fill=DARK, anchor="mm")
img.save(os.path.join(OUT, "01-cover.png"), optimize=True)
print("saved", os.path.join(OUT, "01-cover.png"))

# ================================================================== store SVG (460x345)
SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 460 345" font-family="Helvetica,Arial,sans-serif">
<rect width="460" height="345" fill="#f8f6f0"/>
<rect width="460" height="62" fill="#2e7d6b"/>
<text x="230" y="40" text-anchor="middle" font-size="22" font-weight="bold" fill="#ffffff">THE 10-MINUTE SIMPLE BUDGET</text>
<text x="230" y="92" text-anchor="middle" font-size="15" fill="#282c34">One page. Ten lines. Traffic-light colors</text>
<text x="230" y="112" text-anchor="middle" font-size="15" font-weight="bold" fill="#282c34">tell you exactly where you stand.</text>
<rect x="100" y="132" width="290" height="164" rx="8" fill="#ffffff" stroke="#d9d4c8"/>
<rect x="100" y="132" width="290" height="30" fill="#1f3a5f" rx="8"/>
<rect x="100" y="152" width="290" height="12" fill="#1f3a5f"/>
<text x="115" y="152" font-size="13" font-weight="bold" fill="#ffffff">MONTHLY BUDGET</text>
<text x="120" y="208" font-size="13" fill="#282c34">Money in</text><text x="378" text-anchor="end" y="208" font-size="13" font-weight="bold" fill="#1f3a5f">$2,750</text><text x="120" y="232" font-size="13" fill="#282c34">Money out (10 lines)</text><text x="378" text-anchor="end" y="232" font-size="13" font-weight="bold" fill="#1f3a5f">$2,466</text><text x="120" y="256" font-size="13" fill="#282c34">Bills paid</text><text x="378" text-anchor="end" y="256" font-size="13" font-weight="bold" fill="#1f3a5f">8 of 12</text><text x="120" y="280" font-size="13" fill="#282c34">LEFT OVER</text><text x="378" text-anchor="end" y="280" font-size="13" font-weight="bold" fill="#2e7d6b">$284</text>
<rect y="305" width="460" height="40" fill="#f2c14e"/>
<text x="230" y="330" text-anchor="middle" font-size="13" font-weight="bold" fill="#282c34">One page  •  Traffic lights  •  10 minutes flat</text>
</svg>"""
with open(os.path.join(OUT, "store-card.svg"), "w") as f:
    f.write(SVG)
# store/ is rebuilt by the main session — only copy when the card actually changed
store_card = os.path.join(STORE_IMG, "simple-budget.svg")
if os.path.isdir(STORE_IMG) and not (os.path.exists(store_card) and filecmp.cmp(os.path.join(OUT, "store-card.svg"), store_card, shallow=False)):
    shutil.copy(os.path.join(OUT, "store-card.svg"), store_card)
    print("saved store-card.svg (+ copied to store/site/img/simple-budget.svg)")
else:
    print("saved store-card.svg (store copy already up to date — untouched)")
