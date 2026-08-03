#!/usr/bin/env python3
"""Verify 10-Minute-Simple-Budget.xlsx with a REAL spreadsheet engine.

build_product.py already re-verifies every formula with its own mini evaluator
at build time; this script is the independent ground truth: LibreOffice
headless recalc + the P1 review fix checks (paid-grid dropdown, $0-format
hero cells, live-range mini chart).
"""
import os, subprocess, sys, tempfile, zipfile
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "10-Minute-Simple-Budget.xlsx")
MONEY0 = "$#,##0.00;($#,##0.00)"
failures = []

def check(cond, msg):
    print(("  ok  " if cond else "  FAIL") + " " + msg)
    if not cond:
        failures.append(msg)

wb = load_workbook(XLSX)
mb, bc = wb["Monthly Budget"], wb["Bill Calendar"]

print("[1] P1 fix structure")
check(mb["C26"].number_format == MONEY0 and mb["C27"].number_format == MONEY0,
      "LEFT OVER hero cells render $0.00 at zero (never '-')")
dvs = {str(dv.sqref): dv for dv in bc.data_validations.dataValidation}
check("E7:P26" in dvs and dvs["E7:P26"].formula1 == '"x"', "paid-grid x-dropdown covers E7:P26")
with zipfile.ZipFile(XLSX) as z:
    charts = [n for n in z.namelist() if n.startswith("xl/charts/chart")]
    chart_xml = z.read(charts[0]).decode() if charts else ""
check(len(charts) == 1, "exactly one chart part in the workbook")
check("'Monthly Budget'!$C$13:$C$22" in chart_xml and "'Monthly Budget'!$D$13:$D$22" in chart_xml,
      "Planned + Spent series wired to the live 10 lines")
check("'Monthly Budget'!$B$13:$B$22" in chart_xml, "categories are the line names")
check('val="1F3A5F"' in chart_xml and 'val="F2C14E"' in chart_xml, "house palette (navy/gold) applied")
sh_texts = " | ".join(str(c.value) for row in wb["START HERE"].iter_rows() for c in row if c.value)
check("dropdown" in sh_texts, "START HERE mentions the dropdown")

print("[2] formula recalc (LibreOffice ground truth)")
with tempfile.TemporaryDirectory() as td:
    subprocess.run(["soffice", "--headless", "--convert-to", "xlsx", "--outdir", td, XLSX],
                   check=True, capture_output=True, timeout=180)
    calc = load_workbook(os.path.join(td, os.path.basename(XLSX)), data_only=True)
mbc, bcc = calc["Monthly Budget"], calc["Bill Calendar"]
def near(a, b, eps=0.01):
    return a is not None and isinstance(a, (int, float)) and abs(a - b) < eps
check(near(mbc["C9"].value, 2750), "money in $2,750")
check(near(mbc["C23"].value, 2570) and near(mbc["D23"].value, 2466.21), "money out $2,570 planned / $2,466.21 spent")
check(near(mbc["C26"].value, 180) and near(mbc["C27"].value, 283.79), "LEFT OVER $180 planned / $283.79 now")
check(near(bcc["D27"].value, 2076.99), "bills total $2,076.99")
check(bcc["E28"].value == 8 and bcc["F28"].value == 0, "Jan example 8 paid, Feb 0")

if failures:
    print(f"\nFAILED: {len(failures)} check(s)")
    sys.exit(1)
print("\nALL CHECKS PASSED")
